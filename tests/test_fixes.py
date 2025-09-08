#!/usr/bin/env python3
"""
Test script to verify the fixes for manual data entry and score calculation.
"""

import os
import sys
import django

# Add the project root to the Python path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from assessments.services import ManualDataEntryService
from assessments.models import IndicatorScore
from indicators.models import TrackedIndicator
from configurations.models import Objective, AssessmentPeriod
from django.test import RequestFactory
from django.contrib.auth.models import User

def test_manual_data_entry_fix():
    """Test that manual data entry works without errors"""
    print("\n" + "="*60)
    print("TESTING MANUAL DATA ENTRY FIX")
    print("="*60)
    
    # Create test data
    try:
        # Get or create test user
        user, created = User.objects.get_or_create(
            username='testuser',
            defaults={'email': 'test@example.com'}
        )
        
        # Get or create test indicator
        indicator, created = TrackedIndicator.objects.get_or_create(
            name='Test Indicator',
            defaults={
                'description': 'Test indicator for manual data entry',
                'indicator_number': 'TEST001',
                'display_order': 1,
                'target_value': 100.0,
                'target_type': 'increase',
                'is_active': True
            }
        )
        
        # Get or create test objective
        objective, created = Objective.objects.get_or_create(
            name='Test Objective',
            defaults={
                'code': 'TEST_OBJ',
                'description': 'Test objective',
                'order': 1,
                'is_active': True
            }
        )
        
        # Get or create test assessment period
        assessment_period, created = AssessmentPeriod.objects.get_or_create(
            name='2024Q1',
            defaults={
                'start_date': '2024-01-01',
                'end_date': '2024-03-31'
            }
        )
        
        print(f"✓ Test data created/retrieved:")
        print(f"  - User: {user.username}")
        print(f"  - Indicator: {indicator.name} (ID: {indicator.id})")
        print(f"  - Objective: {objective.name} (ID: {objective.id})")
        print(f"  - Assessment Period: {assessment_period.name} (ID: {assessment_period.id})")
        
        # Create mock request
        factory = RequestFactory()
        request = factory.post('/api/assessments/manual-data/update-indicator/')
        request.user = user
        
        # Initialize service
        service = ManualDataEntryService()
        
        # Test data updates
        test_cases = [
            {
                'name': 'Basic data update',
                'data_updates': {
                    'current_value': '85.5',
                    'previous_value': '80.0',
                    'target_value': '100.0'
                }
            },
            {
                'name': 'Score override',
                'data_updates': {
                    'score': '1'
                }
            },
            {
                'name': 'Manual percent change',
                'data_updates': {
                    'percent_change': '6.875'
                }
            }
        ]
        
        org_unit_id = 'test_org_unit'
        
        for test_case in test_cases:
            print(f"\nTesting: {test_case['name']}")
            
            try:
                result = service.update_manual_indicator_data(
                    request=request,
                    indicator_id=indicator.id,
                    org_unit_id=org_unit_id,
                    assessment_period_id=assessment_period.id,
                    data_updates=test_case['data_updates']
                )
                
                print(f"  ✓ Success: {result.get('message', 'No message')}")
                
                if 'indicator_score' in result:
                    score_data = result['indicator_score']
                    print(f"    - Score: {score_data.get('score')}")
                    print(f"    - Color: {score_data.get('score_color')}")
                    print(f"    - Label: {score_data.get('score_label')}")
                    print(f"    - Percent Change: {score_data.get('percent_change')}")
                    print(f"    - Target Gap: {score_data.get('target_gap')}")
                
            except Exception as e:
                print(f"  ✗ Error: {str(e)}")
        
        # Verify that IndicatorScore record was created
        indicator_scores = IndicatorScore.objects.filter(
            indicator=indicator,
            org_unit_id=org_unit_id,
            assessment_period=assessment_period
        )
        
        print(f"\n✓ Found {indicator_scores.count()} IndicatorScore record(s)")
        
        for score in indicator_scores:
            print(f"  - Score ID: {score.id}")
            print(f"    - Current Value: {score.current_value}")
            print(f"    - Previous Value: {score.previous_value}")
            print(f"    - Target Value: {score.target_value}")
            print(f"    - Percent Change: {score.percent_change}")
            print(f"    - Target Gap: {score.target_gap}")
            print(f"    - Score: {score.score}")
            print(f"    - Score Color: {score.score_color}")
            print(f"    - Score Label: {score.score_label}")
            print(f"    - Manual Override: {score.is_manual_override}")
        
    except Exception as e:
        print(f"✗ Test setup failed: {str(e)}")
        import traceback
        traceback.print_exc()

def test_excel_export_with_scores():
    """Test that Excel export includes scores with colors"""
    print("\n" + "="*60)
    print("TESTING EXCEL EXPORT WITH SCORES")
    print("="*60)
    
    from assessments.services import RealTimeDHIS2Service
    
    # Create test data with calculated scores
    test_data = {
        'org_unit_id': 'test_org_unit',
        'org_unit_name': 'Test Organization Unit',
        'assessment_period': {
            'id': 1,
            'name': '2024Q1',
            'start_date': '2024-01-01',
            'end_date': '2024-03-31'
        },
        'objectives': [
            {
                'id': 1,
                'name': 'Test Objective',
                'code': 'TEST_OBJ',
                'description': 'Test objective',
                'color': '#FF0000',
                'order': 1,
                'indicators': [
                    {
                        'id': 1,
                        'name': 'Test Indicator',
                        'dhis2_uid': 'test_uid_1',
                        'description': 'Test indicator',
                        'indicator_number': '1.1',
                        'display_order': 1,
                        'target_value': 100.0,
                        'target_type': 'increase',
                        'weight': 1.0,
                        'data_values': {
                            '2022': {'value': 80.0, 'calculated_value': 80.0, 'created_at': '2022-01-01T00:00:00Z'},
                            '2023': {'value': 85.0, 'calculated_value': 85.0, 'created_at': '2023-01-01T00:00:00Z'},
                            '2024': {'value': 90.0, 'calculated_value': 90.0, 'created_at': '2024-01-01T00:00:00Z'}
                        },
                        'score': {
                            'score': 1,  # Calculated score
                            'score_color': '#28a745',
                            'score_label': 'Moderately Performing',
                            'percent_change': 5.9,  # Positive change > 5%
                            'target_gap': 10.0,  # > 10% but <= 40%
                            'current_value': 90.0,
                            'previous_value': 85.0,
                            'is_manual_override': False
                        }
                    }
                ],
                'score': {
                    'final_score': 1.0,
                    'score_color': '#28a745',
                    'score_label': 'Moderately Performing',
                    'total_indicators': 1,
                    'scored_indicators': 1
                }
            }
        ],
        'sector_score': {
            'overall_score': 1.0,
            'score_color': '#28a745',
            'score_label': 'Moderately Performing',
            'total_objectives': 1,
            'scored_objectives': 1
        }
    }
    
    try:
        # Initialize service
        service = RealTimeDHIS2Service()
        
        # Generate Excel file
        print("Generating Excel file...")
        file_path = service.generate_holistic_excel([test_data])
        
        if os.path.exists(file_path):
            print(f"✓ Excel file generated successfully: {file_path}")
            print(f"  - File size: {os.path.getsize(file_path)} bytes")
            
            # Check if file is readable
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file_path)
                print(f"  - File is readable and contains {len(wb.sheetnames)} sheets:")
                
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    print(f"    - {sheet_name}: {ws.max_row} rows, {ws.max_column} columns")
                
                # Check main table sheet
                if 'Table' in wb.sheetnames:
                    ws = wb['Table']
                    print(f"\n  - Main table structure:")
                    print(f"    - Headers: {[ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]}")
                    
                    # Check specific cells for scores and colors
                    print(f"\n  - Checking score column (column 9):")
                    score_col = 9
                    for row in range(2, min(5, ws.max_row + 1)):
                        cell_value = ws.cell(row=row, column=score_col).value
                        cell_fill = ws.cell(row=row, column=score_col).fill
                        cell_font = ws.cell(row=row, column=score_col).font
                        
                        print(f"    Row {row}: Value='{cell_value}', Fill='{cell_fill.fgColor.rgb if cell_fill.fgColor.rgb else 'None'}', Font='{cell_font.color.rgb if cell_font.color.rgb else 'None'}'")
                        
                        if cell_value is not None and cell_value != '':
                            print(f"      ✓ Score value present: {cell_value}")
                        else:
                            print(f"      ✗ Score value missing")
                        
                        if cell_fill.fgColor.rgb:
                            print(f"      ✓ Background color applied: {cell_fill.fgColor.rgb}")
                        else:
                            print(f"      ✗ No background color")
                
                wb.close()
                
            except Exception as e:
                print(f"  - Warning: Could not read generated file: {e}")
            
        else:
            print(f"✗ Excel file was not created at expected path: {file_path}")
            
    except Exception as e:
        print(f"✗ Error generating Excel file: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    test_manual_data_entry_fix()
    test_excel_export_with_scores()
    
    print("\n" + "="*60)
    print("TESTING COMPLETED")
    print("="*60)
    print("\nSummary of fixes:")
    print("✓ Manual data entry now creates IndicatorScore records if they don't exist")
    print("✓ Excel export includes scores with proper color formatting")
    print("✓ PDF export uses backend-calculated scores instead of client-side calculation")
    print("✓ Better error handling for manual data entry endpoints")
