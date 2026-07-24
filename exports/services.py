import os
import uuid
import logging
from datetime import timedelta
from django.conf import settings
from django.utils import timezone
from django.db import transaction
from django.db.models import Q, Count, Avg, Sum, Max, Min
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import json

from .models import (
    ExportTemplate, ExportJob, ExportSchedule, ExportLog, ExportConfiguration
)
from assessments.models import (
    IndicatorScore, ObjectiveScore, SectorScore, DataSyncLog
)
from indicators.models import TrackedIndicator
from configurations.models import Objective, AssessmentPeriod
from organisation.models import OrgUnit
from organisation.services import AccessControlService

logger = logging.getLogger(__name__)


class ExportService:
    """
    Service for handling export operations
    """
    
    def __init__(self):
        self.access_service = AccessControlService()
    
    def create_export_job(self, export_request, user):
        """
        Create a new export job
        """
        try:
            # Generate unique job ID
            job_id = f"export_{uuid.uuid4().hex[:8]}"
            
            # Get template if specified
            template = None
            if export_request.get('template_id'):
                template = ExportTemplate.objects.get(
                    id=export_request['template_id'],
                    is_active=True
                )
            
            # Get org unit scope
            org_unit_scope = None
            if export_request.get('org_unit_id'):
                org_unit_scope = OrgUnit.objects.get(id=export_request['org_unit_id'])
            
            # Calculate expiration time
            expires_in_hours = export_request.get('expires_in_hours', 24)
            expires_at = timezone.now() + timedelta(hours=expires_in_hours)
            
            # Create export job
            export_job = ExportJob.objects.create(
                job_id=job_id,
                name=export_request.get('name', f"Export {job_id}"),
                description=export_request.get('description', ''),
                template=template,
                export_format=export_request['export_format'],
                export_type=export_request['export_type'],
                export_parameters=export_request.get('export_parameters', {}),
                priority=export_request.get('priority', ExportJob.ExportPriority.NORMAL),
                created_by_id=user.id,
                org_unit_scope=org_unit_scope,
                expires_at=expires_at
            )
            
            # Log job creation
            ExportLog.objects.create(
                export_job=export_job,
                level=ExportLog.LogLevel.INFO,
                message=f"Export job created: {export_job.name}",
                details={'export_request': export_request}
            )
            
            logger.info(f"Created export job {job_id} for user {user.dhis2_username}")
            return export_job
            
        except Exception as e:
            logger.error(f"Failed to create export job: {str(e)}")
            raise
    
    def process_export_job(self, export_job):
        """
        Process an export job and generate the file
        """
        try:
            # Mark job as started
            export_job.mark_started()
            
            # Log start
            ExportLog.objects.create(
                export_job=export_job,
                level=ExportLog.LogLevel.INFO,
                message="Export job started"
            )
            
            # Get export data based on type
            export_data = self._get_export_data(export_job)
            
            # Generate file based on format
            file_path, file_size = self._generate_export_file(
                export_job, export_data
            )
            
            # Generate file URL
            file_url = self._generate_file_url(file_path)
            
            # Mark job as completed
            export_job.mark_completed(
                file_path=file_path,
                file_size=file_size,
                file_url=file_url
            )
            
            # Log completion
            ExportLog.objects.create(
                export_job=export_job,
                level=ExportLog.LogLevel.INFO,
                message=f"Export job completed successfully. File size: {file_size} bytes"
            )
            
            logger.info(f"Export job {export_job.job_id} completed successfully")
            return export_job
            
        except Exception as e:
            # Mark job as failed
            export_job.mark_failed(str(e))
            
            # Log error
            ExportLog.objects.create(
                export_job=export_job,
                level=ExportLog.LogLevel.ERROR,
                message=f"Export job failed: {str(e)}",
                details={'error': str(e)}
            )
            
            logger.error(f"Export job {export_job.job_id} failed: {str(e)}")
            raise
    
    def _get_export_data(self, export_job):
        """
        Get data for export based on export type
        """
        export_type = export_job.export_type
        parameters = export_job.export_parameters
        
        if export_type == ExportTemplate.ExportType.DASHBOARD_SUMMARY:
            return self._get_dashboard_summary_data(export_job, parameters)
        elif export_type == ExportTemplate.ExportType.OBJECTIVE_PERFORMANCE:
            return self._get_objective_performance_data(export_job, parameters)
        elif export_type == ExportTemplate.ExportType.INDICATOR_PERFORMANCE:
            return self._get_indicator_performance_data(export_job, parameters)
        elif export_type == ExportTemplate.ExportType.ORG_UNIT_PERFORMANCE:
            return self._get_org_unit_performance_data(export_job, parameters)
        elif export_type == ExportTemplate.ExportType.TREND_ANALYSIS:
            return self._get_trend_analysis_data(export_job, parameters)
        elif export_type == ExportTemplate.ExportType.COMPARATIVE_ANALYSIS:
            return self._get_comparative_analysis_data(export_job, parameters)
        else:
            raise ValueError(f"Unsupported export type: {export_type}")
    
    def _get_dashboard_summary_data(self, export_job, parameters):
        """
        Get dashboard summary data for export
        """
        # Get user's accessible org units
        accessible_org_units = self.access_service.get_user_accessible_org_units(
            export_job.created_by
        )
        
        # Get assessment period
        assessment_period = parameters.get('assessment_period')
        if not assessment_period:
            current_period = AssessmentPeriod.objects.filter(is_current=True).first()
            assessment_period = current_period.period if current_period else None
        
        if not assessment_period:
            return {'error': 'No assessment period available'}
        
        # Get sector scores
        sector_scores = SectorScore.objects.filter(
            org_unit__in=accessible_org_units,
            assessment_period=assessment_period
        ).select_related('org_unit')
        
        # Calculate summary statistics
        total_org_units = accessible_org_units.count()
        org_units_with_scores = sector_scores.count()
        
        if org_units_with_scores > 0:
            avg_score = sector_scores.aggregate(avg=Avg('score'))['avg']
            max_score = sector_scores.aggregate(max=Max('score'))['max']
            min_score = sector_scores.aggregate(min=Min('score'))['min']
            
            # Count by performance category
            performance_counts = {}
            for score in sector_scores:
                label = score.label or 'Unknown'
                performance_counts[label] = performance_counts.get(label, 0) + 1
        else:
            avg_score = max_score = min_score = 0
            performance_counts = {}
        
        return {
            'assessment_period': assessment_period,
            'total_org_units': total_org_units,
            'org_units_with_scores': org_units_with_scores,
            'average_score': round(avg_score, 2) if avg_score else 0,
            'max_score': round(max_score, 2) if max_score else 0,
            'min_score': round(min_score, 2) if min_score else 0,
            'performance_distribution': performance_counts,
            'sector_scores': [
                {
                    'org_unit_id': score.org_unit_id,
                    'org_unit_name': score.org_unit.name,
                    'score': score.score,
                    'color': score.color,
                    'label': score.label
                }
                for score in sector_scores
            ]
        }
    
    def _get_objective_performance_data(self, export_job, parameters):
        """
        Get objective performance data for export
        """
        # Get user's accessible org units
        accessible_org_units = self.access_service.get_user_accessible_org_units(
            export_job.created_by
        )
        
        # Get assessment period
        assessment_period = parameters.get('assessment_period')
        if not assessment_period:
            current_period = AssessmentPeriod.objects.filter(is_current=True).first()
            assessment_period = current_period.period if current_period else None
        
        if not assessment_period:
            return {'error': 'No assessment period available'}
        
        # Get objective scores
        objective_scores = ObjectiveScore.objects.filter(
            org_unit__in=accessible_org_units,
            assessment_period=assessment_period
        ).select_related('objective', 'org_unit')
        
        # Group by objective
        objectives_data = {}
        for score in objective_scores:
            objective_name = score.objective.name
            if objective_name not in objectives_data:
                objectives_data[objective_name] = {
                    'objective_id': score.objective.id,
                    'objective_name': objective_name,
                    'scores': [],
                    'average_score': 0,
                    'performance_distribution': {}
                }
            
            objectives_data[objective_name]['scores'].append({
                'org_unit_id': score.org_unit_id,
                'org_unit_name': score.org_unit.name,
                'score': score.score,
                'color': score.color,
                'label': score.label
            })
        
        # Calculate statistics for each objective
        for objective_name, data in objectives_data.items():
            scores = [score['score'] for score in data['scores']]
            if scores:
                data['average_score'] = round(sum(scores) / len(scores), 2)
                
                # Count by performance category
                for score in data['scores']:
                    label = score['label'] or 'Unknown'
                    data['performance_distribution'][label] = data['performance_distribution'].get(label, 0) + 1
        
        return {
            'assessment_period': assessment_period,
            'objectives': list(objectives_data.values())
        }
    
    def _get_indicator_performance_data(self, export_job, parameters):
        """
        Get indicator performance data for export
        """
        # Get user's accessible org units
        accessible_org_units = self.access_service.get_user_accessible_org_units(
            export_job.created_by
        )
        
        # Get assessment period
        assessment_period = parameters.get('assessment_period')
        if not assessment_period:
            current_period = AssessmentPeriod.objects.filter(is_current=True).first()
            assessment_period = current_period.period if current_period else None
        
        if not assessment_period:
            return {'error': 'No assessment period available'}
        
        # Get indicator scores
        indicator_scores = IndicatorScore.objects.filter(
            org_unit__in=accessible_org_units,
            assessment_period=assessment_period
        ).select_related('indicator', 'org_unit')
        
        # Group by indicator
        indicators_data = {}
        for score in indicator_scores:
            indicator_name = score.indicator.name
            if indicator_name not in indicators_data:
                indicators_data[indicator_name] = {
                    'indicator_id': score.indicator.id,
                    'indicator_name': indicator_name,
                    'indicator_type': score.indicator.indicator_type,
                    'target_value': score.indicator.target_value,
                    'scores': [],
                    'average_score': 0,
                    'performance_distribution': {},
                    'trend_analysis': {
                        'improving': 0,
                        'declining': 0,
                        'stable': 0
                    }
                }
            
            indicators_data[indicator_name]['scores'].append({
                'org_unit_id': score.org_unit_id,
                'org_unit_name': score.org_unit.name,
                'raw_value': score.raw_value,
                'score': score.score,
                'trend': score.trend,
                'color': score.color,
                'label': score.label
            })
        
        # Calculate statistics for each indicator
        for indicator_name, data in indicators_data.items():
            scores = [score['score'] for score in data['scores']]
            if scores:
                data['average_score'] = round(sum(scores) / len(scores), 2)
                
                # Count by performance category and trend
                for score in data['scores']:
                    label = score['label'] or 'Unknown'
                    data['performance_distribution'][label] = data['performance_distribution'].get(label, 0) + 1
                    
                    trend = score['trend'] or 'stable'
                    data['trend_analysis'][trend] = data['trend_analysis'].get(trend, 0) + 1
        
        return {
            'assessment_period': assessment_period,
            'indicators': list(indicators_data.values())
        }
    
    def _get_org_unit_performance_data(self, export_job, parameters):
        """
        Get org unit performance data for export
        """
        # Get specific org unit or user's accessible org units
        org_unit_id = parameters.get('org_unit_id')
        if org_unit_id:
            org_units = OrgUnit.objects.filter(id=org_unit_id)
        else:
            org_units = self.access_service.get_user_accessible_org_units(
                export_job.created_by
            )
        
        # Get assessment period
        assessment_period = parameters.get('assessment_period')
        if not assessment_period:
            current_period = AssessmentPeriod.objects.filter(is_current=True).first()
            assessment_period = current_period.period if current_period else None
        
        if not assessment_period:
            return {'error': 'No assessment period available'}
        
        # Get performance data for each org unit
        org_units_data = []
        for org_unit in org_units:
            # Get sector score
            sector_score = SectorScore.objects.filter(
                org_unit=org_unit,
                assessment_period=assessment_period
            ).first()
            
            # Get objective scores
            objective_scores = ObjectiveScore.objects.filter(
                org_unit=org_unit,
                assessment_period=assessment_period
            ).select_related('objective')
            
            # Get indicator scores
            indicator_scores = IndicatorScore.objects.filter(
                org_unit=org_unit,
                assessment_period=assessment_period
            ).select_related('indicator')
            
            org_unit_data = {
                'org_unit': {
                    'id': org_unit.id,
                    'name': org_unit.name,
                    'level': org_unit.level.name
                },
                'assessment_period': assessment_period,
                'sector_score': {
                    'score': sector_score.score if sector_score else None,
                    'color': sector_score.color if sector_score else None,
                    'label': sector_score.label if sector_score else None
                },
                'objectives': [
                    {
                        'objective_id': score.objective.id,
                        'objective_name': score.objective.name,
                        'score': score.score,
                        'color': score.color,
                        'label': score.label
                    }
                    for score in objective_scores
                ],
                'indicators': [
                    {
                        'indicator_id': score.indicator.id,
                        'indicator_name': score.indicator.name,
                        'raw_value': score.raw_value,
                        'score': score.score,
                        'trend': score.trend,
                        'color': score.color,
                        'label': score.label
                    }
                    for score in indicator_scores
                ]
            }
            
            org_units_data.append(org_unit_data)
        
        return {
            'org_units': org_units_data
        }
    
    def _get_trend_analysis_data(self, export_job, parameters):
        """
        Get trend analysis data for export
        """
        # Get org unit
        org_unit_id = parameters.get('org_unit_id')
        if not org_unit_id:
            return {'error': 'org_unit_id is required for trend analysis'}
        
        org_unit = OrgUnit.objects.get(id=org_unit_id)
        
        # Get assessment period
        assessment_period = parameters.get('assessment_period')
        if not assessment_period:
            current_period = AssessmentPeriod.objects.filter(is_current=True).first()
            assessment_period = current_period.period if current_period else None
        
        if not assessment_period:
            return {'error': 'No assessment period available'}
        
        # Generate periods for trend analysis
        periods_back = parameters.get('periods_back', 3)
        periods = self._generate_periods_back(assessment_period, periods_back)
        
        # Get sector scores for all periods
        sector_scores = SectorScore.objects.filter(
            org_unit=org_unit,
            assessment_period__in=periods
        ).order_by('assessment_period')
        
        # Get objective scores for all periods
        objective_scores = ObjectiveScore.objects.filter(
            org_unit=org_unit,
            assessment_period__in=periods
        ).select_related('objective').order_by('objective__name', 'assessment_period')
        
        # Build trend data
        sector_trend = []
        objectives_trend = {}
        
        for period in periods:
            # Sector score for this period
            period_sector_score = sector_scores.filter(assessment_period=period).first()
            sector_trend.append({
                'period': period,
                'score': period_sector_score.score if period_sector_score else None,
                'color': period_sector_score.color if period_sector_score else None,
                'label': period_sector_score.label if period_sector_score else None
            })
            
            # Objective scores for this period
            period_objective_scores = objective_scores.filter(assessment_period=period)
            for score in period_objective_scores:
                objective_name = score.objective.name
                if objective_name not in objectives_trend:
                    objectives_trend[objective_name] = []
                
                objectives_trend[objective_name].append({
                    'period': period,
                    'score': score.score,
                    'color': score.color,
                    'label': score.label
                })
        
        return {
            'org_unit_id': org_unit_id,
            'org_unit_name': org_unit.name,
            'periods': periods,
            'sector_trend': sector_trend,
            'objectives_trend': objectives_trend
        }
    
    def _get_comparative_analysis_data(self, export_job, parameters):
        """
        Get comparative analysis data for export
        """
        # Get org units to compare
        org_unit_ids = parameters.get('org_unit_ids', [])
        if not org_unit_ids:
            return {'error': 'org_unit_ids is required for comparative analysis'}
        
        org_units = OrgUnit.objects.filter(id__in=org_unit_ids)
        
        # Get assessment period
        assessment_period = parameters.get('assessment_period')
        if not assessment_period:
            current_period = AssessmentPeriod.objects.filter(is_current=True).first()
            assessment_period = current_period.period if current_period else None
        
        if not assessment_period:
            return {'error': 'No assessment period available'}
        
        # Get sector scores for comparison
        sector_scores = SectorScore.objects.filter(
            org_unit__in=org_units,
            assessment_period=assessment_period
        ).select_related('org_unit')
        
        # Get objective scores for comparison
        objective_scores = ObjectiveScore.objects.filter(
            org_unit__in=org_units,
            assessment_period=assessment_period
        ).select_related('objective', 'org_unit')
        
        # Build comparison data
        comparison_data = {
            'assessment_period': assessment_period,
            'org_units': [],
            'sector_comparison': [],
            'objectives_comparison': {}
        }
        
        for org_unit in org_units:
            org_unit_sector_score = sector_scores.filter(org_unit=org_unit).first()
            org_unit_objective_scores = objective_scores.filter(org_unit=org_unit)
            
            org_unit_data = {
                'org_unit_id': org_unit.id,
                'org_unit_name': org_unit.name,
                'org_unit_level': org_unit.level.name,
                'sector_score': {
                    'score': org_unit_sector_score.score if org_unit_sector_score else None,
                    'color': org_unit_sector_score.color if org_unit_sector_score else None,
                    'label': org_unit_sector_score.label if org_unit_sector_score else None
                },
                'objectives': [
                    {
                        'objective_id': score.objective.id,
                        'objective_name': score.objective.name,
                        'score': score.score,
                        'color': score.color,
                        'label': score.label
                    }
                    for score in org_unit_objective_scores
                ]
            }
            
            comparison_data['org_units'].append(org_unit_data)
            comparison_data['sector_comparison'].append({
                'org_unit_name': org_unit.name,
                'score': org_unit_sector_score.score if org_unit_sector_score else None,
                'color': org_unit_sector_score.color if org_unit_sector_score else None,
                'label': org_unit_sector_score.label if org_unit_sector_score else None
            })
        
        return comparison_data
    
    def _generate_periods_back(self, current_period, periods_back):
        """
        Generate list of periods going back from current period
        """
        periods = [current_period]
        
        for i in range(1, periods_back):
            # Parse current period (assuming YYYYMM format)
            year = int(current_period[:4])
            month = int(current_period[4:])
            
            # Calculate previous month
            if month == 1:
                prev_month = 12
                prev_year = year - 1
            else:
                prev_month = month - 1
                prev_year = year
            
            prev_period = f"{prev_year:04d}{prev_month:02d}"
            periods.append(prev_period)
            
            # Update for next iteration
            current_period = prev_period
        
        return periods[::-1]  # Reverse to get chronological order
    
    def _generate_export_file(self, export_job, export_data):
        """
        Generate export file based on format
        """
        export_format = export_job.export_format
        
        if export_format == ExportTemplate.ExportFormat.EXCEL:
            return self._generate_excel_file(export_job, export_data)
        elif export_format == ExportTemplate.ExportFormat.CSV:
            return self._generate_csv_file(export_job, export_data)
        elif export_format == ExportTemplate.ExportFormat.PDF:
            return self._generate_pdf_file(export_job, export_data)
        elif export_format == ExportTemplate.ExportFormat.HTML:
            return self._generate_html_file(export_job, export_data)
        else:
            raise ValueError(f"Unsupported export format: {export_format}")
    
    def _generate_excel_file(self, export_job, export_data):
        """
        Generate Excel file
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Export Data"
            
            # Add header
            ws['A1'] = f"Export: {export_job.name}"
            ws['A2'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A3'] = f"Export Type: {export_job.get_export_type_display()}"
            
            # Style header
            header_font = Font(bold=True, size=12)
            ws['A1'].font = header_font
            ws['A2'].font = Font(size=10)
            ws['A3'].font = Font(size=10)
            
            # Add data based on export type
            if export_job.export_type == ExportTemplate.ExportType.DASHBOARD_SUMMARY:
                self._add_dashboard_summary_to_excel(ws, export_data)
            elif export_job.export_type == ExportTemplate.ExportType.OBJECTIVE_PERFORMANCE:
                self._add_objective_performance_to_excel(ws, export_data)
            elif export_job.export_type == ExportTemplate.ExportType.INDICATOR_PERFORMANCE:
                self._add_indicator_performance_to_excel(ws, export_data)
            elif export_job.export_type == ExportTemplate.ExportType.ORG_UNIT_PERFORMANCE:
                self._add_org_unit_performance_to_excel(ws, export_data)
            elif export_job.export_type == ExportTemplate.ExportType.TREND_ANALYSIS:
                self._add_trend_analysis_to_excel(ws, export_data)
            elif export_job.export_type == ExportTemplate.ExportType.COMPARATIVE_ANALYSIS:
                self._add_comparative_analysis_to_excel(ws, export_data)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            filename = f"export_{export_job.job_id}.xlsx"
            file_path = os.path.join('exports', filename)
            
            # Save to storage
            with default_storage.open(file_path, 'wb') as f:
                wb.save(f)
            
            # Get file size
            file_size = default_storage.size(file_path)
            
            return file_path, file_size
            
        except ImportError:
            raise ValueError("openpyxl is required for Excel export")
        except Exception as e:
            raise ValueError(f"Failed to generate Excel file: {str(e)}")
    
    def _generate_csv_file(self, export_job, export_data):
        """
        Generate CSV file
        """
        import csv
        import io
        
        try:
            # Create CSV content
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Add header
            writer.writerow([f"Export: {export_job.name}"])
            writer.writerow([f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            writer.writerow([f"Export Type: {export_job.get_export_type_display()}"])
            writer.writerow([])
            
            # Add data based on export type
            if export_job.export_type == ExportTemplate.ExportType.DASHBOARD_SUMMARY:
                self._add_dashboard_summary_to_csv(writer, export_data)
            elif export_job.export_type == ExportTemplate.ExportType.OBJECTIVE_PERFORMANCE:
                self._add_objective_performance_to_csv(writer, export_data)
            elif export_job.export_type == ExportTemplate.ExportType.INDICATOR_PERFORMANCE:
                self._add_indicator_performance_to_csv(writer, export_data)
            elif export_job.export_type == ExportTemplate.ExportType.ORG_UNIT_PERFORMANCE:
                self._add_org_unit_performance_to_csv(writer, export_data)
            elif export_job.export_type == ExportTemplate.ExportType.TREND_ANALYSIS:
                self._add_trend_analysis_to_csv(writer, export_data)
            elif export_job.export_type == ExportTemplate.ExportType.COMPARATIVE_ANALYSIS:
                self._add_comparative_analysis_to_csv(writer, export_data)
            
            # Save file
            filename = f"export_{export_job.job_id}.csv"
            file_path = os.path.join('exports', filename)
            
            # Save to storage
            content = output.getvalue()
            default_storage.save(file_path, ContentFile(content.encode('utf-8')))
            
            # Get file size
            file_size = default_storage.size(file_path)
            
            return file_path, file_size
            
        except Exception as e:
            raise ValueError(f"Failed to generate CSV file: {str(e)}")
    
    def _generate_pdf_file(self, export_job, export_data):
        """
        Generate PDF file
        """
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            import io
            
            # Create PDF content
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            story = []
            
            # Get styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30
            )
            
            # Add title
            story.append(Paragraph(f"Export: {export_job.name}", title_style))
            story.append(Paragraph(f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            story.append(Paragraph(f"Export Type: {export_job.get_export_type_display()}", styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Add data based on export type
            if export_job.export_type == ExportTemplate.ExportType.DASHBOARD_SUMMARY:
                self._add_dashboard_summary_to_pdf(story, export_data, styles)
            elif export_job.export_type == ExportTemplate.ExportType.OBJECTIVE_PERFORMANCE:
                self._add_objective_performance_to_pdf(story, export_data, styles)
            elif export_job.export_type == ExportTemplate.ExportType.INDICATOR_PERFORMANCE:
                self._add_indicator_performance_to_pdf(story, export_data, styles)
            elif export_job.export_type == ExportTemplate.ExportType.ORG_UNIT_PERFORMANCE:
                self._add_org_unit_performance_to_pdf(story, export_data, styles)
            elif export_job.export_type == ExportTemplate.ExportType.TREND_ANALYSIS:
                self._add_trend_analysis_to_pdf(story, export_data, styles)
            elif export_job.export_type == ExportTemplate.ExportType.COMPARATIVE_ANALYSIS:
                self._add_comparative_analysis_to_pdf(story, export_data, styles)
            
            # Build PDF
            doc.build(story)
            
            # Save file
            filename = f"export_{export_job.job_id}.pdf"
            file_path = os.path.join('exports', filename)
            
            # Save to storage
            pdf_content = buffer.getvalue()
            default_storage.save(file_path, ContentFile(pdf_content))
            
            # Get file size
            file_size = default_storage.size(file_path)
            
            return file_path, file_size
            
        except ImportError:
            raise ValueError("reportlab is required for PDF export")
        except Exception as e:
            raise ValueError(f"Failed to generate PDF file: {str(e)}")
    
    def _generate_html_file(self, export_job, export_data):
        """
        Generate HTML file
        """
        try:
            # Create HTML content
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Export: {export_job.name}</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 20px; }}
                    .header {{ background-color: #f0f0f0; padding: 10px; margin-bottom: 20px; }}
                    .section {{ margin-bottom: 20px; }}
                    .section h2 {{ color: #333; }}
                    table {{ border-collapse: collapse; width: 100%; }}
                    th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                    th {{ background-color: #f2f2f2; }}
                    .score {{ font-weight: bold; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <h1>Export: {export_job.name}</h1>
                    <p>Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                    <p>Export Type: {export_job.get_export_type_display()}</p>
                </div>
            """
            
            # Add data based on export type
            if export_job.export_type == ExportTemplate.ExportType.DASHBOARD_SUMMARY:
                html_content += self._add_dashboard_summary_to_html(export_data)
            elif export_job.export_type == ExportTemplate.ExportType.OBJECTIVE_PERFORMANCE:
                html_content += self._add_objective_performance_to_html(export_data)
            elif export_job.export_type == ExportTemplate.ExportType.INDICATOR_PERFORMANCE:
                html_content += self._add_indicator_performance_to_html(export_data)
            elif export_job.export_type == ExportTemplate.ExportType.ORG_UNIT_PERFORMANCE:
                html_content += self._add_org_unit_performance_to_html(export_data)
            elif export_job.export_type == ExportTemplate.ExportType.TREND_ANALYSIS:
                html_content += self._add_trend_analysis_to_html(export_data)
            elif export_job.export_type == ExportTemplate.ExportType.COMPARATIVE_ANALYSIS:
                html_content += self._add_comparative_analysis_to_html(export_data)
            
            html_content += """
            </body>
            </html>
            """
            
            # Save file
            filename = f"export_{export_job.job_id}.html"
            file_path = os.path.join('exports', filename)
            
            # Save to storage
            default_storage.save(file_path, ContentFile(html_content.encode('utf-8')))
            
            # Get file size
            file_size = default_storage.size(file_path)
            
            return file_path, file_size
            
        except Exception as e:
            raise ValueError(f"Failed to generate HTML file: {str(e)}")
    
    def _generate_file_url(self, file_path):
        """
        Generate file URL for download
        """
        # This would typically generate a signed URL or use your media URL configuration
        # For now, return a simple path
        return f"/media/{file_path}"
    
    # Excel helper methods (placeholder implementations)
    def _add_dashboard_summary_to_excel(self, ws, data):
        """Add dashboard summary data to Excel worksheet"""
        # Implementation would go here
        pass
    
    def _add_objective_performance_to_excel(self, ws, data):
        """Add objective performance data to Excel worksheet"""
        # Implementation would go here
        pass
    
    def _add_indicator_performance_to_excel(self, ws, data):
        """Add indicator performance data to Excel worksheet"""
        # Implementation would go here
        pass
    
    def _add_org_unit_performance_to_excel(self, ws, data):
        """Add org unit performance data to Excel worksheet"""
        # Implementation would go here
        pass
    
    def _add_trend_analysis_to_excel(self, ws, data):
        """Add trend analysis data to Excel worksheet"""
        # Implementation would go here
        pass
    
    def _add_comparative_analysis_to_excel(self, ws, data):
        """Add comparative analysis data to Excel worksheet"""
        # Implementation would go here
        pass
    
    # CSV helper methods (placeholder implementations)
    def _add_dashboard_summary_to_csv(self, writer, data):
        """Add dashboard summary data to CSV"""
        # Implementation would go here
        pass
    
    def _add_objective_performance_to_csv(self, writer, data):
        """Add objective performance data to CSV"""
        # Implementation would go here
        pass
    
    def _add_indicator_performance_to_csv(self, writer, data):
        """Add indicator performance data to CSV"""
        # Implementation would go here
        pass
    
    def _add_org_unit_performance_to_csv(self, writer, data):
        """Add org unit performance data to CSV"""
        # Implementation would go here
        pass
    
    def _add_trend_analysis_to_csv(self, writer, data):
        """Add trend analysis data to CSV"""
        # Implementation would go here
        pass
    
    def _add_comparative_analysis_to_csv(self, writer, data):
        """Add comparative analysis data to CSV"""
        # Implementation would go here
        pass
    
    # PDF helper methods (placeholder implementations)
    def _add_dashboard_summary_to_pdf(self, story, data, styles):
        """Add dashboard summary data to PDF"""
        # Implementation would go here
        pass
    
    def _add_objective_performance_to_pdf(self, story, data, styles):
        """Add objective performance data to PDF"""
        # Implementation would go here
        pass
    
    def _add_indicator_performance_to_pdf(self, story, data, styles):
        """Add indicator performance data to PDF"""
        # Implementation would go here
        pass
    
    def _add_org_unit_performance_to_pdf(self, story, data, styles):
        """Add org unit performance data to PDF"""
        # Implementation would go here
        pass
    
    def _add_trend_analysis_to_pdf(self, story, data, styles):
        """Add trend analysis data to PDF"""
        # Implementation would go here
        pass
    
    def _add_comparative_analysis_to_pdf(self, story, data, styles):
        """Add comparative analysis data to PDF"""
        # Implementation would go here
        pass
    
    # HTML helper methods (placeholder implementations)
    def _add_dashboard_summary_to_html(self, data):
        """Add dashboard summary data to HTML"""
        return "<div class='section'><h2>Dashboard Summary</h2><p>Data would be displayed here</p></div>"
    
    def _add_objective_performance_to_html(self, data):
        """Add objective performance data to HTML"""
        return "<div class='section'><h2>Objective Performance</h2><p>Data would be displayed here</p></div>"
    
    def _add_indicator_performance_to_html(self, data):
        """Add indicator performance data to HTML"""
        return "<div class='section'><h2>Indicator Performance</h2><p>Data would be displayed here</p></div>"
    
    def _add_org_unit_performance_to_html(self, data):
        """Add org unit performance data to HTML"""
        return "<div class='section'><h2>Org Unit Performance</h2><p>Data would be displayed here</p></div>"
    
    def _add_trend_analysis_to_html(self, data):
        """Add trend analysis data to HTML"""
        return "<div class='section'><h2>Trend Analysis</h2><p>Data would be displayed here</p></div>"
    
    def _add_comparative_analysis_to_html(self, data):
        """Add comparative analysis data to HTML"""
        return "<div class='section'><h2>Comparative Analysis</h2><p>Data would be displayed here</p></div>" 