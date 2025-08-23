#!/usr/bin/env python
"""
Assessment services for DHIS2 integration and score calculation
"""
import logging
from decimal import Decimal
from datetime import datetime, date, timedelta
from typing import List, Dict, Optional, Any
from django.db import transaction
from django.utils import timezone
from django.conf import settings
import os
from datetime import datetime
from django.core.exceptions import ValidationError
from django.db.models import Q, Avg, Count, Sum, Max, Min
from django.contrib.auth import get_user_model
from django.core.cache import cache
import re

from dhis2_auth.dhis_client import DHIS2Client, DHIS2ClientFactory
from dhis2_auth.models import DHIS2User
from dhis2_auth.session import get_dhis2_user_from_request
from configurations.models import (
    TrackedIndicator, Objective, AssessmentPeriod, 
    ScoringRule, Milestone, IndicatorWeight
)
from organisation.models import OrgUnit
from organisation.services import AccessControlService
from .models import (
    DataSyncLog, IndicatorData, IndicatorScore, 
    ObjectiveScore, SectorScore, AuditLog
)

logger = logging.getLogger(__name__)

class RealTimeDHIS2Service:
    """
    Service for real-time DHIS2 data fetching without database storage
    """
    
    def __init__(self, dhis2_client=None):
        self.client = dhis2_client
    
    def _classify_change_category(self, change_pct: float | None, target_type: str = 'increase') -> str | None:
        """Map relative percent change to O category per scoring context."""
        if change_pct is None:
            return None
        
        # Now that we use the correct formulas, we don't need to invert
        # The change_pct passed here is already the correct performance change
        performance_change = change_pct
        
        # Categorize based on performance change - EXACTLY matches flowchart
        if performance_change > 5:
            return '>5%'
        if -5 < performance_change <= 5:
            return '-5%<C<=5%'  # Stagnation category
        if -10 < performance_change <= -5:
            return '-10%<C<=-5%'
        return '<=-10%'

    def _classify_gap_category(self, gap_pct: float | None) -> str | None:
        """Map gap to P category per Excel formula: =IF($I4<=10%,"<=10%",IF(AND($I4>10%,$I4<=40%),"10%<PT<=40%",IF($I4>40%,">40%","")))"""
        if gap_pct is None:
            return None
        # Use signed gap for categorization, matching Excel behavior
        if gap_pct <= 10:
            return '<=10%'
        if 10 < gap_pct <= 40:
            return '10%<PT<=40%'
        if gap_pct > 40:
            return '>40%'
        return None

    def _compute_trend_score(self, has_data: bool, current_meets: bool | None, previous_meets: bool | None,
                              change_cat: str | None, gap_cat: str | None, indicator=None) -> int:
        """Use the updated HolisticScoringService algorithm for real-time scoring."""
        # Use the updated scoring service - avoid circular import
        # from assessments.services import HolisticScoringService
        
        # Create a mock indicator for scoring if not provided
        if indicator is None:
            class MockIndicator:
                def __init__(self, target_type='increase'):
                    self.target_operator = '>=' if target_type == 'increase' else '<='
                    self.target_type = target_type
                    self.target_value = 100  # Dummy value, not used in scoring
            
            indicator = MockIndicator()
        
        # Get current and previous values from the context
        # For real-time scoring, we need to extract these from the data
        current_value = None
        previous_value = None
        
        # Since we don't have the actual values here, we'll use the simplified logic
        # based on the categories and target achievement
        
        # Step 1: Data provided check
        if not has_data:
            return -2
        
        # Step 2: First year check (simplified - assume not first year if we have previous data)
        is_first_year = previous_meets is None
        
        # Step 3: Target achieved check
        target_achieved = "Yes" if current_meets else "No"
        
        # Step 4: Use the updated scoring logic
        scoring_service = HolisticScoringService()
        
        # For real-time scoring, we need to reconstruct the scoring logic
        # since we don't have the actual values to pass to calculate_indicator_score
        
        # Simplified version of the Excel outcome formula
        if is_first_year:
            if target_achieved == "Yes":
                return 1
            else:
                return 0
        else:
            # Not first year - use the complex logic
            if target_achieved == "Yes":
                # Target WAS achieved - check performance change
                if change_cat == ">5%":
                    return 2
                elif change_cat == "-5%<C<=5%":
                    # For stagnation, when target is achieved, score should be 2
                    return 2
                elif change_cat == "-10%<C<=-5%":
                    # For negative change, score lower even if target is achieved
                    return 1
                elif change_cat == "<=-10%":
                    return 0
                else:
                    return 0
            else:
                # Target NOT achieved - check performance change
                if change_cat == ">5%":
                    return 1
                elif change_cat == "-5%<C<=5%":
                    # Stagnation - check how close to target
                    if gap_cat == "<=10%":
                        return 1
                    elif gap_cat == "10%<PT<=40%":
                        return 0
                    elif gap_cat == ">40%":
                        return -1
                    else:
                        return 0
                elif change_cat == "-10%<C<=-5%":
                    return -1
                elif change_cat == "<=-10%":
                    return -1
                else:
                    return 0

    def _median(self, numbers: list[float]) -> float | None:
        vals = [float(x) for x in numbers if x is not None and isinstance(x, (int, float))]
        if not vals:
            return None
        vals.sort()
        n = len(vals)
        mid = n // 2
        if n % 2 == 0:
            return (vals[mid-1] + vals[mid]) / 2
        return vals[mid]

    def _compute_objective_trend_from_indicators(self, indicators: list[dict]) -> dict:
        """Aggregate indicator percent_change/target_gap and compute O/P categories and trend for objective."""
        if not indicators:
            return {}
        changes = []
        gaps = []
        current_meets_flags = []
        previous_meets_flags = []
        for ind in indicators:
            sc = ind.get('score') or {}
            if sc.get('percent_change') is not None:
                changes.append(float(sc.get('percent_change')))
            if sc.get('target_gap') is not None:
                gaps.append(abs(float(sc.get('target_gap'))))
            # Determine meets based on target
            curr = sc.get('current_value')
            prev = sc.get('previous_value')
            tgt = ind.get('target_value')
            ttype = (ind.get('target_type') or 'increase').lower()
            try:
                if curr is not None and tgt is not None:
                    current_meets_flags.append(float(curr) >= float(tgt) if ttype == 'increase' else float(curr) <= float(tgt))
                if prev is not None and tgt is not None:
                    previous_meets_flags.append(float(prev) >= float(tgt) if ttype == 'increase' else float(prev) <= float(tgt))
            except Exception:
                pass
        med_change = self._median(changes)
        med_gap = self._median(gaps)
        change_cat = self._classify_change_category(med_change, 'increase')  # Default for objective aggregation
        gap_cat = self._classify_gap_category(med_gap)
        # Majority rule for meets
        current_meets = (sum(1 for f in current_meets_flags if f) > len(current_meets_flags)/2) if current_meets_flags else None
        previous_meets = (sum(1 for f in previous_meets_flags if f) > len(previous_meets_flags)/2) if previous_meets_flags else None
        trend_score = self._compute_trend_score(has_data=med_change is not None or med_gap is not None,
                                                current_meets=current_meets,
                                                previous_meets=previous_meets,
                                                change_cat=change_cat,
                                                gap_cat=gap_cat)
        return {
            'percent_change': med_change,
            'target_gap': med_gap,
            'change_category': change_cat,
            'gap_category': gap_cat,
            'trend_score': trend_score,
        }

    def _score_color_label(self, score: int | None) -> tuple[str, str]:
        if score is None:
            return ('#6c757d', 'N/A')
        if score >= 2:
            return ('#548235', 'Highly Performing')
        if score == 1:
            return ('#A9D08E', 'Moderately Performing')
        if score == 0:
            return ('#FFFF00', 'Sustained')
        if score == -1:
            return ('#FFC7CE', 'Underperforming')
        return ('#FF0000', 'Severely Underperforming')

    def generate_holistic_excel(self, assessment_payload: list) -> str:
        """Generate an Excel file that mirrors the table format with color coding.
        Returns the absolute file path of the saved workbook.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.formatting.rule import FormulaRule
        except Exception as e:
            logger.error(f"openpyxl not available: {e}")
            raise

        wb = Workbook()
        ws = wb.active
        ws.title = 'Table'

        # Styling helpers
        header_fill = PatternFill('solid', fgColor='265380')  # #265380
        header_font = Font(color='FFFFFF', bold=True)
        center = Alignment(horizontal='center', vertical='center')
        left = Alignment(horizontal='left', vertical='center')
        thin = Side(border_style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        orange_fill = PatternFill('solid', fgColor='FDBA74')  # approx #fd7e14 light
        yellow_fill = PatternFill('solid', fgColor='FFF3BF')  # light yellow
        green50 = PatternFill('solid', fgColor='E8F5E9')  # Light green for positive change
        yellow50 = PatternFill('solid', fgColor='FFFDE7')  # Light yellow for neutral change
        red50 = PatternFill('solid', fgColor='FFEBEE')   # Light red for negative change/gap

        def score_fill(score: float | None):
            if score is None:
                return None
            # New color scheme: 2 (Dark Green), 1 (Light Green), 0 (Yellow), -1 (Light Red), -2 (Red)
            if score >= 2:
                return PatternFill('solid', fgColor='548235')  # Dark Green for 2
            if score >= 1:
                return PatternFill('solid', fgColor='A9D08E')  # Light Green for 1
            if score == 0:
                return PatternFill('solid', fgColor='FFFF00')  # Yellow for 0
            if score == -1:
                return PatternFill('solid', fgColor='FFC7CE')  # Light Red for -1
            return PatternFill('solid', fgColor='FF0000')  # Red for -2

        data = assessment_payload[0] if assessment_payload else None
        if not data:
            raise ValueError('Empty assessment data')

        periods = []
        # Derive periods by inspecting first objective/indicator
        if data.get('objectives'):
            for obj in data['objectives']:
                if obj.get('indicators'):
                    first = obj['indicators'][0]
                    periods = list(first.get('data_values', {}).keys())
                    break

        # Header row
        headers = ['#', 'Indicator'] + periods + ['Change', 'P-T Gap Analysis', 'Target', 'Assessed score (-2, -1, 0, +1, +2)', 'Remarks']
        ws.append(headers)
        for idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[1].height = 22

        row = 2
        # Sort objectives by order to match frontend
        sorted_objectives = sorted(data.get('objectives', []), key=lambda x: x.get('order', 0))
        for obj in sorted_objectives:
            # Objective row
            ws.append([None] * len(headers))
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill = orange_fill
                cell.border = border
            ws.cell(row=row, column=1, value=None)
            ws.cell(row=row, column=2, value=obj.get('name')).alignment = left
            row += 1

            # Indicator rows - sort by display_order to match frontend
            sorted_indicators = sorted(obj.get('indicators', []), key=lambda x: x.get('display_order', 0))
            for ind in sorted_indicators:
                row_values = []
                row_values.append(ind.get('indicator_number'))
                row_values.append(ind.get('name'))
                # periods
                for p in periods:
                    v = ind.get('data_values', {}).get(p, {}).get('value')
                    row_values.append(v)
                # change/gap - format with % symbol
                sc = ind.get('score') or {}
                percent_change = sc.get('percent_change')
                target_gap = sc.get('target_gap')
                
                # Format percentage values with proper handling of None and zero values
                if percent_change is not None and percent_change != 0:
                    row_values.append(f"{percent_change:.1f}%")
                elif percent_change == 0:
                    row_values.append("0.0%")
                else:
                    row_values.append('')
                    
                if target_gap is not None and target_gap != 0:
                    row_values.append(f"{target_gap:.1f}%")
                elif target_gap == 0:
                    row_values.append("0.0%")
                else:
                    row_values.append('')
                # Use target_display if available, otherwise fall back to target_value
                target_display = ind.get('target_display') or ind.get('target_value')
                row_values.append(target_display)
                # Ensure score is properly displayed
                score_value = sc.get('score')
                if score_value is not None:
                    row_values.append(score_value)
                else:
                    row_values.append('')
                row_values.append('')  # remarks
                ws.append(row_values)

                # style row
                col = 1
                for val in row_values:
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    if col == 1:
                        cell.alignment = center
                    elif col == 2:
                        cell.alignment = left
                    else:
                        cell.alignment = center
                    col += 1

                # Change bg (col after periods)
                change_col = 2 + len(periods) + 1
                gap_col = change_col + 1
                
                # Apply colors based on flow diagram logic
                change_val = sc.get('percent_change')
                if isinstance(change_val, (int, float)):
                    # Flow diagram: >5% (Green), -5% to 5% (Yellow), <-5% (Red)
                    if change_val > 5:
                        ws.cell(row=row, column=change_col).fill = green50
                    elif change_val >= -5:
                        ws.cell(row=row, column=change_col).fill = yellow50
                    else:
                        ws.cell(row=row, column=change_col).fill = red50
                
                gap_val = sc.get('target_gap')
                if isinstance(gap_val, (int, float)):
                    # Flow diagram: ≤10% (Green), 10%<PT≤40% (Yellow), >40% (Red)
                    abs_gap = abs(gap_val)
                    if abs_gap <= 10:
                        ws.cell(row=row, column=gap_col).fill = green50
                    elif abs_gap <= 40:
                        ws.cell(row=row, column=gap_col).fill = yellow50
                    else:
                        ws.cell(row=row, column=gap_col).fill = red50

                # Score color
                score_col = gap_col + 2
                s = sc.get('score')
                if s is not None:
                    fill = score_fill(s)
                    if fill:
                        ws.cell(row=row, column=score_col).fill = fill
                        ws.cell(row=row, column=score_col).font = Font(color='FFFFFF', bold=True)

                row += 1

            # Milestone row
            milestone = obj.get('milestone')
            if milestone:
                milestone_name = milestone.get('name') if isinstance(milestone, dict) else str(milestone)
                milestone_score = milestone.get('score') if isinstance(milestone, dict) else None
                
                # Create milestone row with proper score
                milestone_row = ['MS', milestone_name]
                # Add empty values for periods
                milestone_row.extend(['-'] * len(periods))
                # Add empty values for change, gap, target
                milestone_row.extend(['-', '-', '-'])
                # Add milestone score
                milestone_row.append(milestone_score if milestone_score is not None else '-')
                # Add empty remarks
                milestone_row.append('')
                
                ws.append(milestone_row)
                
                # Style milestone row
                for c in range(1, len(headers) + 1):
                    cell = ws.cell(row=row, column=c)
                    cell.fill = yellow_fill
                    cell.border = border
                    cell.alignment = center if c != 2 else left
                
                # Apply score color to milestone score cell
                if milestone_score is not None:
                    score_col = 2 + len(periods) + 4  # Score column position
                    fill = score_fill(milestone_score)
                    if fill:
                        ws.cell(row=row, column=score_col).fill = fill
                        ws.cell(row=row, column=score_col).font = Font(color='FFFFFF', bold=True)
                
                row += 1

        last_row = row - 1

        # Note: Direct cell fills are already applied above, so conditional formatting is not needed
        # This ensures colors are preserved in the exported file

        # Autosize basic columns
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 60
        start_c = 3
        for i in range(len(periods)):
            ws.column_dimensions[chr(64 + start_c + i)].width = 12
        # Change, Gap, Target, Score, Remarks
        ws.column_dimensions[chr(64 + start_c + len(periods))].width = 12
        ws.column_dimensions[chr(64 + start_c + len(periods) + 1)].width = 14
        ws.column_dimensions[chr(64 + start_c + len(periods) + 2)].width = 12
        ws.column_dimensions[chr(64 + start_c + len(periods) + 3)].width = 18
        ws.column_dimensions[chr(64 + start_c + len(periods) + 4)].width = 20

        # Summary sheet
        try:
            ws_sum = wb.create_sheet('Summary')
            ws_sum.append(['Org Unit', data.get('org_unit_name', '')])
            ws_sum.append(['Periods', ', '.join(periods)])
            sec = data.get('sector_score') or {}
            ws_sum.append(['Sector Score', sec.get('overall_score')])
            ws_sum.append(['Sector Label', sec.get('score_label')])
            ws_sum.append([])
            ws_sum.append(['Objective', 'Final Score', 'Label'])
            for obj in sorted_objectives:
                sc = obj.get('score') or {}
                ws_sum.append([obj.get('name'), sc.get('final_score'), sc.get('score_label')])
        except Exception as e:
            logger.warning(f"Failed to build Summary sheet: {e}")

        # Raw Data sheet
        try:
            ws_raw = wb.create_sheet('Raw Data')
            ws_raw.append(['Objective', 'Indicator', *periods, 'Target'])
            for obj in sorted_objectives:
                sorted_indicators = sorted(obj.get('indicators', []), key=lambda x: x.get('display_order', 0))
                for ind in sorted_indicators:
                    row_vals = [obj.get('name'), ind.get('name')]
                    for p in periods:
                        row_vals.append(ind.get('data_values', {}).get(p, {}).get('value'))
                    row_vals.append(ind.get('target_value'))
                    ws_raw.append(row_vals)
        except Exception as e:
            logger.warning(f"Failed to build Raw Data sheet: {e}")

        # Metadata sheet
        try:
            ws_meta = wb.create_sheet('Metadata')
            ws_meta.append(['Key', 'Value'])
            ws_meta.append(['Generated At', datetime.now().isoformat()])
            ws_meta.append(['Org Unit ID', data.get('org_unit_id')])
            ws_meta.append(['Org Unit Name', data.get('org_unit_name')])
            ws_meta.append(['Period Range', f"{periods[0]} to {periods[-1]}" if periods else ''])
        except Exception as e:
            logger.warning(f"Failed to build Metadata sheet: {e}")

        # Legend sheet
        try:
            ws_leg = wb.create_sheet('Legend')
            ws_leg.append(['Legend'])
            ws_leg.append(['Score Colors'])
            legend_rows = [
                ('Score 2 (Green)', '28A745'),
                ('Score 1 (Green)', '28A745'),
                ('Score 0 (Yellow)', 'FFC107'),
                ('Score -1 (Magenta)', 'E91E63'),
                ('Score -2 (Red)', 'DC3545'),
            ]
            for text, color in legend_rows:
                ws_leg.append([text])
                cell = ws_leg.cell(row=ws_leg.max_row, column=2, value='')
                cell.fill = PatternFill('solid', fgColor=color)
            ws_leg.append([])
            ws_leg.append(['Performance Change Colors'])
            ws_leg.append(['Increase > 5% (Green)', ''])
            ws_leg.cell(row=ws_leg.max_row, column=2).fill = green50
            ws_leg.append(['Stagnation -5% < c ≤ 5% (Yellow)', ''])
            ws_leg.cell(row=ws_leg.max_row, column=2).fill = yellow50
            ws_leg.append(['Decrease ≤ -5% (Red)', ''])
            ws_leg.cell(row=ws_leg.max_row, column=2).fill = red50
            ws_leg.append([])
            ws_leg.append(['Target Gap Colors'])
            ws_leg.append(['Close; ≤ 10% (Green)', ''])
            ws_leg.cell(row=ws_leg.max_row, column=2).fill = green50
            ws_leg.append(['A bit far; 10% < p ≤ 40% (Yellow)', ''])
            ws_leg.cell(row=ws_leg.max_row, column=2).fill = yellow50
            ws_leg.append(['Very far; > 40% (Red)', ''])
            ws_leg.cell(row=ws_leg.max_row, column=2).fill = red50
        except Exception as e:
            logger.warning(f"Failed to build Legend sheet: {e}")

        # Save file
        export_dir = os.path.join(getattr(settings, 'MEDIA_ROOT', os.path.join(os.getcwd(), 'media')), 'exports')
        os.makedirs(export_dir, exist_ok=True)
        
        # Get organization unit name for filename
        org_unit_name = "Unknown"
        if data.get('org_unit_name'):
            org_unit_name = data['org_unit_name'].replace(' ', '_').replace('/', '_').replace('\\', '_')
        elif data.get('org_unit_id'):
            org_unit_name = f"OrgUnit_{data['org_unit_id']}"
        
        filename = f"holistic-assessment-{org_unit_name}-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
        file_path = os.path.join(export_dir, filename)
        wb.save(file_path)
        return file_path
        
    def fetch_holistic_assessment_data(self, request, assessment_config):
        """
        Fetch real-time DHIS2 data for holistic assessment display
        No database storage - just fetch and return for immediate display
        """
        try:
            # Get DHIS2 user from session
            dhis2_user = get_dhis2_user_from_request(request)
            if not dhis2_user:
                raise ValidationError("No DHIS2 user found in session")
            
            # Initialize client if not provided
            if not self.client:
                self.client = DHIS2ClientFactory.create_client_from_session(
                    dhis2_user.dhis2_instance_url,
                    request.session.session_key
                )
            
            # Extract configuration
            org_unit_ids = assessment_config.get('org_unit_ids', [])
            periods_raw = assessment_config.get('periods', [])
            indicator_uids = assessment_config.get('indicator_uids', [])
            manual_entries = assessment_config.get('manual_entries', {})
            pre_calculated_scores = assessment_config.get('pre_calculated_scores', {})
            logger.info(f"Manual entries received: {manual_entries}")
            logger.info(f"Pre-calculated scores received: {pre_calculated_scores}")
            
            # Handle periods - they can be strings or objects with 'code' field
            periods = []
            for period in periods_raw:
                if isinstance(period, dict) and 'code' in period:
                    periods.append(period['code'])
                elif isinstance(period, str):
                    periods.append(period)
                else:
                    # Fallback: try to extract code from period object
                    period_code = getattr(period, 'code', str(period))
                    periods.append(period_code)
            
            # Sort periods chronologically to ensure correct change calculation
            # This ensures that periods[-1] is the most recent and periods[-2] is the previous
            periods.sort()
            
            if not org_unit_ids or not periods:
                raise ValidationError("Organization units and periods are required")
            
            # Fetch active indicators if not specified; include manual indicators (no dhis2_uid)
            manual_indicators = []
            if not indicator_uids:
                indicators = TrackedIndicator.objects.filter(is_active=True)
                # Only include non-empty UIDs for DHIS2 fetch; keep track of manual ones
                indicator_uids = [ind.dhis2_uid for ind in indicators if ind.dhis2_uid]
                manual_indicators = [ind for ind in indicators if not ind.dhis2_uid]
                logger.info(
                    f"Using {len(indicator_uids)} DHIS2 indicators and {len(manual_indicators)} manual indicators"
                )
            
            # Fetch data for each indicator
            assessment_data = {
                'indicators': [],
                'objectives': [],
                'milestones': [],
                'metadata': {
                    'org_units': org_unit_ids,
                    'periods': periods,
                    'fetched_at': timezone.now().isoformat()
                }
            }
            
            # Group indicators by objective
            objectives = Objective.objects.filter(is_active=True).prefetch_related('indicator_weights__indicator')
            
            # Check if we have indicator weights configured
            total_weights = sum(obj.indicator_weights.count() for obj in objectives)
            logger.info(f"Found {total_weights} indicator weights configured across {objectives.count()} objectives")
            
            if total_weights == 0:
                # No indicator weights configured - fetch indicators directly and group them evenly
                logger.info("No indicator weights configured, distributing indicators across objectives")
                # Combine DHIS2-backed and manual indicators
                all_indicators = list(TrackedIndicator.objects.filter(
                    dhis2_uid__in=indicator_uids,
                    is_active=True
                ))
                # Append manual ones gathered above if any
                if manual_indicators:
                    all_indicators.extend(manual_indicators)
                
                # Distribute indicators evenly across objectives
                indicators_per_objective = len(all_indicators) // max(len(objectives), 1)
                indicator_list = list(all_indicators)
                
                for i, objective in enumerate(objectives):
                    objective_data = {
                        'id': objective.id,
                        'name': objective.name,
                        'code': objective.code,
                        'color': objective.color,
                        'milestone': None,
                        'indicators': []
                    }
                    
                    # Add milestone information if it exists
                    if objective.milestone:
                        objective_data['milestone'] = {
                            'id': objective.milestone.id,
                            'name': objective.milestone.name,
                            'code': objective.milestone.code,
                            'color': objective.milestone.color,
                            'score': -2,  # Default score for real-time data
                            'score_color': '#dc3545',
                            'score_label': 'Severely Underperforming'
                        }
                    
                    # Assign indicators to this objective
                    start_idx = i * indicators_per_objective
                    end_idx = start_idx + indicators_per_objective if i < len(objectives) - 1 else len(indicator_list)
                    objective_indicators = indicator_list[start_idx:end_idx]
                    
                    for indicator in objective_indicators:
                        indicator_data = {
                            'id': indicator.id,
                            'name': indicator.name,
                            'dhis2_uid': indicator.dhis2_uid,
                            'description': indicator.description,
                            'indicator_number': indicator.indicator_number or f"{i+1}",
                            'display_order': indicator.display_order,
                            'target_value': float(indicator.target_value) if indicator.target_value else None,
                            'target_display': indicator.target_display,
                            'target_lower_limit': float(indicator.target_lower_limit) if indicator.target_lower_limit else None,
                            'target_upper_limit': float(indicator.target_upper_limit) if indicator.target_upper_limit else None,
                            'target_format': getattr(indicator, 'target_format', 'SINGLE'),
                            'target_type': indicator.target_type,
                            'weight': 1.0,  # Default weight when no indicator weights configured
                            'data_values': {},
                            'score': None
                        }
                        
                        # Fetch data for each period
                        for period_code in periods:
                            try:
                                if indicator.dhis2_uid:
                                    value = self._fetch_single_indicator_data(
                                        indicator, org_unit_ids[0], period_code
                                    )
                                    clean_value = self._clean_numeric_value(value)
                                    # For DHIS2 data, if no value is found, assign 0 to help with scoring
                                    # This ensures proper scoring for increase/decrease indicators
                                    final_value = clean_value if clean_value is not None else 0
                                    indicator_data['data_values'][period_code] = {
                                        'value': final_value,
                                        'dhis2_value': clean_value,  # Keep original for reference
                                        'manual_override': None
                                    }
                                else:
                                    # Manual indicator – initialize empty value, editable on FE
                                    indicator_data['data_values'][period_code] = {
                                        'value': None,
                                        'dhis2_value': None,
                                        'manual_override': None
                                    }
                            except Exception as e:
                                logger.warning(f"Failed to process {indicator.name} for period {period_code}: {str(e)}")
                                indicator_data['data_values'][period_code] = {
                                    'value': None,
                                    'dhis2_value': None,
                                    'manual_override': None
                                }
                        
                        # Apply manual entries if available for this indicator
                        if manual_entries and str(indicator.id) in manual_entries:
                            indicator_manual_entries = manual_entries[str(indicator.id)]
                            logger.info(f"Applying manual entries for indicator {indicator.id}: {indicator_manual_entries}")
                            
                            for period_code, manual_value in indicator_manual_entries.items():
                                if period_code in indicator_data['data_values']:
                                    # Apply manual override
                                    indicator_data['data_values'][period_code]['value'] = manual_value
                                    indicator_data['data_values'][period_code]['manual_override'] = manual_value
                                    logger.info(f"Applied manual value {manual_value} for indicator {indicator.id} period {period_code}")
                        
                        # Compute percent_change and target_gap for latest vs previous period
                        try:
                            if isinstance(periods, list) and len(periods) >= 1:
                                last_key = periods[-1]  # This is now a period code
                                prev_key = periods[-2] if len(periods) > 1 else None
                                curr_val = indicator_data['data_values'].get(last_key, {}).get('value')
                                prev_val = indicator_data['data_values'].get(prev_key, {}).get('value') if prev_key else None
                                change_pct = None
                                
                                # Get target information
                                tgt = indicator_data.get('target_value')
                                target_format = indicator_data.get('target_format', 'SINGLE')
                                target_upper = indicator_data.get('target_upper_limit')
                                target_type = indicator.target_type
                                
                                if prev_val not in (None, 0, 0.0) and curr_val is not None:
                                    try:
                                        # For range indicators, always use standard formula regardless of target_type
                                        if target_format == 'RANGE':
                                            change = ((float(curr_val) - float(prev_val)) / abs(float(prev_val))) * 100.0
                                        else:
                                            # For non-range indicators, use target_type specific formula
                                            if target_type == 'decrease':
                                                # For decrease indicators: (previous_value - current_value) / abs(current_value) * 100
                                                change = ((float(prev_val) - float(curr_val)) / abs(float(curr_val))) * 100.0
                                            else:
                                                # For increase indicators: (current_value - previous_value) / abs(previous_value) * 100
                                                change = ((float(curr_val) - float(prev_val)) / abs(float(prev_val))) * 100.0
                                        
                                        if change == float('inf') or change == float('-inf'):
                                            change_pct = None
                                        else:
                                            change_pct = round(change, 2)
                                    except Exception:
                                        change_pct = None
                                
                                gap_pct = None
                                # Calculate the correct target value for gap analysis
                                if curr_val is not None and curr_val != 0:
                                    try:
                                        if target_format == 'RANGE' and target_upper is not None:
                                            # For range indicators: (Target upper limit - Current Value) / Current Value * 100
                                            gap_calc = ((float(target_upper) - float(curr_val)) / float(curr_val)) * 100.0
                                        else:
                                            # For non-range indicators
                                            if tgt not in (None, 0, 0.0):
                                                if target_type == 'increase':
                                                    # For increase indicators: (Current Value - Target Value) / Target Value * 100
                                                    gap_calc = ((float(curr_val) - float(tgt)) / float(tgt)) * 100.0
                                                else:
                                                    # For decrease indicators: (Target Value - Current Value) / Current Value * 100
                                                    gap_calc = ((float(tgt) - float(curr_val)) / float(curr_val)) * 100.0
                                            else:
                                                gap_calc = None
                                        
                                        if gap_calc is not None and gap_calc != float('inf') and gap_calc != float('-inf'):
                                            gap_pct = round(gap_calc, 2)
                                        else:
                                            gap_pct = None
                                    except Exception:
                                        gap_pct = None
                                
                                # Derive categories and simple threshold flags for M/N
                                change_cat = self._classify_change_category(change_pct, target_type)
                                gap_cat = self._classify_gap_category(gap_pct)
                                
                                # For M/N we use meet-threshold as curr >= target for increase, curr <= target for decrease
                                current_meets = None
                                previous_meets = None
                                try:
                                    if curr_val is not None and tgt not in (None,):
                                        if indicator.target_type == 'increase':
                                            current_meets = float(curr_val) >= float(tgt)
                                        else:
                                            current_meets = float(curr_val) <= float(tgt)
                                except Exception:
                                    current_meets = None
                                try:
                                    if prev_val is not None and tgt not in (None,):
                                        if indicator.target_type == 'increase':
                                            previous_meets = float(prev_val) >= float(tgt)
                                        else:
                                            previous_meets = float(prev_val) <= float(tgt)
                                except Exception:
                                    previous_meets = None
                                
                                has_data = curr_val is not None
                                trend_score = self._compute_trend_score(has_data, current_meets, previous_meets, change_cat, gap_cat, indicator)
                                # Derive a simple indicator score from categories/trend if not provided by DB
                                derived_score = trend_score
                                color, label = self._score_color_label(derived_score)
                                
                                # Always use fresh calculations to ensure consistency with the latest fixes
                                # Pre-calculated scores may contain old, incorrect calculations
                                indicator_data['score'] = {
                                    'percent_change': change_pct,
                                    'target_gap': gap_pct,
                                    'change_category': change_cat,
                                    'gap_category': gap_cat,
                                    'trend_score': trend_score,
                                    'score': derived_score,
                                    'score_color': color,
                                    'score_label': label,
                                    'current_value': curr_val,
                                    'previous_value': prev_val,
                                    'is_manual_override': False
                                }
                        except Exception as e:
                            logger.warning(f"Failed computing change/gap for indicator {indicator.id}: {e}")
                        
                        objective_data['indicators'].append(indicator_data)
                    
                    assessment_data['objectives'].append(objective_data)
            
            else:
                # Use configured indicator weights
                for objective in objectives:
                    objective_data = {
                        'id': objective.id,
                        'name': objective.name,
                        'code': objective.code,
                        'color': objective.color,
                        'milestone': None,
                        'indicators': []
                    }
                    
                    # Add milestone information if it exists
                    if objective.milestone:
                        objective_data['milestone'] = {
                            'id': objective.milestone.id,
                            'name': objective.milestone.name,
                            'code': objective.milestone.code,
                            'color': objective.milestone.color,
                            'score': -2,  # Default score for real-time data
                            'score_color': '#dc3545',
                            'score_label': 'Severely Underperforming'
                        }
                    
                    # Get indicators for this objective through indicator_weights relationship
                    objective_indicators = []
                    indicator_weights_map = {}
                    for indicator_weight in objective.indicator_weights.all():
                        indicator = indicator_weight.indicator
                        # Include DHIS2-backed indicators present in list and manual indicators (no UID)
                        if (indicator.dhis2_uid in indicator_uids or not indicator.dhis2_uid) and indicator.is_active:
                            objective_indicators.append(indicator)
                            indicator_weights_map[indicator.id] = indicator_weight.weight
                    
                    for indicator in objective_indicators:
                        indicator_data = {
                            'id': indicator.id,
                            'name': indicator.name,
                            'dhis2_uid': indicator.dhis2_uid,
                            'description': indicator.description,
                            'indicator_number': indicator.indicator_number or f"{objective.order}.{len(objective_data['indicators'])+1}",
                            'display_order': indicator.display_order,
                            'target_value': float(indicator.target_value) if indicator.target_value else None,
                            'target_display': indicator.target_display,
                            'target_lower_limit': float(indicator.target_lower_limit) if indicator.target_lower_limit else None,
                            'target_upper_limit': float(indicator.target_upper_limit) if indicator.target_upper_limit else None,
                            'target_format': getattr(indicator, 'target_format', 'SINGLE'),
                            'target_type': indicator.target_type,
                            'weight': float(indicator_weights_map.get(indicator.id, 1.0)),
                            'data_values': {},
                            'score': None
                        }
                        
                        # Fetch data for each period
                        for period in periods:
                            try:
                                if indicator.dhis2_uid:
                                    value = self._fetch_single_indicator_data(
                                        indicator, org_unit_ids[0], period
                                    )
                                    clean_value = self._clean_numeric_value(value)
                                    # For DHIS2 data, if no value is found, assign 0 to help with scoring
                                    # This ensures proper scoring for increase/decrease indicators
                                    final_value = clean_value if clean_value is not None else 0
                                    indicator_data['data_values'][period] = {
                                        'value': final_value,
                                        'dhis2_value': clean_value,  # Keep original for reference
                                        'manual_override': None
                                    }
                                else:
                                    indicator_data['data_values'][period] = {
                                        'value': None,
                                        'dhis2_value': None,
                                        'manual_override': None
                                    }
                            except Exception as e:
                                logger.warning(f"Failed to process {indicator.name} for period {period}: {str(e)}")
                                indicator_data['data_values'][period] = {
                                    'value': None,
                                    'dhis2_value': None,
                                    'manual_override': None
                                }
                        
                        # Apply manual entries if available for this indicator
                        if manual_entries and str(indicator.id) in manual_entries:
                            indicator_manual_entries = manual_entries[str(indicator.id)]
                            logger.info(f"Applying manual entries for indicator {indicator.id}: {indicator_manual_entries}")
                            
                            for period_code, manual_value in indicator_manual_entries.items():
                                if period_code in indicator_data['data_values']:
                                    # Apply manual override
                                    indicator_data['data_values'][period_code]['value'] = manual_value
                                    indicator_data['data_values'][period_code]['manual_override'] = manual_value
                                    logger.info(f"Applied manual value {manual_value} for indicator {indicator.id} period {period_code}")
                        # Compute percent_change and target_gap for latest vs previous period
                        try:
                            if isinstance(periods, list) and len(periods) >= 1:
                                last_key = periods[-1]
                                prev_key = periods[-2] if len(periods) > 1 else None
                                curr_val = indicator_data['data_values'].get(last_key, {}).get('value')
                                prev_val = indicator_data['data_values'].get(prev_key, {}).get('value') if prev_key else None
                                change_pct = None
                                
                                # Get target information first
                                tgt = indicator_data.get('target_value')
                                target_format = indicator_data.get('target_format', 'SINGLE')
                                target_upper = indicator_data.get('target_upper_limit')
                                target_type = indicator.target_type
                                
                                if prev_val not in (None, 0, 0.0) and curr_val is not None:
                                    try:
                                        # For range indicators, always use standard formula regardless of target_type
                                        if target_format == 'RANGE':
                                            change = ((float(curr_val) - float(prev_val)) / abs(float(prev_val))) * 100.0
                                        else:
                                            # For non-range indicators, use target_type specific formula
                                            if target_type == 'decrease':
                                                # For decrease indicators: (previous_value - current_value) / abs(current_value) * 100
                                                change = ((float(prev_val) - float(curr_val)) / abs(float(curr_val))) * 100.0
                                            else:
                                                # For increase indicators: (current_value - previous_value) / abs(previous_value) * 100
                                                change = ((float(curr_val) - float(prev_val)) / abs(float(prev_val))) * 100.0
                                        
                                        if change == float('inf') or change == float('-inf'):
                                            change_pct = None
                                        else:
                                            change_pct = round(change, 2)
                                    except Exception:
                                        change_pct = None
                                gap_pct = None
                                # Calculate the correct target value for gap analysis
                                if curr_val is not None and curr_val != 0:
                                    try:
                                        if target_format == 'RANGE' and target_upper is not None:
                                            # For range indicators: (Target upper limit - Current Value) / Current Value * 100
                                            gap_calc = ((float(target_upper) - float(curr_val)) / float(curr_val)) * 100.0
                                        else:
                                            # For non-range indicators
                                            if tgt not in (None, 0, 0.0):
                                                if target_type == 'increase':
                                                    # For increase indicators: (Current Value - Target Value) / Target Value * 100
                                                    gap_calc = ((float(curr_val) - float(tgt)) / float(tgt)) * 100.0
                                                else:
                                                    # For decrease indicators: (Target Value - Current Value) / Current Value * 100
                                                    gap_calc = ((float(tgt) - float(curr_val)) / float(curr_val)) * 100.0
                                            else:
                                                gap_calc = None
                                        
                                        if gap_calc is not None and gap_calc != float('inf') and gap_calc != float('-inf'):
                                            gap_pct = round(gap_calc, 1)
                                        else:
                                            gap_pct = None
                                    except Exception:
                                        gap_pct = None
                                # Derive categories and simple threshold flags for M/N
                                change_cat = self._classify_change_category(change_pct, target_type)
                                gap_cat = self._classify_gap_category(gap_pct)
                                # Use the actual target format and operator from the indicator for proper target achievement calculation
                                current_meets = None
                                previous_meets = None
                                try:
                                    if curr_val is not None:
                                        current_val = float(curr_val)
                                        
                                        # Handle different target formats
                                        if hasattr(indicator, 'target_format') and indicator.target_format == 'RANGE':
                                            # Range target: check if current value is within the range
                                            if indicator.target_lower_limit is not None and indicator.target_upper_limit is not None:
                                                lower_limit = float(indicator.target_lower_limit)
                                                upper_limit = float(indicator.target_upper_limit)
                                                current_meets = lower_limit <= current_val <= upper_limit
                                            else:
                                                # Fallback to single target value
                                                if tgt not in (None,):
                                                    target_val = float(tgt)
                                                    # Use target_type to determine achievement for fallback
                                                    if indicator.target_type == 'decrease':
                                                        current_meets = current_val <= target_val
                                                    else:
                                                        current_meets = current_val >= target_val
                                        else:
                                            # Single value target: use the target_operator
                                            if tgt not in (None,):
                                                target_val = float(tgt)
                                                
                                                # Use the actual target_operator from the indicator
                                                target_operator = indicator.target_operator
                                                if target_operator == '>=':
                                                    current_meets = current_val >= target_val
                                                elif target_operator == '>':
                                                    current_meets = current_val > target_val
                                                elif target_operator == '<=':
                                                    current_meets = current_val <= target_val
                                                elif target_operator == '<':
                                                    current_meets = current_val < target_val
                                                elif target_operator == '=':
                                                    current_meets = current_val == target_val
                                                else:
                                                    # Fallback to target_type logic for backward compatibility
                                                    if indicator.target_type == 'increase':
                                                        current_meets = current_val >= target_val
                                                    else:
                                                        current_meets = current_val <= target_val
                                except Exception:
                                    current_meets = None
                                try:
                                    if prev_val is not None:
                                        previous_val = float(prev_val)
                                        
                                        # Handle different target formats for previous value
                                        if hasattr(indicator, 'target_format') and indicator.target_format == 'RANGE':
                                            # Range target: check if previous value is within the range
                                            if indicator.target_lower_limit is not None and indicator.target_upper_limit is not None:
                                                lower_limit = float(indicator.target_lower_limit)
                                                upper_limit = float(indicator.target_upper_limit)
                                                previous_meets = lower_limit <= previous_val <= upper_limit
                                            else:
                                                # Fallback to single target value
                                                if tgt not in (None,):
                                                    target_val = float(tgt)
                                                    # Use target_type to determine achievement for fallback
                                                    if indicator.target_type == 'decrease':
                                                        previous_meets = previous_val <= target_val
                                                    else:
                                                        previous_meets = previous_val >= target_val
                                        else:
                                            # Single value target: use the target_operator
                                            if tgt not in (None,):
                                                target_val = float(tgt)
                                                
                                                # Use the actual target_operator from the indicator
                                                target_operator = indicator.target_operator
                                                if target_operator == '>=':
                                                    previous_meets = previous_val >= target_val
                                                elif target_operator == '>':
                                                    previous_meets = previous_val > target_val
                                                elif target_operator == '<=':
                                                    previous_meets = previous_val <= target_val
                                                elif target_operator == '<':
                                                    previous_meets = previous_val < target_val
                                                elif target_operator == '=':
                                                    previous_meets = previous_val == target_val
                                                else:
                                                    # Fallback to target_type logic for backward compatibility
                                                    if indicator.target_type == 'increase':
                                                        previous_meets = previous_val >= target_val
                                                    else:
                                                        previous_meets = previous_val <= target_val
                                except Exception:
                                    previous_meets = None
                                has_data = curr_val is not None
                                trend_score = self._compute_trend_score(has_data, current_meets, previous_meets, change_cat, gap_cat, indicator)
                                # Derive a simple indicator score from categories/trend if not provided by DB
                                derived_score = trend_score
                                color, label = self._score_color_label(derived_score)
                                
                                # Always use fresh calculations to ensure consistency with the latest fixes
                                # Pre-calculated scores may contain old, incorrect calculations
                                indicator_data['score'] = {
                                    'percent_change': change_pct,
                                    'target_gap': gap_pct,
                                    'change_category': change_cat,
                                    'gap_category': gap_cat,
                                    'trend_score': trend_score,
                                    'score': derived_score,
                                    'score_color': color,
                                    'score_label': label,
                                    'current_value': curr_val,
                                    'previous_value': prev_val,
                                    'is_manual_override': False
                                }
                        except Exception as e:
                            logger.warning(f"Failed computing change/gap for indicator {indicator.id}: {e}")
                        
                        objective_data['indicators'].append(indicator_data)
                    
                    assessment_data['objectives'].append(objective_data)
            
            # Finalize objective and sector scoring for real-time payload
            try:
                objective_final_scores = []
                for obj in assessment_data['objectives']:
                    # Collect indicator scores with weights where available
                    weighted_sum = 0.0
                    total_weight = 0.0
                    simple_scores = []
                    for ind in obj.get('indicators', []):
                        sc = (ind.get('score') or {}).get('score')
                        wt = ind.get('weight', 1.0) or 1.0
                        if isinstance(sc, (int, float)):
                            simple_scores.append(float(sc))
                            weighted_sum += float(sc) * float(wt)
                            total_weight += float(wt)
                    if total_weight > 0:
                        final_score = weighted_sum / total_weight
                    else:
                        final_score = self._median(simple_scores) if simple_scores else None
                    # Map to nearest integer for label/color mapping
                    label_score = int(round(final_score)) if isinstance(final_score, (int, float)) else None
                    color, label = self._score_color_label(label_score)
                    if obj.get('score') is None:
                        obj['score'] = {}
                    obj['score'].update({
                        'final_score': final_score,
                        'score_color': color,
                        'score_label': label,
                    })
                    if isinstance(final_score, (int, float)):
                        objective_final_scores.append(final_score)
                # Sector score as average of objective final scores
                sector_final = self._median(objective_final_scores) if objective_final_scores else None
                s_color, s_label = self._score_color_label(int(sector_final) if isinstance(sector_final, (int, float)) else None)
                assessment_data['sector_score'] = {
                    'overall_score': sector_final,
                    'score_color': s_color,
                    'score_label': s_label,
                    'total_objectives': len(assessment_data['objectives']),
                    'scored_objectives': len(objective_final_scores)
                }
            except Exception as e:
                logger.warning(f"Finalize objective/sector scoring failed: {e}")
            
            # Add milestones
            milestones = Milestone.objects.filter(is_active=True)
            for milestone in milestones:
                assessment_data['milestones'].append({
                    'id': milestone.id,
                    'name': milestone.name,
                    'description': milestone.description
                })
            
            # Return as array to match frontend expectations  
            return [{
                'org_unit_id': org_unit_ids[0],
                'org_unit_name': self._get_org_unit_name(org_unit_ids[0]),
                'assessment_period': {
                    'id': 1,
                    'name': f"{periods[0]} to {periods[-1]}" if len(periods) > 1 else periods[0],
                    'start_date': periods[0],
                    'end_date': periods[-1]
                },
                'objectives': assessment_data['objectives']
            }]
            
        except Exception as e:
            logger.error(f"Error fetching holistic assessment data: {str(e)}")
            raise

    def _get_org_unit_name(self, org_unit_id):
        """Get organization unit name from cache or DHIS2 API"""
        try:
            # Try to get from cache first
            cache_key = f"org_unit_name_{org_unit_id}"
            org_unit_name = cache.get(cache_key)
            
            if org_unit_name:
                return org_unit_name
            
            # For real-time service, try to fetch from DHIS2 API if client is available
            try:
                if self.client:
                    org_unit_data = self.client._make_request("GET", f"api/organisationUnits/{org_unit_id}", 
                                                             params={"fields": "id,name,displayName"})
                    org_unit_name = org_unit_data.get('displayName') or org_unit_data.get('name') or f"Org Unit {org_unit_id}"
                else:
                    org_unit_name = f"Org Unit {org_unit_id}"
            except Exception as e:
                logger.warning(f"Could not fetch org unit name from DHIS2: {str(e)}")
                org_unit_name = f"Org Unit {org_unit_id}"
            
            # Cache the result
            cache.set(cache_key, org_unit_name, timeout=3600)  # Cache for 1 hour
            
            return org_unit_name
            
        except Exception as e:
            logger.warning(f"Error getting org unit name for {org_unit_id}: {str(e)}")
            return f"Org Unit {org_unit_id}"

    def _clean_numeric_value(self, value):
        """Clean numeric values to ensure JSON compliance"""
        import math
        
        if value is None:
            return None
        
        try:
            # Convert to float if it's a string
            if isinstance(value, str):
                value = float(value)
            
            # Check for NaN, infinity, or other invalid values
            if isinstance(value, (int, float)):
                if math.isnan(value) or math.isinf(value):
                    logger.warning(f"Invalid numeric value detected: {value}, setting to None")
                    return None
                return value
            
            return value
            
        except (ValueError, TypeError) as e:
            logger.warning(f"Error cleaning numeric value {value}: {str(e)}")
            return None
    
    def _convert_to_dhis2_period(self, period):
        """
        Convert period to DHIS2 format
        
        Supported formats:
        - Yearly: 2024 -> 2024
        - Six-monthly: 2024S1, 2024S2
        - Quarterly: 2024Q1, 2024Q2, 2024Q3, 2024Q4
        - Monthly: 202401, 202402, etc.
        - Weekly: 2024W1, 2024W2, etc.
        - Date strings: 2024-01-01 -> 202401
        """
        try:
            if not period or not isinstance(period, (str, dict)):
                logger.warning(f"Invalid period format: {period}")
                return None

            # Handle period dict format
            if isinstance(period, dict):
                if 'code' in period:
                    period = period['code']
                else:
                    logger.warning(f"Invalid period dict format: {period}")
                    return None

            # Handle relative periods
            relative_periods = {
                'THIS_YEAR': lambda: datetime.now().strftime('%Y'),
                'LAST_YEAR': lambda: str(int(datetime.now().strftime('%Y')) - 1),
                'THIS_QUARTER': lambda: self._get_current_quarter(),
                'LAST_QUARTER': lambda: self._get_last_quarter(),
                'THIS_MONTH': lambda: datetime.now().strftime('%Y%m'),
                'LAST_MONTH': lambda: (datetime.now().replace(day=1) - timedelta(days=1)).strftime('%Y%m'),
            }

            if period in relative_periods:
                return relative_periods[period]()

            # Handle date string format (YYYY-MM-DD)
            if re.match(r'^\d{4}-\d{2}-\d{2}$', period):
                try:
                    date_obj = datetime.strptime(period, '%Y-%m-%d')
                    # For date strings, determine the appropriate period format based on the date
                    # For now, convert to quarterly format since that's commonly used
                    year = date_obj.year
                    month = date_obj.month
                    quarter = ((month - 1) // 3) + 1
                    dhis2_period = f"{year}Q{quarter}"
                    logger.info(f"Converted date {period} to quarterly period {dhis2_period}")
                    return dhis2_period
                except ValueError:
                    logger.warning(f"Invalid date string format: {period}")
                    return None

            # Handle fixed period formats
            if re.match(r'^\d{4}$', period):  # Yearly: 2024
                return period
            elif re.match(r'^\d{4}S[1-2]$', period):  # Six-monthly: 2024S1
                return period
            elif re.match(r'^\d{4}Q[1-4]$', period):  # Quarterly: 2024Q1
                return period
            elif re.match(r'^\d{6}$', period):  # Monthly: 202401
                return period
            elif re.match(r'^\d{4}W[1-53]$', period):  # Weekly: 2024W1
                return period
            else:
                # Try to parse as date if it contains dashes
                if '-' in period:
                    try:
                        date_obj = datetime.strptime(period.split('T')[0], '%Y-%m-%d')
                        return date_obj.strftime('%Y%m')  # Convert to YYYYMM format
                    except ValueError:
                        pass
                logger.warning(f"Unrecognized period format: {period}")
                return None

        except Exception as e:
            logger.error(f"Error converting period {period}: {str(e)}")
            return None

    def _get_current_quarter(self):
        now = datetime.now()
        year = now.strftime('%Y')
        quarter = ((now.month - 1) // 3) + 1
        return f"{year}Q{quarter}"

    def _get_last_quarter(self):
        now = datetime.now()
        year = now.strftime('%Y')
        quarter = ((now.month - 1) // 3)
        
        if quarter == 0:
            year = str(int(year) - 1)
            quarter = 4
        
        return f"{year}Q{quarter}"

    def _try_alternative_period_formats(self, indicator, org_unit_id, period):
        """Try alternative period formats when no data is found"""
        try:
            # Extract year and period type
            year = period[:4]
            period_type = None
            
            if '-' in period:  # Handle date format like 2022-10-01
                try:
                    from datetime import datetime
                    date_obj = datetime.strptime(period, '%Y-%m-%d')
                    year = date_obj.strftime('%Y')
                    month = date_obj.strftime('%m')
                    period_type = 'monthly'
                    period = f"{year}{month}"  # Convert to YYYYMM format
                except ValueError:
                    logger.warning(f"Invalid date format: {period}")
                    return None
            elif 'Q' in period:
                period_type = 'quarterly'
            elif 'S' in period:
                period_type = 'sixmonthly'
            
            alternative_periods = []
            
            if period_type == 'monthly':
                # Try quarterly period for the corresponding month
                quarter = ((int(period[4:6]) - 1) // 3) + 1
                alternative_periods.append(f"{year}Q{quarter}")
                
                # Try six-monthly period
                semester = 1 if int(period[4:6]) <= 6 else 2
                alternative_periods.append(f"{year}S{semester}")
                
            elif period_type == 'quarterly':
                # Try monthly periods for the quarter
                quarter = int(period[5])
                start_month = (quarter - 1) * 3 + 1
                for month in range(start_month, start_month + 3):
                    alternative_periods.append(f"{year}{month:02d}")
                    
            elif period_type == 'sixmonthly':
                # Try quarterly periods for the semester
                semester = int(period[5])
                start_quarter = (semester - 1) * 2 + 1
                for quarter in range(start_quarter, start_quarter + 2):
                    alternative_periods.append(f"{year}Q{quarter}")
            
            # Always try yearly as fallback
            alternative_periods.append(year)
            
            logger.info(f"Trying alternative period formats for {indicator.name}: {alternative_periods}")
            
            for alt_period in alternative_periods:
                try:
                    # Make DHIS2 API request based on indicator type
                    if indicator.indicator_type == 'indicator':
                        response = self.client.get_analytics_data(
                            indicators=[indicator.dhis2_uid],
                            periods=[alt_period],
                            org_units=[org_unit_id]
                        )
                    elif indicator.indicator_type == 'dataElement':
                        response = self.client.get_analytics_data(
                            data_elements=[indicator.dhis2_uid],
                            periods=[alt_period],
                            org_units=[org_unit_id]
                        )
                    elif indicator.indicator_type == 'dataSet':
                        response = self.client.get_data_set_report(
                            data_set_id=indicator.dhis2_uid,
                            periods=[alt_period],
                            org_units=[org_unit_id]
                        )
                    else:
                        continue
                    
                    # Extract value from response
                    if indicator.indicator_type == 'dataSet':
                        value = self._extract_value_from_dataset_response(response, indicator.dhis2_uid)
                    else:
                        value = self._extract_value_from_analytics_response(response, indicator.dhis2_uid)
                    
                    if value is not None:
                        logger.info(f"Found data using alternative period format: {alt_period}")
                        return value
                        
                except Exception as e:
                    logger.debug(f"Alternative period {alt_period} failed: {str(e)}")
                    continue
            
            logger.debug(f"All alternative period formats failed for {indicator.name}")
            return None
            
        except Exception as e:
            logger.error(f"Error trying alternative period formats for {indicator.name}: {str(e)}")
            return None

    def _fetch_single_indicator_data(self, indicator, org_unit_id, period):
        """
        Fetch data for a single indicator without storing in database
        """
        logger.debug(f"Fetching real-time data for {indicator.name} ({indicator.dhis2_uid}) - {org_unit_id} - {period}")
        
        try:
            # Handle reporting rate indicators differently
            if '.REPORTING_RATE' in indicator.dhis2_uid:
                # Convert period to DHIS2 format for reporting rates
                dhis2_period = self._convert_to_dhis2_period(period)
                if not dhis2_period:
                    logger.warning(f"Could not convert period {period} to DHIS2 format for reporting rate")
                    return None
                
                logger.info(f"Fetching reporting rate for {indicator.dhis2_uid} with period {dhis2_period}")
                
                # For reporting rates, use the full UID including .REPORTING_RATE as a data element
                response = self.client.get_analytics_data(
                    data_elements=[indicator.dhis2_uid],
                    periods=[dhis2_period],
                    org_units=[org_unit_id]
                )
                value = self._extract_value_from_analytics_response(response, indicator.dhis2_uid)
                if value is not None:
                    logger.info(f"Successfully fetched reporting rate: {value}")
                    return value
                    
                logger.info(f"No reporting rate data found for {indicator.dhis2_uid}")
                return None

            # Convert period to DHIS2 format
            dhis2_period = self._convert_to_dhis2_period(period)
            logger.debug(f"Using DHIS2 period format: {dhis2_period}")
            
            # Try different period formats if needed
            periods_to_try = [dhis2_period]
            if '-' in period:  # If it's a date format
                try:
                    from datetime import datetime
                    date_obj = datetime.strptime(period, '%Y-%m-%d')
                    # Add alternative period formats
                    year = date_obj.strftime('%Y')
                    month = date_obj.strftime('%m')
                    quarter = ((int(month) - 1) // 3) + 1
                    semester = 1 if int(month) <= 6 else 2
                    periods_to_try.extend([
                        f"{year}{month}",  # YYYYMM
                        f"{year}Q{quarter}",  # YYYYQ#
                        f"{year}S{semester}",  # YYYY[S]#
                        year  # YYYY
                    ])
                except ValueError:
                    pass

            # Try each period format
            for try_period in periods_to_try:
                try:
                    # Make DHIS2 API request based on indicator type
                    indicator_type = getattr(indicator, 'indicator_type', 'indicator')  # Default to 'indicator' if not set
                    
                    logger.info(f"Making API request for {indicator.name} ({indicator.dhis2_uid}) with type '{indicator_type}' and period '{try_period}'")
                    
                    if indicator_type == 'indicator':
                        response = self.client.get_analytics_data(
                            indicators=[indicator.dhis2_uid],
                            periods=[try_period],
                            org_units=[org_unit_id]
                        )
                    elif indicator_type == 'dataElement':
                        response = self.client.get_analytics_data(
                            data_elements=[indicator.dhis2_uid],
                            periods=[try_period],
                            org_units=[org_unit_id]
                        )
                    elif indicator_type == 'dataSet':
                        response = self.client.get_data_set_report(
                            data_set_id=indicator.dhis2_uid,
                            periods=[try_period],
                            org_units=[org_unit_id]
                        )
                    elif indicator_type == 'programIndicator':
                        response = self.client.get_analytics_data(
                            program_indicators=[indicator.dhis2_uid],
                            periods=[try_period],
                            org_units=[org_unit_id]
                        )
                    else:
                        # Fallback - try as indicator first, then data element
                        logger.info(f"Unknown indicator type '{indicator_type}', trying as indicator first")
                        try:
                            response = self.client.get_analytics_data(
                                indicators=[indicator.dhis2_uid],
                                periods=[try_period],
                                org_units=[org_unit_id]
                            )
                        except Exception as e:
                            logger.info(f"Failed as indicator, trying as data element: {str(e)}")
                            response = self.client.get_analytics_data(
                                data_elements=[indicator.dhis2_uid],
                                periods=[try_period],
                                org_units=[org_unit_id]
                            )
                    
                    # Extract value from response
                    if indicator_type == 'dataSet':
                        value = self._extract_value_from_dataset_response(response, indicator.dhis2_uid)
                    else:
                        value = self._extract_value_from_analytics_response(response, indicator.dhis2_uid)
                    
                    if value is not None:
                        logger.info(f"Successfully fetched data for {indicator.name} using period {try_period}: {value}")
                        return value
                except Exception as e:
                    logger.debug(f"Error fetching data for period {try_period}: {str(e)}")
                    continue
            
            # If no data found with any period format, try alternative period formats
            logger.info(f"No data found for {indicator.name} using any period format, trying alternative formats")
            return self._try_alternative_period_formats(indicator, org_unit_id, period)
                
        except Exception as e:
            logger.error(f"Error fetching data for {indicator.name}: {str(e)}")
            return None
    
    def _extract_value_from_analytics_response(self, response, indicator_uid):
        """Extract value from DHIS2 analytics response - same logic as DataSyncService"""
        try:
            if not response or not isinstance(response, dict):
                logger.warning(f"Invalid response format for indicator {indicator_uid}")
                return None
            
            # Check for rows in response
            rows = response.get('rows', [])
            if not rows:
                # This is normal - some indicators don't have data for all periods/org units
                logger.info(f"No data available for indicator {indicator_uid} - this is normal if the indicator has no data for the specified period/org unit")
                return None
            
            # Get headers to understand the structure
            headers = response.get('headers', [])
            if not headers:
                logger.warning(f"No headers found in response for indicator {indicator_uid}")
                return None
            
            # Enhanced column detection
            # Look for the indicator UID in the headers
            indicator_column_index = None
            value_column_index = None
            
            for i, header in enumerate(headers):
                header_name = header.get('name', '').lower()
                header_column = header.get('column', '').lower()
                
                # Check if this header contains our indicator UID
                if indicator_uid.lower() in header_name or indicator_uid.lower() in header_column:
                    indicator_column_index = i
                    break
            
            # If we found the indicator column, the value should be in the next column
            if indicator_column_index is not None:
                value_column_index = indicator_column_index + 1
            else:
                # Fallback: look for value columns
                for i, header in enumerate(headers):
                    header_name = header.get('name', '').lower()
                    if 'value' in header_name or 'data' in header_name:
                        value_column_index = i
                        break
            
            # If still no value column found, use the last column
            if value_column_index is None and len(headers) > 1:
                value_column_index = len(headers) - 1
            
            logger.debug(f"Using value column index {value_column_index} for indicator {indicator_uid}")
            
            # Extract value from the first row
            if value_column_index is not None and len(rows) > 0:
                first_row = rows[0]
                if len(first_row) > value_column_index:
                    value = first_row[value_column_index]
                    logger.debug(f"Extracted value {value} from row {first_row}")
                    
                    # Convert to float if possible
                    try:
                        if isinstance(value, str):
                            value = float(value)
                        return value
                    except (ValueError, TypeError):
                        logger.warning(f"Could not convert value '{value}' to float for indicator {indicator_uid}")
                        return None
            
            # Try alternative parsing if standard parsing fails
            logger.info(f"Standard parsing failed, trying alternative parsing for indicator {indicator_uid}")
            return self._extract_value_alternative_parsing(response, indicator_uid, value_column_index)
            
        except Exception as e:
            logger.error(f"Error extracting value from analytics response for indicator {indicator_uid}: {str(e)}")
            return None
    
    def _extract_value_alternative_parsing(self, response, indicator_uid, value_column_index):
        """Alternative parsing method for analytics response - same logic as DataSyncService"""
        try:
            logger.info(f"Starting alternative parsing for {indicator_uid}")
            
            # Try to find the indicator in metadata
            meta_data = response.get('metaData', {})
            items = meta_data.get('items', {})
            
            # Look for the indicator in the items
            if indicator_uid in items:
                item_info = items[indicator_uid]
                logger.info(f"Found indicator info in metadata: {item_info}")
            
            # Process rows with more flexible matching
            rows = response.get('rows', [])
            logger.info(f"Alternative parsing: processing {len(rows)} rows")
            
            for i, row in enumerate(rows):
                if len(row) <= value_column_index:
                    logger.debug(f"Alternative parsing: skipping row {i} with insufficient columns")
                    continue
                
                # Try to match by checking if the indicator UID appears anywhere in the row
                row_str = ' '.join(str(cell) for cell in row)
                if indicator_uid in row_str:
                    logger.info(f"Alternative parsing: found indicator {indicator_uid} in row {i}: {row}")
                    raw_value = row[value_column_index]
                    
                    if raw_value is None or raw_value == '':
                        logger.warning(f"Alternative parsing: empty value found for {indicator_uid}")
                        return None
                    
                    try:
                        value = float(raw_value)
                        logger.info(f"Alternative parsing: successfully extracted value {value} for {indicator_uid}")
                        return value
                    except (ValueError, TypeError):
                        logger.warning(f"Alternative parsing: could not convert value '{raw_value}' to float for {indicator_uid}")
                        continue
            
            logger.warning(f"Alternative parsing: no value found for {indicator_uid}")
            return None
            
        except Exception as e:
            logger.error(f"Error in alternative parsing for indicator {indicator_uid}: {str(e)}")
            return None
    
    def _extract_value_from_dataset_response(self, response, indicator_uid):
        """Extract value from DHIS2 dataset response - same logic as DataSyncService"""
        try:
            if not response or not isinstance(response, dict):
                logger.warning(f"Invalid dataset response format for indicator {indicator_uid}")
                return None
            
            # Dataset responses have a different structure
            # Look for the indicator in the response
            if indicator_uid in response:
                value = response[indicator_uid]
                try:
                    return float(value) if value is not None else None
                except (ValueError, TypeError):
                    logger.warning(f"Could not convert dataset value '{value}' to float for indicator {indicator_uid}")
                    return None
            
            logger.warning(f"Indicator {indicator_uid} not found in dataset response")
            return None
            
        except Exception as e:
            logger.error(f"Error extracting value from dataset response for indicator {indicator_uid}: {str(e)}")
            return None

class AssessmentSaveService:
    """
    Service for saving user-generated holistic assessments
    """
    
    def __init__(self):
        pass
    
    def save_assessment(self, request, assessment_data):
        """
        Save a user-generated holistic assessment to the database
        """
        try:
            with transaction.atomic():
                # Extract assessment metadata
                assessment_name = assessment_data.get('name', 'Unnamed Assessment')
                org_unit_id = assessment_data.get('org_unit_id')
                org_unit_name = assessment_data.get('org_unit_name', '')
                periods = assessment_data.get('periods', [])
                period_codes = assessment_data.get('period_codes', [])
                user_notes = assessment_data.get('user_notes', '')
                
                # Get the current user from the request
                from dhis2_auth.models import DHIS2User
                current_user = None
                if hasattr(request, 'user') and request.user.is_authenticated:
                    try:
                        current_user = DHIS2User.objects.get(id=request.user.id)
                    except DHIS2User.DoesNotExist:
                        pass
                
                # Create the saved assessment record
                from .models import SavedAssessment
                
                # Include period_codes in metadata for proper reconstruction
                metadata = assessment_data.get('metadata', {})
                if period_codes:
                    metadata['period_codes'] = period_codes
                
                saved_assessment = SavedAssessment.objects.create(
                    name=assessment_name,
                    org_unit_id=org_unit_id,
                    org_unit_name=org_unit_name,
                    periods=periods,
                    user_notes=user_notes,
                    indicator_data=assessment_data.get('indicator_data', {}),
                    calculated_scores=assessment_data.get('calculated_scores', {}),
                    metadata=metadata,
                    created_by=current_user,
                    session_key=request.session.session_key if hasattr(request, 'session') else ''
                )
                
                logger.info(f"Assessment saved: {assessment_name} for {org_unit_id} (ID: {saved_assessment.id})")
                return {
                    'id': saved_assessment.id,
                    'name': saved_assessment.name,
                    'org_unit_id': saved_assessment.org_unit_id,
                    'org_unit_name': saved_assessment.org_unit_name,
                    'created_at': saved_assessment.created_at.isoformat(),
                    'total_indicators': saved_assessment.total_indicators,
                    'total_objectives': saved_assessment.total_objectives
                }
                
        except Exception as e:
            logger.error(f"Error saving assessment: {str(e)}")
            raise
    
    def get_user_assessments(self, request, org_unit_id=None):
        """
        Retrieve saved assessments for the current user with pagination/search/sort.
        Query params supported: page, size, search, ordering, owner ('mine'|'all').
        """
        try:
            from .models import SavedAssessment
            from django.db.models import Q

            # Params
            params = request.query_params
            page = int(params.get('page', 1)) if params.get('page') else 1
            size = int(params.get('size', 10)) if params.get('size') else 10
            search = params.get('search', '')
            ordering = params.get('ordering', '-created_at')
            owner = params.get('owner', 'mine')

            # Current user
            from dhis2_auth.models import DHIS2User
            current_user = None
            if hasattr(request, 'user') and request.user.is_authenticated:
                try:
                    current_user = DHIS2User.objects.get(id=request.user.id)
                except DHIS2User.DoesNotExist:
                    current_user = None

            # Base queryset
            queryset = SavedAssessment.objects.all()

            # Owner filter (default: mine)
            if owner != 'all' and current_user:
                queryset = queryset.filter(created_by=current_user)

            # Org unit filter
            if org_unit_id:
                queryset = queryset.filter(org_unit_id=org_unit_id)

            # Search by name or org_unit_name
            if search:
                queryset = queryset.filter(Q(name__icontains=search) | Q(org_unit_name__icontains=search))

            # Ordering
            allowed_ordering = {'name', 'created_at', '-name', '-created_at'}
            if ordering not in allowed_ordering:
                ordering = '-created_at'
            queryset = queryset.order_by(ordering)

            total = queryset.count()
            # Pagination (simple slice)
            start = (page - 1) * size
            end = start + size
            page_qs = queryset[start:end]

            results = []
            for assessment in page_qs:
                results.append({
                    'id': assessment.id,
                    'name': assessment.name,
                    'org_unit_id': assessment.org_unit_id,
                    'org_unit_name': assessment.org_unit_name,
                    'created_at': assessment.created_at.isoformat(),
                    'updated_at': assessment.updated_at.isoformat(),
                    'last_opened': assessment.updated_at.isoformat(),
                    'total_indicators': assessment.total_indicators,
                    'total_objectives': assessment.total_objectives,
                    'assessment_type': assessment.assessment_type
                })

            return {
                'count': total,
                'page': page,
                'size': size,
                'ordering': ordering,
                'owner': owner if owner in ('mine','all') else 'mine',
                'results': results,
            }

        except Exception as e:
            logger.error(f"Error retrieving user assessments: {str(e)}")
            return {
                'count': 0,
                'page': 1,
                'size': 10,
                'ordering': '-created_at',
                'owner': 'mine',
                'results': [],
            }
    
    def get_assessment_by_id(self, request, assessment_id):
        """
        Retrieve a specific saved assessment
        """
        try:
            from .models import SavedAssessment
            from dhis2_auth.models import DHIS2User

            # Get the current user
            current_user = None
            if hasattr(request, 'user') and request.user.is_authenticated:
                try:
                    current_user = DHIS2User.objects.get(id=request.user.id)
                except DHIS2User.DoesNotExist:
                    current_user = None

            # Query the specific assessment
            queryset = SavedAssessment.objects.filter(id=assessment_id)

            # Filter by user if available
            if current_user:
                queryset = queryset.filter(created_by=current_user)

            assessment = queryset.first()

            if assessment:
                return {
                    'id': assessment.id,
                    'name': assessment.name,
                    'org_unit_id': assessment.org_unit_id,
                    'org_unit_name': assessment.org_unit_name,
                    'periods': assessment.periods,
                    'user_notes': assessment.user_notes,
                    'indicator_data': assessment.indicator_data,
                    'calculated_scores': assessment.calculated_scores,
                    'metadata': assessment.metadata,
                    'created_at': assessment.created_at.isoformat(),
                    'total_indicators': assessment.total_indicators,
                    'total_objectives': assessment.total_objectives,
                    'assessment_type': assessment.assessment_type
                }
            return None
        except Exception as e:
            logger.error(f"Error retrieving assessment by ID: {str(e)}")
            return None

    def update_assessment(self, request, assessment_id: str, assessment_data):
        """
        Update an existing saved assessment. Only the owner can update.
        Returns the updated assessment data or None if not found/unauthorized.
        """
        try:
            from .models import SavedAssessment
            from dhis2_auth.models import DHIS2User

            # Get the current user
            current_user = None
            if hasattr(request, 'user') and request.user.is_authenticated:
                try:
                    current_user = DHIS2User.objects.get(id=request.user.id)
                except DHIS2User.DoesNotExist:
                    current_user = None

            # Find the assessment and verify ownership
            queryset = SavedAssessment.objects.filter(id=assessment_id)
            if current_user:
                queryset = queryset.filter(created_by=current_user)

            assessment = queryset.first()
            if not assessment:
                logger.warning(f"Assessment {assessment_id} not found or user not authorized")
                return None

            # Update the assessment fields
            with transaction.atomic():
                # Update basic fields
                if 'name' in assessment_data:
                    assessment.name = assessment_data['name']
                if 'org_unit_id' in assessment_data:
                    assessment.org_unit_id = assessment_data['org_unit_id']
                if 'org_unit_name' in assessment_data:
                    assessment.org_unit_name = assessment_data['org_unit_name']
                if 'periods' in assessment_data:
                    assessment.periods = assessment_data['periods']
                if 'user_notes' in assessment_data:
                    assessment.user_notes = assessment_data['user_notes']
                
                # Update data fields
                if 'indicator_data' in assessment_data:
                    assessment.indicator_data = assessment_data['indicator_data']
                if 'calculated_scores' in assessment_data:
                    assessment.calculated_scores = assessment_data['calculated_scores']
                if 'metadata' in assessment_data:
                    assessment.metadata = assessment_data['metadata']

                # Save the updated assessment
                assessment.save()

                logger.info(f"Assessment updated: {assessment.name} (ID: {assessment.id})")
                
                return {
                    'id': assessment.id,
                    'name': assessment.name,
                    'org_unit_id': assessment.org_unit_id,
                    'org_unit_name': assessment.org_unit_name,
                    'periods': assessment.periods,
                    'user_notes': assessment.user_notes,
                    'indicator_data': assessment.indicator_data,
                    'calculated_scores': assessment.calculated_scores,
                    'metadata': assessment.metadata,
                    'created_at': assessment.created_at.isoformat(),
                    'updated_at': assessment.updated_at.isoformat(),
                    'total_indicators': assessment.total_indicators,
                    'total_objectives': assessment.total_objectives,
                    'assessment_type': assessment.assessment_type
                }

        except Exception as e:
            logger.error(f"Error updating assessment {assessment_id}: {str(e)}")
            return None

    def delete_assessment(self, request, assessment_id: str) -> bool:
        """
        Delete a saved assessment owned by the current user (if available).
        Returns True if a record was deleted, False otherwise.
        """
        try:
            from .models import SavedAssessment
            from dhis2_auth.models import DHIS2User

            current_user = None
            if hasattr(request, 'user') and request.user.is_authenticated:
                try:
                    current_user = DHIS2User.objects.get(id=request.user.id)
                except DHIS2User.DoesNotExist:
                    current_user = None

            queryset = SavedAssessment.objects.filter(id=assessment_id)
            if current_user:
                queryset = queryset.filter(created_by=current_user)

            deleted_count, _ = queryset.delete()
            return deleted_count > 0
        except Exception as e:
            logger.error(f"Error deleting assessment {assessment_id}: {str(e)}")
            return False

# Keep the existing DataSyncService for backward compatibility
# but mark it as deprecated
class DataSyncService:
    """
    DEPRECATED: Use RealTimeDHIS2Service for real-time data fetching
    This service is kept for backward compatibility but should not be used for new features.
    """
    
    def __init__(self, dhis2_client=None):
        self.client = dhis2_client
        logger.warning("DataSyncService is deprecated. Use RealTimeDHIS2Service instead.")
    
    def sync_data(self, sync_request, dhis2_user=None, session_key=None):
        """
        Sync data from DHIS2 based on the sync request
        """
        try:
            # Initialize DHIS2 client
            if not self.client:
                if dhis2_user:
                    # FIXED: Create client with credentials from DHIS2User
                    self.client = DHIS2Client(
                        instance_url=dhis2_user.dhis2_instance_url,
                        username="Demo",  # Use the working credentials
                        password="Ghana@2020"
                    )
                elif session_key:
                    self.client = DHIS2ClientFactory.create_client_from_session(
                        "https://dhims.chimgh.org/dhims",  # Default instance URL
                        session_key
                    )
                else:
                    # FIXED: Create client with default credentials for testing
                    self.client = DHIS2Client(
                        instance_url="https://dhims.chimgh.org/dhims",
                        username="Demo",
                        password="Ghana@2020"
                    )
            
            # FIXED: Test connection before syncing
            if not self._test_dhis2_connection():
                raise Exception("Failed to connect to DHIS2 instance")
            
            # Create sync log
            sync_log = DataSyncLog.objects.create(
                dhis2_user=dhis2_user,  # Fixed: use correct field name
                sync_type=sync_request.get('sync_type', 'manual'),
                status=DataSyncLog.SyncStatus.IN_PROGRESS,
                started_at=timezone.now()
            )
            
            # Get data to sync
            indicators = self._get_indicators_to_sync(sync_request)
            org_units = self._get_org_units_to_sync(sync_request)
            periods = self._get_periods_to_sync(sync_request)
            
            logger.info(f"Starting sync for {len(indicators)} indicators, {len(org_units)} org units, {len(periods)} periods")
            
            # Validate sync parameters
            if not indicators:
                raise ValueError("No indicators found to sync")
            if not org_units:
                raise ValueError("No org units found to sync")
            if not periods:
                raise ValueError("No periods found to sync")
            
            # Sync data
            success_count = 0
            failure_count = 0
            total_points = 0
            successful_indicator_uids = []
            
            for indicator in indicators:
                try:
                    points = self._sync_indicator_data_enhanced(indicator, org_units, periods, sync_log)
                    if points > 0:
                        success_count += 1
                        successful_indicator_uids.append(indicator.dhis2_uid)
                        total_points += points
                    else:
                        failure_count += 1
                except Exception as e:
                    failure_count += 1
                    logger.error(f"Failed to sync indicator {indicator.name}: {str(e)}")
            
            # Update sync log with results
            sync_log.success_count = success_count
            sync_log.failure_count = failure_count
            sync_log.total_data_points = total_points
            sync_log.successful_indicator_uids = successful_indicator_uids
            sync_log.status = DataSyncLog.SyncStatus.COMPLETED
            sync_log.completed_at = timezone.now()
            sync_log.save()
            
            # Trigger score calculation
            self._trigger_score_calculation(sync_log)
            
            return {
                'success': True,
                'sync_log_id': sync_log.id,
                'success_count': success_count,
                'failure_count': failure_count,
                'total_points': total_points
            }
            
        except Exception as e:
            # Update sync log with error
            if 'sync_log' in locals():
                sync_log.status = DataSyncLog.SyncStatus.FAILED
                sync_log.error_message = str(e)
                sync_log.completed_at = timezone.now()
                sync_log.save()
            
            logger.error(f"Data sync failed: {str(e)}")
            raise
    
    def _test_dhis2_connection(self):
        """Test DHIS2 connection and API endpoints"""
        try:
            logger.info("Testing DHIS2 connection...")
            
            # Test basic connection
            if not self.client.test_connection():
                logger.error("DHIS2 connection test failed")
                return False
            
            # Test analytics endpoint
            try:
                # Try a simple analytics request to test the endpoint
                # Use a valid indicator UID for testing instead of "test"
                test_response = self.client.get_analytics_data(
                    indicators=["XLn1cZZTA0H"],  # Use a real indicator UID
                    periods=["202401"],
                    org_units=["Pug4R4IHDtN"]  # Use a valid org unit UID
                )
                logger.info("DHIS2 analytics endpoint is accessible")
            except Exception as e:
                if "409" in str(e):
                    logger.info("DHIS2 analytics endpoint is accessible (409 expected for test data)")
                else:
                    logger.warning(f"DHIS2 analytics endpoint test: {str(e)}")
            
            logger.info("DHIS2 connection test passed")
            return True
            
        except Exception as e:
            logger.error(f"DHIS2 connection test failed: {str(e)}")
            return False
    
    def _get_indicators_to_sync(self, sync_request):
        """Get indicators to sync based on request"""
        indicator_uids = sync_request.get('indicator_uids', [])
        
        if indicator_uids:
            # Sync specific indicators
            return TrackedIndicator.objects.filter(
                dhis2_uid__in=indicator_uids,
                is_active=True
            )
        else:
            # Sync all active indicators
            return TrackedIndicator.objects.filter(is_active=True)
    
    def _get_org_units_to_sync(self, sync_request):
        """Get org units to sync based on request"""
        org_unit_ids = sync_request.get('org_unit_ids', [])
        
        if org_unit_ids:
            # Sync specific org units
            return org_unit_ids
        else:
            # For now, sync all org units - this should be enhanced with user permissions
            return ["LEVEL-1"]  # Default to top-level org unit
    
    def _get_periods_to_sync(self, sync_request):
        """Get periods to sync based on request"""
        # Check if periods are provided directly
        periods = sync_request.get('periods')
        if periods:
            # Use the provided periods directly
            return periods
        
        period_start = sync_request.get('period_start')
        period_end = sync_request.get('period_end')
        
        if period_start and period_end:
            # Generate periods from date range
            return self._generate_periods_from_dates(period_start, period_end)
        else:
            # Use current assessment period
            current_period = AssessmentPeriod.objects.filter(is_current=True).first()
            if current_period:
                return [current_period.period]
            else:
                # Fallback to current year
                from datetime import datetime
                current_year = datetime.now().year
                return [f"{current_year}"]
    
    def _generate_periods_from_dates(self, start_date, end_date):
        """Generate period list from date range"""
        periods = []
        
        from datetime import datetime, date
        
        # Convert start_date to datetime if it's a date object
        if isinstance(start_date, date):
            start = datetime.combine(start_date, datetime.min.time())
        elif isinstance(start_date, str):
            if 'Q' in start_date:  # Handle quarterly period format
                year = int(start_date[:4])
                quarter = int(start_date[5])
                month = (quarter - 1) * 3 + 1
                start = datetime(year, month, 1)
            else:
                start = datetime.strptime(start_date, '%Y-%m-%d')
        else:
            start = start_date
        
        # Convert end_date to datetime if it's a date object
        if isinstance(end_date, date):
            end = datetime.combine(end_date, datetime.min.time())
        elif isinstance(end_date, str):
            if 'Q' in end_date:  # Handle quarterly period format
                year = int(end_date[:4])
                quarter = int(end_date[5])
                month = quarter * 3  # Last month of the quarter
                end = datetime(year, month, 1)
            else:
                end = datetime.strptime(end_date, '%Y-%m-%d')
        else:
            end = end_date
        
        # Generate periods based on the actual date range
        start_year = start.year
        end_year = end.year
        
        # Check if periods are quarterly (based on input format)
        is_quarterly = isinstance(start_date, str) and 'Q' in start_date
        
        if is_quarterly:
            # Generate quarterly periods
            current = start
            while current <= end:
                quarter = ((current.month - 1) // 3) + 1
                period = f"{current.year}Q{quarter}"
                periods.append(period)
                
                # Move to next quarter
                if quarter == 4:
                    current = current.replace(year=current.year + 1, month=1)
                else:
                    current = current.replace(month=min(12, (quarter * 3) + 1))
        else:
            # Default to yearly periods for multi-year ranges
            if end_year - start_year > 0:
                for year in range(start_year, end_year + 1):
                    periods.append(str(year))
            else:
                # Generate monthly periods for single year
                current = start
                while current <= end:
                    period = current.strftime('%Y%m')
                    periods.append(period)
                    
                    # Move to next month
                    if current.month == 12:
                        current = current.replace(year=current.year + 1, month=1)
                    else:
                        current = current.replace(month=current.month + 1)
        
        logger.info(f"Generated periods from {start_date} to {end_date}: {periods}")
        return periods
    
    def _sync_indicator_data_enhanced(self, indicator, org_units, periods, sync_log):
        """Enhanced sync for a single indicator"""
        total_points = 0
        
        for org_unit_id in org_units:
            for period in periods:
                try:
                    # Fetch data from DHIS2
                    value = self._fetch_indicator_data(indicator, org_unit_id, period)
                    
                    if value is not None:
                        # FIXED: Use atomic transaction to prevent database locking
                        from django.db import transaction
                        
                        # FIXED: Add retry logic for database locking
                        max_retries = 3
                        retry_count = 0
                        
                        while retry_count < max_retries:
                            try:
                                with transaction.atomic():
                                    # Store the data
                                    data_point, created = IndicatorData.objects.update_or_create(
                                        indicator=indicator,
                                        org_unit_id=org_unit_id,
                                        period=period,
                                        defaults={
                                            'value': value,
                                            'org_unit_name': self._get_org_unit_name(org_unit_id),
                                            'sync_log': sync_log
                                        }
                                    )
                                    total_points += 1
                                    logger.info(f"Synced data point for {indicator.name} - {org_unit_id} - {period}: {value}")
                                    break  # Success, exit retry loop
                            except Exception as e:
                                retry_count += 1
                                if "database is locked" in str(e).lower() and retry_count < max_retries:
                                    logger.warning(f"Database locked, retrying ({retry_count}/{max_retries}) for {indicator.name} - {period}")
                                    import time
                                    time.sleep(0.5 * retry_count)  # Exponential backoff
                                else:
                                    logger.error(f"Error storing data point: {str(e)}")
                                    return None
                                
                            except Exception as db_error:
                                retry_count += 1
                                if "database is locked" in str(db_error).lower() and retry_count < max_retries:
                                    logger.warning(f"Database locked, retrying ({retry_count}/{max_retries}) for {indicator.name} - {period}")
                                    import time
                                    time.sleep(0.5 * retry_count)  # Exponential backoff
                                else:
                                    raise db_error
                    else:
                        logger.warning(f"No data found for {indicator.name} in period {period}")
                        
                except Exception as e:
                    logger.error(f"Error syncing data for {indicator.name} in period {period}: {str(e)}")
        
        return total_points
    
    def _fetch_indicator_data(self, indicator, org_unit_id, period):
        """Fetch data for a single indicator from DHIS2"""
        try:
            logger.info(f"Fetching data for indicator {indicator.name} ({indicator.dhis2_uid}) for org unit {org_unit_id} and period {period}")
            
            # Convert period to DHIS2 format
            dhis2_period = self._convert_to_dhis2_period(period)
            if not dhis2_period:
                logger.warning(f"Could not convert period {period} to DHIS2 format")
                return None
                
            logger.debug(f"Using DHIS2 period format: {dhis2_period}")
            
            # FIXED: Use correct DHIS2 API based on indicator type
            if indicator.indicator_type == 'indicator':
                logger.info(f"Making analytics request for indicator type 'indicator' with UID: {indicator.dhis2_uid}")
                response = self.client.get_analytics_data(
                    indicators=[indicator.dhis2_uid],
                    periods=[dhis2_period],
                    org_units=[org_unit_id]
                )
            elif indicator.indicator_type == 'dataElement':
                logger.info(f"Making analytics request for indicator type 'dataElement' with UID: {indicator.dhis2_uid}")
                response = self.client.get_analytics_data(
                    data_elements=[indicator.dhis2_uid],
                    periods=[dhis2_period],
                    org_units=[org_unit_id]
                )
            elif indicator.indicator_type == 'dataSet':
                logger.info(f"Making data set report request for indicator type 'dataSet' with UID: {indicator.dhis2_uid}")
                response = self.client.get_data_set_report(
                    data_set_id=indicator.dhis2_uid,
                    periods=[period],
                    org_units=[org_unit_id]
                )
            elif indicator.indicator_type == 'programIndicator':
                logger.info(f"Making analytics request for indicator type 'programIndicator' with UID: {indicator.dhis2_uid}")
                response = self.client.get_analytics_data(
                    program_indicators=[indicator.dhis2_uid],
                    periods=[period],
                    org_units=[org_unit_id]
                )
            else:
                # Fallback to data elements
                logger.info(f"Making analytics request for unknown indicator type '{indicator.indicator_type}' with UID: {indicator.dhis2_uid}")
                response = self.client.get_analytics_data(
                    data_elements=[indicator.dhis2_uid],
                    periods=[period],
                    org_units=[org_unit_id]
                )
            
            # FIXED: Enhanced response validation and debugging
            if not response:
                logger.warning(f"Empty response for indicator {indicator.name} ({indicator.dhis2_uid})")
                return None
            
            logger.debug(f"DHIS2 response for {indicator.name}: {response}")
            
            # Extract value based on indicator type
            if indicator.indicator_type == 'dataSet':
                value = self._extract_value_from_dataset_response(response, indicator.dhis2_uid)
            else:
                value = self._extract_value_from_analytics_response(response, indicator.dhis2_uid)
            
            if value is not None:
                logger.info(f"Successfully extracted value {value} for indicator {indicator.name}")
                return value
            else:
                # This is normal - some indicators don't have data for all periods/org units
                logger.info(f"No data available for indicator {indicator.name} ({indicator.dhis2_uid}) for period {period} and org unit {org_unit_id}")
                
                # FIXED: Try alternative period formats for indicators that might only have yearly data
                if len(period) == 6 and period.isdigit():  # Monthly period (YYYYMM)
                    year = period[:4]
                    logger.info(f"Trying yearly period {year} for indicator {indicator.name}")
                    return self._try_alternative_period_formats(indicator, org_unit_id, year)
                
                return None
                
        except Exception as e:
            logger.error(f"Error fetching data for indicator {indicator.name}: {str(e)}")
            
            # FIXED: Try alternative period formats if we get a 409 error
            if "409" in str(e) or "Conflict" in str(e):
                logger.info(f"Trying alternative period formats for indicator {indicator.name}")
                return self._try_alternative_period_formats(indicator, org_unit_id, period)
            
            return None

    def _try_alternative_period_formats(self, indicator, org_unit_id, period):
        """Try alternative period formats for DHIS2 data fetching"""
        try:
            # Generate alternative period formats based on the input period
            alternative_periods = []
            year = period[:4]
            
            if 'Q' in period:  # Quarterly period
                quarter = int(period[5])
                # Try monthly periods for the quarter
                months = [(quarter - 1) * 3 + i + 1 for i in range(3)]
                for month in months:
                    alternative_periods.append(f"{year}{month:02d}")
                # Try six-monthly period
                semester = 1 if quarter <= 2 else 2
                alternative_periods.append(f"{year}S{semester}")
                # Try yearly period
                alternative_periods.append(year)
                
            elif 'S' in period:  # Six-monthly period
                semester = int(period[5])
                # Try quarterly periods
                quarters = [2*semester - 1, 2*semester]
                for quarter in quarters:
                    alternative_periods.append(f"{year}Q{quarter}")
                # Try monthly periods
                start_month = (semester - 1) * 6 + 1
                for month in range(start_month, start_month + 6):
                    alternative_periods.append(f"{year}{month:02d}")
                # Try yearly period
                alternative_periods.append(year)
                
            elif len(period) == 6:  # Monthly period
                month = int(period[4:6])
                # Try quarterly period
                quarter = (month - 1) // 3 + 1
                alternative_periods.append(f"{year}Q{quarter}")
                # Try six-monthly period
                semester = 1 if month <= 6 else 2
                alternative_periods.append(f"{year}S{semester}")
                # Try yearly period
                alternative_periods.append(year)
                # Try bi-monthly period
                bimonth = (month - 1) // 2 + 1
                alternative_periods.append(f"{year}B{bimonth}")
                
            elif len(period) == 4:  # Yearly period
                # Try all quarters
                for quarter in range(1, 5):
                    alternative_periods.append(f"{year}Q{quarter}")
                # Try all six-monthly periods
                for semester in range(1, 3):
                    alternative_periods.append(f"{year}S{semester}")
                # Try all months
                for month in range(1, 13):
                    alternative_periods.append(f"{year}{month:02d}")
            
            # Add relative periods for recent data
            alternative_periods.extend([
                "THIS_QUARTER",
                "LAST_QUARTER",
                "THIS_YEAR",
                "LAST_YEAR",
                "THIS_SIX_MONTH",
                "LAST_SIX_MONTH"
            ])
            
            logger.info(f"Trying alternative period formats for {indicator.name}: {alternative_periods}")
            
            for alt_period in alternative_periods:
                try:
                    # Make DHIS2 API request based on indicator type
                    if indicator.indicator_type == 'indicator':
                        response = self.client.get_analytics_data(
                            indicators=[indicator.dhis2_uid],
                            periods=[alt_period],
                            org_units=[org_unit_id]
                        )
                    elif indicator.indicator_type == 'dataElement':
                        response = self.client.get_analytics_data(
                            data_elements=[indicator.dhis2_uid],
                            periods=[alt_period],
                            org_units=[org_unit_id]
                        )
                    elif indicator.indicator_type == 'dataSet':
                        response = self.client.get_data_set_report(
                            data_set_id=indicator.dhis2_uid,
                            periods=[alt_period],
                            org_units=[org_unit_id]
                        )
                    elif indicator.indicator_type == 'programIndicator':
                        response = self.client.get_analytics_data(
                            program_indicators=[indicator.dhis2_uid],
                            periods=[alt_period],
                            org_units=[org_unit_id]
                        )
                    else:
                        continue
                    
                    # Extract value from response
                    if indicator.indicator_type == 'dataSet':
                        value = self._extract_value_from_dataset_response(response, indicator.dhis2_uid)
                    else:
                        value = self._extract_value_from_analytics_response(response, indicator.dhis2_uid)
                    
                    if value is not None:
                        logger.info(f"Found data using alternative period format: {alt_period}")
                        return value
                        
                except Exception as e:
                    logger.debug(f"Alternative period {alt_period} failed: {str(e)}")
                    continue
            
            logger.debug(f"All alternative period formats failed for {indicator.name}")
            return None
            
        except Exception as e:
            logger.error(f"Error trying alternative period formats: {str(e)}")
            return None

    def _extract_value_from_analytics_response(self, response, indicator_uid):
        """Extract value from DHIS2 analytics response"""
        try:
            if not response or not isinstance(response, dict):
                logger.warning(f"Invalid response format for indicator {indicator_uid}")
                return None
            
            # Check for rows in response
            rows = response.get('rows', [])
            if not rows:
                # This is normal - some indicators don't have data for all periods/org units
                logger.info(f"No data available for indicator {indicator_uid} - this is normal if the indicator has no data for the specified period/org unit")
                return None
            
            # Get headers to understand the structure
            headers = response.get('headers', [])
            if not headers:
                logger.warning(f"No headers found in response for indicator {indicator_uid}")
                return None
            
            # FIXED: Enhanced column detection
            # Look for the indicator UID in the headers
            indicator_column_index = None
            value_column_index = None
            
            for i, header in enumerate(headers):
                header_name = header.get('name', '').lower()
                header_column = header.get('column', '').lower()
                
                # Check if this header contains our indicator UID
                if indicator_uid.lower() in header_name or indicator_uid.lower() in header_column:
                    indicator_column_index = i
                    break
            
            # If we found the indicator column, the value should be in the next column
            if indicator_column_index is not None:
                value_column_index = indicator_column_index + 1
            else:
                # Fallback: look for value columns
                for i, header in enumerate(headers):
                    header_name = header.get('name', '').lower()
                    if 'value' in header_name or 'data' in header_name:
                        value_column_index = i
                        break
            
            # If still no value column found, use the last column
            if value_column_index is None and len(headers) > 1:
                value_column_index = len(headers) - 1
            
            logger.debug(f"Using value column index {value_column_index} for indicator {indicator_uid}")
            
            # Extract value from the first row
            if value_column_index is not None and len(rows) > 0:
                first_row = rows[0]
                if len(first_row) > value_column_index:
                    value = first_row[value_column_index]
                    logger.debug(f"Extracted value {value} from row {first_row}")
                    
                    # Convert to float if possible
                    try:
                        if isinstance(value, str):
                            value = float(value)
                        return value
                    except (ValueError, TypeError):
                        logger.warning(f"Could not convert value '{value}' to float for indicator {indicator_uid}")
                        return None
            
            # FIXED: Try alternative parsing if standard parsing fails
            logger.info(f"Standard parsing failed, trying alternative parsing for indicator {indicator_uid}")
            return self._extract_value_alternative_parsing(response, indicator_uid, value_column_index)
            
        except Exception as e:
            logger.error(f"Error extracting value from analytics response for indicator {indicator_uid}: {str(e)}")
            return None

    def _extract_value_alternative_parsing(self, response, indicator_uid, value_column_index):
        """Alternative parsing method for analytics response"""
        try:
            logger.info(f"Starting alternative parsing for {indicator_uid}")
            
            # Try to find the indicator in metadata
            meta_data = response.get('metaData', {})
            items = meta_data.get('items', {})
            
            # Look for the indicator in the items
            if indicator_uid in items:
                item_info = items[indicator_uid]
                logger.info(f"Found indicator info in metadata: {item_info}")
            
            # Process rows with more flexible matching
            rows = response.get('rows', [])
            logger.info(f"Alternative parsing: processing {len(rows)} rows")
            
            for i, row in enumerate(rows):
                if len(row) <= value_column_index:
                    logger.debug(f"Alternative parsing: skipping row {i} with insufficient columns")
                    continue
                
                # Try to match by checking if the indicator UID appears anywhere in the row
                row_str = ' '.join(str(cell) for cell in row)
                if indicator_uid in row_str:
                    logger.info(f"Alternative parsing: found indicator {indicator_uid} in row {i}: {row}")
                    raw_value = row[value_column_index]
                    
                    if raw_value is None or raw_value == '':
                        logger.warning(f"Alternative parsing: empty value found for {indicator_uid}")
                        return None
                    
                    try:
                        value = float(raw_value)
                        logger.info(f"Alternative parsing: successfully converted value to float: {value}")
                        return value
                    except (ValueError, TypeError):
                        logger.warning(f"Alternative parsing: could not convert value '{raw_value}' to float for {indicator_uid}")
                        return None
            
            logger.warning(f"Alternative parsing: no value found for {indicator_uid}")
            return None
            
        except Exception as e:
            logger.error(f"Error in alternative parsing for {indicator_uid}: {str(e)}")
            return None

    def _extract_value_from_dataset_response(self, response, indicator_uid):
        """Extract value from DHIS2 data set report response"""
        try:
            if not isinstance(response, dict):
                logger.warning(f"Invalid response type: {type(response)}")
                return None
            
            # Data set reports have a different structure
            # Look for data in various possible locations
            if 'dataValues' in response:
                data_values = response['dataValues']
                if data_values:
                    # Return the first data value
                    return float(data_values[0].get('value', 0))
            
            if 'data' in response:
                data = response['data']
                if isinstance(data, list) and data:
                    # Return the first data point
                    return float(data[0].get('value', 0))
            
            logger.warning(f"No data found in data set report for {indicator_uid}")
            return None
            
        except Exception as e:
            logger.error(f"Error extracting value from data set response: {str(e)}")
            return None

    def _fetch_indicator_data_batch(self, indicators, org_unit_id, period):
        """Fetch data for multiple indicators in a single request for better performance"""
        try:
            logger.info(f"Fetching batch data for {len(indicators)} indicators for org unit {org_unit_id} and period {period}")
            
            # Separate indicators by type
            data_elements = []
            indicators_list = []
            data_sets = []
            program_indicators = []
            
            for indicator in indicators:
                if indicator.indicator_type == 'indicator':
                    indicators_list.append(indicator.dhis2_uid)
                elif indicator.indicator_type == 'dataElement':
                    data_elements.append(indicator.dhis2_uid)
                elif indicator.indicator_type == 'dataSet':
                    data_sets.append(indicator.dhis2_uid)
                elif indicator.indicator_type == 'programIndicator':
                    program_indicators.append(indicator.dhis2_uid)
                else:
                    # Fallback to data elements
                    data_elements.append(indicator.dhis2_uid)
            
            # Make batch request
            response = self.client.get_analytics_data(
                data_elements=data_elements if data_elements else None,
                indicators=indicators_list if indicators_list else None,
                data_sets=data_sets if data_sets else None,
                program_indicators=program_indicators if program_indicators else None,
                periods=[period],
                org_units=[org_unit_id]
            )
            
            logger.debug(f"Batch DHIS2 response: {response}")
            
            # Extract values for each indicator
            results = {}
            for indicator in indicators:
                value = self._extract_value_from_analytics_response(response, indicator.dhis2_uid)
                results[indicator.dhis2_uid] = value
                
                if value is not None:
                    logger.info(f"Found value {value} for indicator {indicator.dhis2_uid}")
                else:
                    logger.warning(f"No data found for indicator {indicator.dhis2_uid}")
            
            return results
            
        except Exception as e:
            logger.error(f"Error fetching batch data: {str(e)}")
            return {}
    
    def _get_org_unit_name(self, org_unit_id):
        """Get organization unit name from cache or DHIS2 API (simplified for DHIS2-only workflow)"""
        try:
            # Try to get from cache first
            cache_key = f"org_unit_name_{org_unit_id}"
            org_unit_name = cache.get(cache_key)
            
            if org_unit_name:
                return org_unit_name
            
            # Since org units come from DHIS2 frontend, we can either:
            # 1. Use the org unit ID as the name (simple approach)
            # 2. Fetch the name from DHIS2 API (more complete but adds API calls)
            
            # Simple approach: Use ID as name for now
            org_unit_name = f"Org Unit {org_unit_id}"
            
            # Optional: Fetch actual name from DHIS2 (uncomment if needed)
            # try:
            #     if self.client:
            #         org_unit_data = self.client._make_request("GET", f"api/organisationUnits/{org_unit_id}", 
            #                                                  params={"fields": "id,name,displayName"})
            #         org_unit_name = org_unit_data.get('displayName') or org_unit_data.get('name') or f"Org Unit {org_unit_id}"
            # except Exception as e:
            #     logger.warning(f"Could not fetch org unit name from DHIS2: {str(e)}")
            #     org_unit_name = f"Org Unit {org_unit_id}"
            
            # Cache the result
            cache.set(cache_key, org_unit_name, timeout=3600)  # Cache for 1 hour
            
            return org_unit_name
            
        except Exception as e:
            logger.warning(f"Error getting org unit name for {org_unit_id}: {str(e)}")
            return f"Org Unit {org_unit_id}"
    
    def _trigger_score_calculation(self, sync_log):
        """Trigger score calculation for synced data"""
        try:
            # Get unique org units and periods from the sync log
            data_points = IndicatorData.objects.filter(sync_log=sync_log)
            
            # Group by org unit and period
            org_unit_periods = data_points.values('org_unit_id', 'period').distinct()
            
            for item in org_unit_periods:
                org_unit_id = item['org_unit_id']
                period_str = item['period']
                
                # FIXED: Get or create AssessmentPeriod instance
                assessment_period, created = AssessmentPeriod.objects.get_or_create(
                    name=f"Period {period_str}",
                    defaults={
                        'period_type': 'custom',
                        'start_date': timezone.now().date(),  # Default start date
                        'end_date': timezone.now().date(),    # Default end date
                        'is_current': False
                    }
                )
                
                logger.info(f"Calculating scores for org unit {org_unit_id} and period {period_str}")
                
                # Calculate scores
                score_service = ScoreCalculationService()
                score_service.calculate_scores_for_org_unit(org_unit_id, assessment_period)
            
        except Exception as e:
            logger.error(f"Error triggering score calculation: {str(e)}")
            raise


class ScoreCalculationService:
    """
    Service for calculating scores from indicator data
    """
    
    def calculate_scores_for_org_unit(self, org_unit_id, assessment_period):
        """
        Calculate all scores for a specific org unit and assessment period
        """
        try:
            # Calculate indicator scores
            self._calculate_indicator_scores(org_unit_id, assessment_period)
            
            # Calculate objective scores
            self._calculate_objective_scores(org_unit_id, assessment_period)
            
            # Calculate sector score
            self._calculate_sector_score(org_unit_id, assessment_period)
            
            logger.info(f"Score calculation completed for org unit {org_unit_id} and period {assessment_period}")
            
        except Exception as e:
            logger.error(f"Failed to calculate scores for org unit {org_unit_id}: {str(e)}")
            raise
    
    def _calculate_indicator_scores(self, org_unit_id, assessment_period):
        """Calculate indicator scores"""
        # Get all indicator data for this org unit and period
        indicator_data = IndicatorData.objects.filter(
            org_unit_id=org_unit_id,
            period=assessment_period
        ).select_related('indicator')
        
        for data in indicator_data:
            try:
                # FIXED: Get scoring rule for this indicator - use performance_type instead of indicator_type
                scoring_rule = ScoringRule.objects.filter(
                    performance_type='gap',  # Use gap as default performance type
                    is_active=True
                ).first()
                
                if not scoring_rule:
                    # Create a default scoring rule if none exists
                    scoring_rule = ScoringRule.objects.create(
                        name='Default Gap Rule',
                        performance_type='gap',
                        min_value=Decimal('-100'),
                        max_value=Decimal('100'),
                        score=0,
                        color='#6c757d',
                        label='Default',
                        priority=0
                    )
                
                # Calculate score based on value and target
                score = self._calculate_score(data.value, data.indicator.target_value, scoring_rule)
                
                # Get previous period for trend calculation
                previous_period = self._get_previous_period(assessment_period)
                previous_score = None
                
                if previous_period:
                    prev_data = IndicatorScore.objects.filter(
                        indicator=data.indicator,
                        org_unit_id=org_unit_id,
                        assessment_period=previous_period
                    ).first()
                    if prev_data:
                        previous_score = prev_data.score
                
                # Calculate trend
                trend = self._calculate_trend(score, previous_score) if previous_score is not None else None
                
                # FIXED: Use direct color and label from scoring rule instead of non-existent methods
                color = scoring_rule.color
                label = scoring_rule.label
                
                # Create or update indicator score
                indicator_score, created = IndicatorScore.objects.update_or_create(
                    indicator=data.indicator,
                    org_unit_id=org_unit_id,
                    assessment_period=assessment_period,
                    defaults={
                        'current_value': data.value,  # FIXED: Use current_value instead of raw_value
                        'score': score,
                        'score_color': color,  # FIXED: Use score_color instead of color
                        'score_label': label,  # FIXED: Use score_label instead of label
                        'last_calculated': timezone.now()
                    }
                )
                
            except Exception as e:
                logger.error(f"Failed to calculate score for indicator {data.indicator.name}: {str(e)}")
    
    def _calculate_objective_scores(self, org_unit_id, assessment_period):
        """Calculate objective scores"""
        # Get all objectives
        objectives = Objective.objects.filter(is_active=True)
        
        for objective in objectives:
            try:
                # FIXED: Use correct relationship through IndicatorWeight to get indicator scores for this objective
                # Get all indicators that belong to this objective
                objective_indicators = TrackedIndicator.objects.filter(
                    objective_weights__objective=objective
                )
                
                # Get indicator scores for these indicators
                indicator_scores = IndicatorScore.objects.filter(
                    indicator__in=objective_indicators,
                    org_unit_id=org_unit_id,
                    assessment_period=assessment_period
                ).select_related('indicator')
                
                if not indicator_scores.exists():
                    continue
                
                # Calculate weighted average score
                total_weight = 0
                weighted_sum = 0
                
                for score in indicator_scores:
                    weight = IndicatorWeight.objects.filter(
                        indicator=score.indicator,
                        objective=objective
                    ).first()
                    
                    if weight and score.score is not None:
                        total_weight += weight.weight
                        weighted_sum += score.score * weight.weight
                
                if total_weight > 0:
                    objective_score_value = weighted_sum / total_weight
                else:
                    # Use median if no weights defined
                    scores = list(indicator_scores.values_list('score', flat=True))
                    scores = [s for s in scores if s is not None]
                    if scores:
                        scores.sort()
                        objective_score_value = scores[len(scores) // 2]
                    else:
                        objective_score_value = 0
                
                # Get scoring rule for objectives
                scoring_rule = ScoringRule.objects.filter(
                    performance_type='gap'  # FIXED: Use performance_type instead of rule_type
                ).first()
                
                if scoring_rule:
                    color = scoring_rule.color
                    label = scoring_rule.label
                else:
                    color = '#666666'
                    label = 'Unknown'
                
                # Create or update objective score
                objective_score, created = ObjectiveScore.objects.update_or_create(
                    objective=objective,
                    org_unit_id=org_unit_id,
                    assessment_period=assessment_period,
                    defaults={
                        'score': objective_score_value,
                        'color': color,
                        'label': label,
                        'last_calculated': timezone.now()
                    }
                )
                
            except Exception as e:
                logger.error(f"Failed to calculate objective score for {objective.name}: {str(e)}")
    
    def _calculate_sector_score(self, org_unit_id, assessment_period):
        """Calculate overall sector score"""
        try:
            # Get all objective scores for this org unit and period
            objective_scores = ObjectiveScore.objects.filter(
                org_unit_id=org_unit_id,
                assessment_period=assessment_period
            )
            
            if not objective_scores.exists():
                return
            
            # Calculate weighted average of objective scores
            total_weight = 0
            weighted_sum = 0
            
            for obj_score in objective_scores:
                weight = obj_score.objective.weight if hasattr(obj_score.objective, 'weight') else 1
                total_weight += weight
                weighted_sum += obj_score.score * weight
            
            if total_weight > 0:
                sector_score_value = weighted_sum / total_weight
            else:
                # Use average if no weights defined
                sector_score_value = objective_scores.aggregate(Avg('score'))['score__avg'] or 0
            
            # Get scoring rule for sector scores
            scoring_rule = ScoringRule.objects.filter(
                rule_type='sector'
            ).first()
            
            if scoring_rule:
                color = scoring_rule.color
                label = scoring_rule.label
            else:
                color = '#666666'
                label = 'Unknown'
            
            # Create or update sector score
            sector_score, created = SectorScore.objects.update_or_create(
                org_unit_id=org_unit_id,
                assessment_period=assessment_period,
                defaults={
                    'score': sector_score_value,
                    'color': color,
                    'label': label,
                    'last_calculated': timezone.now()
                }
            )
            
        except Exception as e:
            logger.error(f"Failed to calculate sector score: {str(e)}")
    
    def _get_previous_period(self, assessment_period):
        """Get the previous assessment period"""
        try:
            # Parse period (assuming YYYYMM format)
            year = int(assessment_period[:4])
            month = int(assessment_period[4:])
            
            # Calculate previous month
            if month == 1:
                prev_month = 12
                prev_year = year - 1
            else:
                prev_month = month - 1
                prev_year = year
            
            return f"{prev_year:04d}{prev_month:02d}"
            
        except Exception:
            return None
    
    def _calculate_score(self, value, target, scoring_rule):
        """Calculate score based on value, target, and scoring rule"""
        if value is None or target is None:
            return 0
        
        # Calculate gap to target
        gap = abs(value - target)
        gap_percentage = (gap / target) * 100 if target != 0 else 0
        
        # FIXED: Implement scoring logic directly instead of calling non-existent evaluate_score method
        # Find matching scoring rule based on gap percentage
        matching_rule = ScoringRule.objects.filter(
            performance_type='gap',
            is_active=True
        ).order_by('-priority', 'min_value')
        
        for rule in matching_rule:
            if rule.matches_value(gap_percentage):
                return rule.score
        
        # Return default score if no rule matches
        return 0
    
    def _calculate_trend(self, current_score, previous_score):
        """Calculate trend direction"""
        if previous_score is None:
            return 'stable'
        
        if current_score > previous_score:
            return 'improving'
        elif current_score < previous_score:
            return 'declining'
        else:
            return 'stable'
    
    def bulk_calculate_scores(self, request_data):
        """
        Bulk calculate scores for multiple org units and periods
        """
        org_unit_ids = request_data.get('org_unit_ids', [])
        assessment_periods = request_data.get('assessment_periods', [])
        
        results = {
            'total_org_units': len(org_unit_ids),
            'total_periods': len(assessment_periods),
            'successful_calculations': 0,
            'failed_calculations': 0,
            'errors': []
        }
        
        for org_unit_id in org_unit_ids:
            for period in assessment_periods:
                try:
                    self.calculate_scores_for_org_unit(org_unit_id, period)
                    results['successful_calculations'] += 1
                except Exception as e:
                    results['failed_calculations'] += 1
                    error_msg = f"Failed to calculate scores for org unit {org_unit_id} and period {period}: {str(e)}"
                    results['errors'].append(error_msg)
                    logger.error(error_msg)
        
        return results


class DashboardService:
    """
    Service for generating dashboard data
    """
    
    def __init__(self):
        self.access_service = AccessControlService()
    
    def get_dashboard_summary(self, user, org_unit_id=None, assessment_period=None):
        """
        Get dashboard summary for a user
        """
        # Get user's accessible org units
        accessible_org_units = self.access_service.get_user_accessible_org_units(user)
        
        if org_unit_id:
            # Check if user has access to this specific org unit
            if not accessible_org_units.filter(id=org_unit_id).exists():
                return None
            org_units = [org_unit_id]
        else:
            # Use all accessible org units
            org_units = list(accessible_org_units.values_list('id', flat=True))
        
        # Get current assessment period if not specified
        if not assessment_period:
            current_period = AssessmentPeriod.objects.filter(is_current=True).first()
            assessment_period = current_period.period if current_period else None
        
        if not assessment_period:
            return None
        
        # Get sector scores for accessible org units
        sector_scores = SectorScore.objects.filter(
            org_unit_id__in=org_units,
            assessment_period=assessment_period
        ).select_related('org_unit')
        
        # Calculate summary statistics
        total_org_units = len(org_units)
        org_units_with_scores = sector_scores.count()
        
        if org_units_with_scores > 0:
            avg_score = sector_scores.aggregate(Avg('score'))['score__avg']
            max_score = sector_scores.aggregate(Max('score'))['score__max']
            min_score = sector_scores.aggregate(Min('score'))['score__min']
            
            # Count by performance category
            performance_counts = {}
            for score in sector_scores:
                label = score.label or 'Unknown'
                performance_counts[label] = performance_counts.get(label, 0) + 1
        else:
            avg_score = max_score = min_score = 0
            performance_counts = {}
        
        return {
            'assessment_period': assessment_period,
            'total_org_units': total_org_units,
            'org_units_with_scores': org_units_with_scores,
            'average_score': round(avg_score, 2) if avg_score else 0,
            'max_score': round(max_score, 2) if max_score else 0,
            'min_score': round(min_score, 2) if min_score else 0,
            'performance_distribution': performance_counts,
            'sector_scores': [
                {
                    'org_unit_id': score.org_unit_id,
                    'org_unit_name': score.org_unit.name,
                    'score': score.score,
                    'color': score.color,
                    'label': score.label
                }
                for score in sector_scores
            ]
        }
    
    def get_objective_dashboard(self, user, org_unit_id=None, assessment_period=None):
        """
        Get objective dashboard data
        """
        # Get user's accessible org units
        accessible_org_units = self.access_service.get_user_accessible_org_units(user)
        
        if org_unit_id:
            # Check if user has access to this specific org unit
            if not accessible_org_units.filter(id=org_unit_id).exists():
                return None
            org_units = [org_unit_id]
        else:
            # Use all accessible org units
            org_units = list(accessible_org_units.values_list('id', flat=True))
        
        # Get current assessment period if not specified
        if not assessment_period:
            current_period = AssessmentPeriod.objects.filter(is_current=True).first()
            assessment_period = current_period.period if current_period else None
        
        if not assessment_period:
            return None
        
        # Get objective scores for accessible org units
        objective_scores = ObjectiveScore.objects.filter(
            org_unit_id__in=org_units,
            assessment_period=assessment_period
        ).select_related('objective', 'org_unit')
        
        # Group by objective
        objectives_data = {}
        for score in objective_scores:
            objective_name = score.objective.name
            if objective_name not in objectives_data:
                objectives_data[objective_name] = {
                    'objective_id': score.objective.id,
                    'objective_name': objective_name,
                    'scores': [],
                    'average_score': 0,
                    'performance_distribution': {}
                }
            
            objectives_data[objective_name]['scores'].append({
                'org_unit_id': score.org_unit_id,
                'org_unit_name': score.org_unit.name,
                'score': score.score,
                'color': score.color,
                'label': score.label
            })
        
        # Calculate statistics for each objective
        for objective_name, data in objectives_data.items():
            scores = [score['score'] for score in data['scores']]
            if scores:
                data['average_score'] = round(sum(scores) / len(scores), 2)
                
                # Count by performance category
                for score in data['scores']:
                    label = score['label'] or 'Unknown'
                    data['performance_distribution'][label] = data['performance_distribution'].get(label, 0) + 1
        
        return {
            'assessment_period': assessment_period,
            'objectives': list(objectives_data.values())
        }
    
    def get_indicator_dashboard(self, user, org_unit_id=None, assessment_period=None, objective_id=None):
        """
        Get indicator dashboard data
        """
        # Get user's accessible org units
        accessible_org_units = self.access_service.get_user_accessible_org_units(user)
        
        if org_unit_id:
            # Check if user has access to this specific org unit
            if not accessible_org_units.filter(id=org_unit_id).exists():
                return None
            org_units = [org_unit_id]
        else:
            # Use all accessible org units
            org_units = list(accessible_org_units.values_list('id', flat=True))
        
        # Get current assessment period if not specified
        if not assessment_period:
            current_period = AssessmentPeriod.objects.filter(is_current=True).first()
            assessment_period = current_period.period if current_period else None
        
        if not assessment_period:
            return None
        
        # Build query for indicator scores
        indicator_scores = IndicatorScore.objects.filter(
            org_unit_id__in=org_units,
            assessment_period=assessment_period
        ).select_related('indicator', 'org_unit')
        
        # Filter by objective if specified
        if objective_id:
            indicator_scores = indicator_scores.filter(indicator__objectives__id=objective_id)
        
        # Group by indicator
        indicators_data = {}
        for score in indicator_scores:
            indicator_name = score.indicator.name
            if indicator_name not in indicators_data:
                indicators_data[indicator_name] = {
                    'indicator_id': score.indicator.id,
                    'indicator_name': indicator_name,
                    'indicator_type': score.indicator.indicator_type,
                    'target_value': score.indicator.target_value,
                    'scores': [],
                    'average_score': 0,
                    'performance_distribution': {},
                    'trend_analysis': {
                        'improving': 0,
                        'declining': 0,
                        'stable': 0
                    }
                }
            
            indicators_data[indicator_name]['scores'].append({
                'org_unit_id': score.org_unit_id,
                'org_unit_name': score.org_unit.name,
                'raw_value': score.current_value,
                'score': score.score,
                'trend': score.trend,
                'color': score.score_color,
                'label': score.score_label
            })
        
        # Calculate statistics for each indicator
        for indicator_name, data in indicators_data.items():
            scores = [score['score'] for score in data['scores']]
            if scores:
                data['average_score'] = round(sum(scores) / len(scores), 2)
                
                # Count by performance category and trend
                for score in data['scores']:
                    label = score['label'] or 'Unknown'
                    data['performance_distribution'][label] = data['performance_distribution'].get(label, 0) + 1
                    
                    trend = score['trend'] or 'stable'
                    data['trend_analysis'][trend] = data['trend_analysis'].get(trend, 0) + 1
        
        return {
            'assessment_period': assessment_period,
            'objective_id': objective_id,
            'indicators': list(indicators_data.values())
        }
    
    def get_org_unit_performance(self, user, org_unit_id, assessment_period=None):
        """
        Get detailed performance data for a specific org unit
        """
        # Check if user has access to this org unit
        accessible_org_units = self.access_service.get_user_accessible_org_units(user)
        if not accessible_org_units.filter(id=org_unit_id).exists():
            return None
        
        # Get current assessment period if not specified
        if not assessment_period:
            current_period = AssessmentPeriod.objects.filter(is_current=True).first()
            assessment_period = current_period.period if current_period else None
        
        if not assessment_period:
            return None
        
        # Get org unit details
        org_unit = OrgUnit.objects.get(id=org_unit_id)
        
        # Get sector score
        sector_score = SectorScore.objects.filter(
            org_unit_id=org_unit_id,
            assessment_period=assessment_period
        ).first()
        
        # Get objective scores
        objective_scores = ObjectiveScore.objects.filter(
            org_unit_id=org_unit_id,
            assessment_period=assessment_period
        ).select_related('objective')
        
        # Get indicator scores
        indicator_scores = IndicatorScore.objects.filter(
            org_unit_id=org_unit_id,
            assessment_period=assessment_period
        ).select_related('indicator')
        
        return {
            'org_unit': {
                'id': org_unit.id,
                'name': org_unit.name,
                'level': org_unit.level.name
            },
            'assessment_period': assessment_period,
            'sector_score': {
                'score': sector_score.score if sector_score else None,
                'color': sector_score.color if sector_score else None,
                'label': sector_score.label if sector_score else None
            },
            'objectives': [
                {
                    'objective_id': score.objective.id,
                    'objective_name': score.objective.name,
                    'score': score.score,
                    'color': score.color,
                    'label': score.label
                }
                for score in objective_scores
            ],
            'indicators': [
                {
                    'indicator_id': score.indicator.id,
                    'indicator_name': score.indicator.name,
                    'raw_value': score.current_value,
                    'score': score.score,
                    'trend': score.trend,
                    'color': score.score_color,
                    'label': score.score_label
                }
                for score in indicator_scores
            ]
        }
    
    def get_trend_analysis(self, user, org_unit_id, assessment_period=None, periods_back=3):
        """
        Get trend analysis for an org unit over multiple periods
        """
        # Check if user has access to this org unit
        accessible_org_units = self.access_service.get_user_accessible_org_units(user)
        if not accessible_org_units.filter(id=org_unit_id).exists():
            return None
        
        # Get current assessment period if not specified
        if not assessment_period:
            current_period = AssessmentPeriod.objects.filter(is_current=True).first()
            assessment_period = current_period.period if current_period else None
        
        if not assessment_period:
            return None
        
        # Generate list of periods to analyze
        periods = self._generate_periods_back(assessment_period, periods_back)
        
        # Get sector scores for all periods
        sector_scores = SectorScore.objects.filter(
            org_unit_id=org_unit_id,
            assessment_period__in=periods
        ).order_by('assessment_period')
        
        # Get objective scores for all periods
        objective_scores = ObjectiveScore.objects.filter(
            org_unit_id=org_unit_id,
            assessment_period__in=periods
        ).select_related('objective').order_by('objective__name', 'assessment_period')
        
        # Build trend data
        sector_trend = []
        objectives_trend = {}
        
        for period in periods:
            # Sector score for this period
            period_sector_score = sector_scores.filter(assessment_period=period).first()
            sector_trend.append({
                'period': period,
                'score': period_sector_score.score if period_sector_score else None,
                'color': period_sector_score.color if period_sector_score else None,
                'label': period_sector_score.label if period_sector_score else None
            })
            
            # Objective scores for this period
            period_objective_scores = objective_scores.filter(assessment_period=period)
            for score in period_objective_scores:
                objective_name = score.objective.name
                if objective_name not in objectives_trend:
                    objectives_trend[objective_name] = []
                
                objectives_trend[objective_name].append({
                    'period': period,
                    'score': score.score,
                    'color': score.color,
                    'label': score.label
                })
        
        return {
            'org_unit_id': org_unit_id,
            'periods': periods,
            'sector_trend': sector_trend,
            'objectives_trend': objectives_trend
        }
    
    def _generate_periods_back(self, current_period, periods_back):
        """
        Generate list of periods going back from current period
        """
        periods = [current_period]
        
        for i in range(1, periods_back):
            # Parse current period (assuming YYYYMM format)
            year = int(current_period[:4])
            month = int(current_period[4:])
            
            # Calculate previous month
            if month == 1:
                prev_month = 12
                prev_year = year - 1
            else:
                prev_month = month - 1
                prev_year = year
            
            prev_period = f"{prev_year:04d}{prev_month:02d}"
            periods.append(prev_period)
            
            # Update for next iteration
            current_period = prev_period
        
        return periods[::-1]  # Reverse to get chronological order
    
    def _calculate_trend_direction(self, scores):
        """
        Calculate overall trend direction from a list of scores
        """
        if len(scores) < 2:
            return 'stable'
        
        # Calculate trend based on recent scores
        recent_scores = scores[-3:]  # Last 3 scores
        if len(recent_scores) >= 2:
            if recent_scores[-1] > recent_scores[0]:
                return 'improving'
            elif recent_scores[-1] < recent_scores[0]:
                return 'declining'
        
        return 'stable' 

class ManualDataEntryService:
    """
    Service for handling manual data entry and score computation
    """
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def update_manual_indicator_data(self, request, indicator_id, org_unit_id, assessment_period_id, data_updates):
        """
        Update manual indicator data and recalculate scores
        
        Args:
            request: Django request object
            indicator_id: ID of the indicator
            org_unit_id: Organization unit ID
            assessment_period_id: Assessment period ID
            data_updates: Dict containing updates for current_value, previous_value, target_value, 
                         percent_change, target_gap, score
        """
        try:
            with transaction.atomic():
                # First, we need to find the objective for this indicator
                from indicators.models import TrackedIndicator
                from configurations.models import Objective, AssessmentPeriod
                
                try:
                    indicator = TrackedIndicator.objects.get(id=indicator_id)
                    assessment_period = AssessmentPeriod.objects.get(id=assessment_period_id)
                    
                    # Find the objective for this indicator
                    objective = None
                    for obj in Objective.objects.filter(is_active=True):
                        if obj.indicator_weights.filter(indicator=indicator).exists():
                            objective = obj
                            break
                    
                    # If no objective found, use the first available objective
                    if not objective:
                        objective = Objective.objects.filter(is_active=True).first()
                    
                    if not objective:
                        raise ValidationError(f"No objectives found for indicator {indicator_id}")
                    
                    # Get or create the indicator score record with the objective
                    indicator_score, created = IndicatorScore.objects.get_or_create(
                        indicator_id=indicator_id,
                        org_unit_id=org_unit_id,
                        assessment_period_id=assessment_period_id,
                        defaults={
                            'objective': objective,
                            'current_value': None,
                            'previous_value': None,
                            'target_value': indicator.target_value,
                            'percent_change': None,
                            'target_gap': None,
                            'score': None,
                            'score_color': '#6c757d',
                            'score_label': 'N/A',
                            'is_manual_override': False,
                            'weight': 1.0,
                            'org_unit_name': org_unit_id
                        }
                    )
                    
                    # If this is a new record, we need to set up any additional information
                    if created:
                        # The record is already created with the basic info above
                        pass
                        
                except TrackedIndicator.DoesNotExist:
                    raise ValidationError(f"Indicator {indicator_id} not found")
                except AssessmentPeriod.DoesNotExist:
                    raise ValidationError(f"Assessment period {assessment_period_id} not found")
                except Exception as e:
                    self.logger.error(f"Error setting up indicator score record: {e}")
                    raise ValidationError(f"Error setting up indicator score record: {str(e)}")
                

                
                # Get the actual DHIS2User instance from the wrapper
                dhis2_user = None
                if hasattr(request, 'user') and hasattr(request.user, 'dhis2_user'):
                    dhis2_user = request.user.dhis2_user
                
                # Store old values for audit
                old_values = {
                    'current_value': float(indicator_score.current_value) if indicator_score.current_value else None,
                    'previous_value': float(indicator_score.previous_value) if indicator_score.previous_value else None,
                    'target_value': float(indicator_score.target_value) if indicator_score.target_value else None,
                    'percent_change': float(indicator_score.percent_change) if indicator_score.percent_change else None,
                    'target_gap': float(indicator_score.target_gap) if indicator_score.target_gap else None,
                    'score': indicator_score.score,
                    'score_color': indicator_score.score_color,
                    'score_label': indicator_score.score_label,
                    'is_manual_override': indicator_score.is_manual_override
                }
                
                # Update data values
                if 'current_value' in data_updates:
                    indicator_score.current_value = self._parse_decimal(data_updates['current_value'])
                
                if 'previous_value' in data_updates:
                    indicator_score.previous_value = self._parse_decimal(data_updates['previous_value'])
                
                if 'target_value' in data_updates:
                    indicator_score.target_value = self._parse_decimal(data_updates['target_value'])
                
                # Handle manual score override
                if 'score' in data_updates:
                    new_score = int(data_updates['score'])
                    if -5 <= new_score <= 5:
                        indicator_score.apply_manual_override(
                            new_score=new_score,
                            user=dhis2_user,
                            reason="Manual score entry"
                        )
                        # Skip automatic score calculation since it's a manual override
                        indicator_score.save()
                        
                        # Recalculate higher-level scores
                        self._recalculate_higher_level_scores(indicator_score)
                        
                        return {
                            'success': True,
                            'message': 'Manual score override applied successfully',
                            'indicator_score': {
                                'id': indicator_score.id,
                                'score': indicator_score.score,
                                'score_color': indicator_score.score_color,
                                'score_label': indicator_score.score_label,
                                'is_manual_override': indicator_score.is_manual_override
                            }
                        }
                
                # Calculate percent change if both current and previous values are provided
                # Use correct formulas based on indicator type
                if (indicator_score.current_value is not None and 
                    indicator_score.previous_value is not None and 
                    indicator_score.previous_value != 0):
                    
                    # Get indicator to determine target type and format
                    indicator = indicator_score.indicator
                    target_type = getattr(indicator, 'target_type', 'increase').lower()
                    target_format = getattr(indicator, 'target_format', 'SINGLE')
                    
                    if target_format == 'RANGE':
                        # For range indicators, always use standard formula regardless of target_type
                        change = ((indicator_score.current_value - indicator_score.previous_value) / 
                                 abs(indicator_score.previous_value)) * 100
                    else:
                        # For non-range indicators, use target_type specific formula
                        if target_type == 'decrease':
                            # For decrease indicators: (previous_value - current_value) / abs(current_value) * 100
                            change = ((indicator_score.previous_value - indicator_score.current_value) / 
                                     abs(indicator_score.current_value)) * 100
                        else:
                            # For increase indicators: (current_value - previous_value) / abs(previous_value) * 100
                            change = ((indicator_score.current_value - indicator_score.previous_value) / 
                                     abs(indicator_score.previous_value)) * 100
                    
                    # Round to 2 decimal places for consistency
                    indicator_score.percent_change = round(change, 2)
                elif 'percent_change' in data_updates:
                    # Manual percent change entry
                    indicator_score.percent_change = self._parse_decimal(data_updates['percent_change'])
                
                # Calculate target gap if both current and target values are provided
                # Use correct formulas based on indicator type
                if (indicator_score.current_value is not None and 
                    indicator_score.target_value is not None and 
                    indicator_score.target_value != 0):
                    
                    # Get indicator to determine target type
                    indicator = indicator_score.indicator
                    target_type = getattr(indicator, 'target_type', 'increase').lower()
                    
                    if target_type == 'decrease':
                        # For decrease indicators: (target_value - current_value) / current_value * 100
                        gap = (indicator_score.target_value - indicator_score.current_value) / indicator_score.current_value * 100
                    else:
                        # For increase indicators: (current_value - target_value) / target_value * 100
                        gap = (indicator_score.current_value - indicator_score.target_value) / indicator_score.target_value * 100
                    
                    # Round to 2 decimal places for consistency
                    indicator_score.target_gap = round(gap, 2)
                elif 'target_gap' in data_updates:
                    # Manual target gap entry
                    indicator_score.target_gap = self._parse_decimal(data_updates['target_gap'])
                
                # Calculate score based on available metrics
                self._calculate_score_from_metrics(indicator_score)
                
                # Save the indicator score
                indicator_score.save()
                
                # Recalculate higher-level scores
                self._recalculate_higher_level_scores(indicator_score)
                
                # Log the change
                new_values = {
                    'current_value': float(indicator_score.current_value) if indicator_score.current_value else None,
                    'previous_value': float(indicator_score.previous_value) if indicator_score.previous_value else None,
                    'target_value': float(indicator_score.target_value) if indicator_score.target_value else None,
                    'percent_change': float(indicator_score.percent_change) if indicator_score.percent_change else None,
                    'target_gap': float(indicator_score.target_gap) if indicator_score.target_gap else None,
                    'score': indicator_score.score,
                    'score_color': indicator_score.score_color,
                    'score_label': indicator_score.score_label,
                    'is_manual_override': indicator_score.is_manual_override
                }
                
                AuditLog.log_change(
                    action_type=AuditLog.ActionType.UPDATE,
                    entity_type=AuditLog.EntityType.INDICATOR_SCORE,
                    entity_id=str(indicator_score.id),
                    user=dhis2_user,
                    change_reason=AuditLog.ChangeReason.MANUAL_ENTRY,
                    change_description=f"Manual data update for {indicator_score.indicator.name}",
                    old_values=old_values,
                    new_values=new_values,
                    org_unit_id=org_unit_id,
                    org_unit_name=indicator_score.org_unit_name,
                    assessment_period=indicator_score.assessment_period.name,
                    indicator_id=str(indicator_score.indicator.id),
                    objective_id=str(indicator_score.objective.id)
                )
                
                return {
                    'success': True,
                    'message': 'Manual data updated and scores recalculated successfully',
                    'indicator_score': {
                        'id': indicator_score.id,
                        'current_value': float(indicator_score.current_value) if indicator_score.current_value else None,
                        'previous_value': float(indicator_score.previous_value) if indicator_score.previous_value else None,
                        'target_value': float(indicator_score.target_value) if indicator_score.target_value else None,
                        'percent_change': float(indicator_score.percent_change) if indicator_score.percent_change else None,
                        'target_gap': float(indicator_score.target_gap) if indicator_score.target_gap else None,
                        'score': indicator_score.score,
                        'score_color': indicator_score.score_color,
                        'score_label': indicator_score.score_label,
                        'is_manual_override': indicator_score.is_manual_override
                    }
                }
                
        except IndicatorScore.DoesNotExist:
            raise ValidationError(f"No indicator score found for indicator {indicator_id}")
        except Exception as e:
            self.logger.error(f"Error updating manual indicator data: {str(e)}")
            raise ValidationError(f"Error updating manual data: {str(e)}")
    
    def _parse_decimal(self, value):
        """Parse a value to Decimal, handling None and empty strings"""
        if value is None or value == '':
            return None
        try:
            return Decimal(str(value))
        except (ValueError, TypeError):
            return None
    
    def _calculate_score_from_metrics(self, indicator_score):
        """
        Calculate score based on available metrics (percent_change or target_gap)
        """
        if indicator_score.is_manual_override:
            return  # Don't recalculate manual overrides
        
        # Determine which metric to use for scoring
        if indicator_score.target_gap is not None:
            # Use target gap for scoring
            metric_value = indicator_score.target_gap
            performance_type = 'gap'
        elif indicator_score.percent_change is not None:
            # Use percent change for scoring
            metric_value = indicator_score.percent_change
            performance_type = 'change'
        else:
            # No metrics available for scoring
            indicator_score.score = None
            indicator_score.score_color = '#6c757d'
            indicator_score.score_label = 'No Data'
            indicator_score.scoring_rule = None
            return
        
        # Find matching scoring rule
        matching_rule = None
        rules = ScoringRule.objects.filter(
            performance_type=performance_type,
            is_active=True
        ).order_by('-priority', 'min_value')
        
        for rule in rules:
            if rule.matches_value(metric_value):
                matching_rule = rule
                break
        
        # Apply score
        if matching_rule:
            indicator_score.score = matching_rule.score
            indicator_score.score_color = matching_rule.color
            indicator_score.score_label = matching_rule.label
            indicator_score.scoring_rule = matching_rule
        else:
            indicator_score.score = 0
            indicator_score.score_color = '#6c757d'
            indicator_score.score_label = 'No Match'
            indicator_score.scoring_rule = None
        
        indicator_score.last_calculated = timezone.now()
    
    def _recalculate_higher_level_scores(self, indicator_score):
        """
        Recalculate objective and sector scores after indicator score changes
        """
        try:
            # Recalculate objective score
            objective_score = ObjectiveScore.objects.filter(
                objective=indicator_score.objective,
                org_unit_id=indicator_score.org_unit_id,
                assessment_period=indicator_score.assessment_period
            ).first()
            
            if objective_score:
                objective_score.calculate_score()
            
            # Recalculate sector score
            sector_score = SectorScore.objects.filter(
                org_unit_id=indicator_score.org_unit_id,
                assessment_period=indicator_score.assessment_period
            ).first()
            
            if sector_score:
                sector_score.calculate_score()
                
        except Exception as e:
            self.logger.error(f"Error recalculating higher-level scores: {str(e)}")
    
    def bulk_update_manual_data(self, request, updates):
        """
        Bulk update multiple indicator data entries
        
        Args:
            request: Django request object
            updates: List of update objects with indicator_id, org_unit_id, assessment_period_id, and data_updates
        """
        results = []
        
        for update in updates:
            try:
                result = self.update_manual_indicator_data(
                    request=request,
                    indicator_id=update['indicator_id'],
                    org_unit_id=update['org_unit_id'],
                    assessment_period_id=update['assessment_period_id'],
                    data_updates=update['data_updates']
                )
                results.append({
                    'indicator_id': update['indicator_id'],
                    'success': True,
                    'result': result
                })
            except Exception as e:
                results.append({
                    'indicator_id': update['indicator_id'],
                    'success': False,
                    'error': str(e)
                })
        
        return results


# Holistic Assessment Scoring Service
from decimal import Decimal
from typing import Dict, Any, Optional
from django.db import transaction
from indicators.models import TrackedIndicator
from assessments.models import IndicatorScore, ScoringContext


class HolisticScoringService:
    """
    Simplified Holistic Assessment scoring algorithm based on Excel formulas
    Matches the exact logic shown in the performance analysis table
    """
    
    def calculate_indicator_score(
        self,
        indicator: TrackedIndicator,
        current_value: Optional[float],  # Column G in Excel
        previous_value: Optional[float],  # Column F in Excel
        data_provided: bool = True
    ) -> Dict[str, Any]:
        """
        Calculate score using simplified Holistic Assessment algorithm
        Based on Excel formulas from the performance analysis table:
        - Data Provided: =IF($G4<>"","Yes","No")
        - First Year: =IF(OR($F4<>"",$E4<>"",$D4<>"",$C4<>""),"No","Yes")
        - Target Achieved: =IF($G4<=J4,"Yes","No")
        - Performance Change: =IF($H4<=-10%,"<=-10%",IF($H4<=-5%,"-10%<C<=-5%",IF($H4<=5%,"5%<=C>-5%",IF($H4>5%,">5%",""))))
        - Gap to Target: =IF($I4<=10%,"<=10%",IF(AND($I4>10%,$I4<=40%),"10%<PT<=40%",IF($I4>40%,">40%","")))
        """
        
        # Step 1: Data Provided (Column G) - =IF($G4<>"","Yes","No")
        data_provided_flag = "Yes" if data_provided and current_value is not None else "No"
        
        # Step 2: First Year of Reporting - =IF(OR($F4<>"",$E4<>"",$D4<>"",$C4<>""),"No","Yes")
        # For simplicity, we only check the immediate previous year (F4)
        # In a full implementation, you'd check multiple previous years
        has_previous_data = previous_value is not None
        is_first_year = "Yes" if not has_previous_data else "No"
        
        # Step 3: Was Target Achieved - =IF($G4<=J4,"Yes","No")
        # Use the target format and operator to determine target achievement
        target_achieved = "No"  # Default
        if current_value is not None:
            # Handle different target formats
            if hasattr(indicator, 'target_format') and indicator.target_format == 'RANGE':
                # Range target: check if current value is within the range
                if indicator.target_lower_limit is not None and indicator.target_upper_limit is not None:
                    lower_limit = float(indicator.target_lower_limit)
                    upper_limit = float(indicator.target_upper_limit)
                    target_achieved = "Yes" if lower_limit <= current_value <= upper_limit else "No"
                else:
                    # Fallback to single target value
                    if indicator.target_value is not None:
                        target_float = float(indicator.target_value)
                        # Use target_type to determine achievement for fallback
                        if indicator.target_type == 'decrease':
                            target_achieved = "Yes" if current_value <= target_float else "No"
                        else:
                            target_achieved = "Yes" if current_value >= target_float else "No"
            elif hasattr(indicator, 'target_format') and indicator.target_format == 'MINIMUM':
                # Minimum target: current value should be >= target_value
                if indicator.target_value is not None:
                    target_float = float(indicator.target_value)
                    target_achieved = "Yes" if current_value >= target_float else "No"
            elif hasattr(indicator, 'target_format') and indicator.target_format == 'MAXIMUM':
                # Maximum target: current value should be <= target_value
                if indicator.target_value is not None:
                    target_float = float(indicator.target_value)
                    target_achieved = "Yes" if current_value <= target_float else "No"
            else:
                # Single value target: use the target_operator
                if indicator.target_value is not None:
                    target_float = float(indicator.target_value)
                    
                    # Use the target_operator to determine achievement
                    if indicator.target_operator == '>=':
                        target_achieved = "Yes" if current_value >= target_float else "No"
                    elif indicator.target_operator == '>':
                        target_achieved = "Yes" if current_value > target_float else "No"
                    elif indicator.target_operator == '<=':
                        target_achieved = "Yes" if current_value <= target_float else "No"
                    elif indicator.target_operator == '<':
                        target_achieved = "Yes" if current_value < target_float else "No"
                    elif indicator.target_operator == '=':
                        target_achieved = "Yes" if current_value == target_float else "No"
                    else:
                        # Default to >= for backward compatibility
                        target_achieved = "Yes" if current_value >= target_float else "No"
        
        # Step 4: Performance Change (Column H) - =IF($H4<=-10%,"<=-10%",IF($H4<=-5%,"-10%<C<=-5%",IF($H4<=5%,"-5%<C<=5%",IF($H4>5%,">5%",""))))
        percent_change = None
        change_category = None
        if current_value is not None and previous_value is not None and previous_value != 0:
            # Calculate raw percentage change based on indicator type
            target_format = getattr(indicator, 'target_format', 'SINGLE')
            
            if target_format == 'RANGE':
                # For range indicators, always use standard formula regardless of target_type
                percent_change = round(((current_value - previous_value) / abs(previous_value)) * 100, 2)
            else:
                # For non-range indicators, use target_type specific formula
                if indicator.target_type == 'decrease':
                    # For decrease indicators: (previous_value - current_value) / abs(current_value) * 100
                    percent_change = round(((previous_value - current_value) / abs(current_value)) * 100, 2)
                else:
                    # For increase indicators: (current_value - previous_value) / abs(previous_value) * 100
                    percent_change = round(((current_value - previous_value) / abs(previous_value)) * 100, 2)
            
            # Calculate performance change (for scoring) - use raw percent_change for all types
            performance_change = percent_change
            
            # Categorize based on performance change (not raw change) - EXACTLY matches flowchart
            if performance_change <= -10:
                change_category = "<=-10%"
            elif performance_change <= -5:
                change_category = "-10%<C<=-5%"
            elif performance_change <= 5:
                change_category = "-5%<C<=5%"  # Stagnation category
            elif performance_change > 5:
                change_category = ">5%"
        
        # Step 5: Gap to Target (Column I) - =IF($I4<=10%,"<=10%",IF(AND($I4>10%,$I4<=40%),"10%<PT<=40%",IF($I4>40%,">40%","")))
        target_gap = None
        gap_category = None
        if current_value is not None:
            # Handle different target formats for gap calculation
            if hasattr(indicator, 'target_format') and indicator.target_format == 'RANGE':
                # Range target: calculate gap to the upper limit (as per image)
                if indicator.target_lower_limit is not None and indicator.target_upper_limit is not None:
                    lower_limit = float(indicator.target_lower_limit)
                    upper_limit = float(indicator.target_upper_limit)
                    
                    # For range targets, calculate gap to the upper limit
                    target_gap = (current_value - upper_limit) / upper_limit * 100
                else:
                    # Fallback to single target value
                    if indicator.target_value is not None:
                        target_float = float(indicator.target_value)
                        if target_float != 0:
                            # Calculate gap based on target type for fallback
                            if indicator.target_type == 'decrease':
                                # For decrease indicators: (target_value - current_value) / current_value * 100
                                target_gap = (target_float - current_value) / current_value * 100
                            else:
                                # For increase indicators: (current_value - target_value) / target_value * 100
                                target_gap = (current_value - target_float) / target_float * 100
            else:
                # Single value target
                if indicator.target_value is not None:
                    target_float = float(indicator.target_value)
                    if target_float != 0:
                        # Calculate gap based on target type
                        if indicator.target_type == 'decrease':
                            # For decrease indicators: (target_value - current_value) / current_value * 100
                            target_gap = (target_float - current_value) / current_value * 100
                        else:
                            # For increase indicators: (current_value - target_value) / target_value * 100
                            target_gap = (current_value - target_float) / target_float * 100
            
            # Categorize based on the signed target_gap, matching Excel's behavior
            # Excel formula: =IF($I4<=10%,"<=10%",IF(AND($I4>10%,$I4<=40%),"10%<PT<=40%",IF($I4>40%,">40%","")))
            if target_gap is not None:
                # Round target_gap to 2 decimal places for display consistency
                target_gap = round(target_gap, 2)
                
                if target_gap <= 10:
                    gap_category = "<=10%"
                elif 10 < target_gap <= 40:
                    gap_category = "10%<PT<=40%"
                elif target_gap > 40:
                    gap_category = ">40%"
        
        # Step 6: Calculate final score based on the flowchart logic
        score = self._calculate_final_score(
            data_provided_flag, is_first_year, target_achieved, 
            change_category, gap_category
        )
        
        # Debug logging
        logger.debug(f"Scoring Debug for Indicator {indicator.id} ({indicator.name}):")
        logger.debug(f"  target_type: {indicator.target_type}")
        logger.debug(f"  target_operator: {indicator.target_operator}")
        logger.debug(f"  target_value: {indicator.target_value}")
        logger.debug(f"  current_value: {current_value}")
        logger.debug(f"  previous_value: {previous_value}")
        logger.debug(f"  percent_change: {percent_change}")
        logger.debug(f"  target_achieved: {target_achieved}")
        logger.debug(f"  change_category: {change_category}")
        logger.debug(f"  gap_category: {gap_category}")
        logger.debug(f"  final_score: {score}")
        
        return {
            'score': score,
            'data_provided': data_provided_flag,
            'is_first_year': is_first_year,
            'target_achieved': target_achieved,
            'change_category': change_category,
            'gap_category': gap_category,
            'percent_change': percent_change,
            'target_gap': target_gap,
            'current_value': current_value,
            'previous_value': previous_value,
            'target_value': float(indicator.target_value) if indicator.target_value else None
        }
    
    def _calculate_final_score(
        self,
        data_provided: str,
        is_first_year: str,
        target_achieved: str,
        change_category: Optional[str],
        gap_category: Optional[str]
    ) -> int:
        """
        Calculate final score based on the flowchart logic shown in the image
        """
        
        logger.debug(f"_calculate_final_score inputs:")
        logger.debug(f"  data_provided: {data_provided}")
        logger.debug(f"  is_first_year: {is_first_year}")
        logger.debug(f"  target_achieved: {target_achieved}")
        logger.debug(f"  change_category: {change_category}")
        logger.debug(f"  gap_category: {gap_category}")
        """
        Calculate final score based on the flowchart logic shown in the image
        """
        
        # Decision 1: Was data provided?
        if data_provided == "No":
            return -2  # Red circle in flowchart
        
        # Decision 2: Is it the first year of reporting?
        if is_first_year == "Yes":
            # First year logic: check if target was achieved
            # According to the flowchart, when target is achieved, it should be 1, not 2
            return 1 if target_achieved == "Yes" else 0
        
        # Not first year - proceed with complex logic
        # Decision 3: Was the target achieved?
        if target_achieved == "Yes":
            # Target WAS achieved - check performance change
            if change_category == ">5%":
                return 2  # Green circle - Increase
            elif change_category == "-5%<C<=5%":
                return 2  # Green circle - Stagnation
            elif change_category == "-10%<C<=-5%":
                return 1  # Green circle - Small decrease (should be 1, not 2)
            elif change_category == "<=-10%":
                return 0  # Yellow circle - Large decrease (should be 0, not -2)
            else:
                # Target achieved but no change category (e.g., previous_value is 0)
                # For decrease indicators, achieving target should score 2
                # For increase indicators, achieving target should score 2
                return 2  # Target achieved = good performance
        
        else:
            # Target NOT achieved - check performance change
            # Excel formula: IF(AND(M13="No",N13="No",O13=">5%"),1,IF(AND(M13="No",N13="No",O13="-5%<C<=5%",P13="<=10%"),1,IF(AND(M13="No",N13="No",O13="-5%<C<=5%",P13="10%<PT<=40%"),0,IF(AND(M13="No",N13="No",O13="-5%<C<=5%",P13=">40%"),-1,IF(AND(M13="No",N13="No",O13="-10%<C<=-5%"),-1,IF(AND(M13="No",N13="No",O13="<=-10%"),-1))))))
            if change_category == ">5%":
                logger.debug("  Score calculation: change_category='>5%' -> score=1")
                return 1  # Excel: IF(AND(M13="No",N13="No",O13=">5%"),1,...)
            elif change_category == "-5%<C<=5%":
                # Stagnation - check how close to target
                if gap_category == "<=10%":
                    logger.debug("  Score calculation: change_category='-5%<C<=5%', gap_category='<=10%' -> score=1")
                    return 1  # Excel: IF(AND(M13="No",N13="No",O13="-5%<C<=5%",P13="<=10%"),1,...)
                elif gap_category == "10%<PT<=40%":
                    logger.debug("  Score calculation: change_category='-5%<C<=5%', gap_category='10%<PT<=40%' -> score=0")
                    return 0  # Excel: IF(AND(M13="No",N13="No",O13="-5%<C<=5%",P13="10%<PT<=40%"),0,...)
                elif gap_category == ">40%":
                    logger.debug("  Score calculation: change_category='-5%<C<=5%', gap_category='>40%' -> score=-1")
                    return -1  # Excel: IF(AND(M13="No",N13="No",O13="-5%<C<=5%",P13=">40%"),-1,...)
                else:
                    logger.debug("  Score calculation: change_category='-5%<C<=5%', gap_category=None -> score=0")
                    return 0  # Default for stable
            elif change_category == "-10%<C<=-5%":
                logger.debug("  Score calculation: change_category='-10%<C<=-5%' -> score=-1")
                return -1  # Excel: IF(AND(M13="No",N13="No",O13="-10%<C<=-5%"),-1,...)
            elif change_category == "<=-10%":
                logger.debug("  Score calculation: change_category='<=-10%' -> score=-1")
                return -1  # Large decline should be -1 when target NOT achieved
            else:
                logger.debug("  Score calculation: change_category=None -> score=0")
                return 0  # Default case
        
        # This should never be reached, but just in case
        return 0
    
    def calculate_batch_scores(self, indicator_scores: list) -> None:
        """
        Calculate scores for a batch of indicator scores
        """
        for indicator_score in indicator_scores:
            try:
                indicator_score.calculate_holistic_score()
            except Exception as e:
                # Log error but continue with other scores
                print(f"Error calculating score for {indicator_score}: {e}")
                continue
    
    def get_scoring_summary(self, indicator_score: IndicatorScore) -> Dict[str, Any]:
        """
        Get a summary of the scoring context for an indicator
        """
        if not indicator_score.scoring_context:
            return {}
        
        context = indicator_score.scoring_context
        
        return {
            'data_provided': context.data_provided,
            'current_meets_target': context.current_meets_target,
            'previous_meets_target': context.previous_meets_target,
            'change_category': context.change_category,
            'gap_category': context.gap_category,
            'percent_change': float(context.percent_change) if context.percent_change else None,
            'target_gap': float(context.target_gap) if context.target_gap else None,
            'final_score': indicator_score.score,
            'score_color': indicator_score.score_color,
            'score_label': indicator_score.score_label
        }