"""
Excel Export Service

This module handles Excel export functionality for assessments and reports.
"""

import io
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from django.http import HttpResponse
from django.conf import settings

logger = logging.getLogger(__name__)


class ExcelExportService:
    """
    Service for exporting assessment data to Excel format.
    
    This service handles the creation of Excel workbooks with formatted
    assessment data, scores, and reports.
    """
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
        # Define styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.subheader_font = Font(bold=True, color="FFFFFF")
        self.subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
    
    def export_assessment_to_excel(self, assessment_data: Dict[str, Any], 
                                 filename: Optional[str] = None) -> HttpResponse:
        """
        Export assessment data to Excel format.
        
        Args:
            assessment_data: Complete assessment data including scores and indicators
            filename: Optional filename for the Excel file
            
        Returns:
            HttpResponse with Excel file attachment
        """
        try:
            workbook = Workbook()
            workbook.remove(workbook.active)  # Remove default sheet
            
            # Create main assessment sheet
            self._create_assessment_sheet(workbook, assessment_data)
            
            # Create detailed scores sheet
            self._create_scores_sheet(workbook, assessment_data)
            
            # Create summary sheet
            self._create_summary_sheet(workbook, assessment_data)
            
            # Generate filename if not provided
            if not filename:
                org_unit = assessment_data.get('org_unit', {}).get('name', 'Unknown')
                period = assessment_data.get('period', 'Unknown')
                filename = f"Assessment_{org_unit}_{period}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            # Save workbook to response
            workbook.save(response)
            return response
            
        except Exception as e:
            self.logger.error(f"Error exporting assessment to Excel: {str(e)}")
            raise
    
    def _create_assessment_sheet(self, workbook: Workbook, assessment_data: Dict[str, Any]) -> None:
        """Create the main assessment sheet with all indicator data."""
        ws = workbook.create_sheet("Assessment Data")
        
        # Add title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"Holistic Assessment Report - {assessment_data.get('org_unit', {}).get('name', 'Unknown')}"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = self.center_alignment
        
        # Add metadata
        metadata_row = 3
        ws[f'A{metadata_row}'] = "Organization Unit:"
        ws[f'B{metadata_row}'] = assessment_data.get('org_unit', {}).get('name', 'N/A')
        ws[f'C{metadata_row}'] = "Period:"
        ws[f'D{metadata_row}'] = assessment_data.get('period', 'N/A')
        ws[f'E{metadata_row}'] = "Generated:"
        ws[f'F{metadata_row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Add headers
        headers = [
            'Indicator Name', 'Current Value', 'Previous Value', 'Target Value',
            'Percentage Change', 'Target Gap', 'Score', 'Category'
        ]
        
        header_row = 5
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment
        
        # Add data
        indicators = assessment_data.get('indicators', [])
        data_row = header_row + 1
        
        for indicator in indicators:
            ws.cell(row=data_row, column=1, value=indicator.get('name', 'N/A'))
            ws.cell(row=data_row, column=2, value=indicator.get('current_value', 'N/A'))
            ws.cell(row=data_row, column=3, value=indicator.get('previous_value', 'N/A'))
            ws.cell(row=data_row, column=4, value=indicator.get('target_value', 'N/A'))
            ws.cell(row=data_row, column=5, value=indicator.get('percentage_change', 'N/A'))
            ws.cell(row=data_row, column=6, value=indicator.get('target_gap', 'N/A'))
            ws.cell(row=data_row, column=7, value=indicator.get('score', 'N/A'))
            ws.cell(row=data_row, column=8, value=indicator.get('category', 'N/A'))
            
            # Apply borders to all cells in the row
            for col in range(1, 9):
                ws.cell(row=data_row, column=col).border = self.border
            
            data_row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_scores_sheet(self, workbook: Workbook, assessment_data: Dict[str, Any]) -> None:
        """Create detailed scores sheet with scoring breakdown."""
        ws = workbook.create_sheet("Detailed Scores")
        
        # Add title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = "Detailed Scoring Breakdown"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = self.center_alignment
        
        # Add headers
        headers = [
            'Indicator', 'Current Score', 'Previous Score', 'Score Change',
            'Weight', 'Weighted Score'
        ]
        
        header_row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment
        
        # Add data
        indicators = assessment_data.get('indicators', [])
        data_row = header_row + 1
        
        for indicator in indicators:
            ws.cell(row=data_row, column=1, value=indicator.get('name', 'N/A'))
            ws.cell(row=data_row, column=2, value=indicator.get('score', 'N/A'))
            ws.cell(row=data_row, column=3, value=indicator.get('previous_score', 'N/A'))
            ws.cell(row=data_row, column=4, value=indicator.get('score_change', 'N/A'))
            ws.cell(row=data_row, column=5, value=indicator.get('weight', 'N/A'))
            ws.cell(row=data_row, column=6, value=indicator.get('weighted_score', 'N/A'))
            
            # Apply borders
            for col in range(1, 7):
                ws.cell(row=data_row, column=col).border = self.border
            
            data_row += 1
        
        # Add total row
        total_row = data_row + 1
        ws.cell(row=total_row, column=1, value="TOTAL")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=assessment_data.get('total_score', 'N/A'))
        ws.cell(row=total_row, column=6).font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_summary_sheet(self, workbook: Workbook, assessment_data: Dict[str, Any]) -> None:
        """Create summary sheet with key metrics and charts."""
        ws = workbook.create_sheet("Summary")
        
        # Add title
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = "Assessment Summary"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = self.center_alignment
        
        # Add key metrics
        metrics_row = 3
        metrics = [
            ("Organization Unit", assessment_data.get('org_unit', {}).get('name', 'N/A')),
            ("Assessment Period", assessment_data.get('period', 'N/A')),
            ("Total Indicators", len(assessment_data.get('indicators', []))),
            ("Overall Score", f"{assessment_data.get('total_score', 'N/A')}%"),
            ("Assessment Date", datetime.now().strftime('%Y-%m-%d')),
            ("Generated By", assessment_data.get('user', 'System'))
        ]
        
        for i, (label, value) in enumerate(metrics):
            ws.cell(row=metrics_row + i, column=1, value=label)
            ws.cell(row=metrics_row + i, column=1).font = Font(bold=True)
            ws.cell(row=metrics_row + i, column=2, value=value)
            ws.cell(row=metrics_row + i, column=2).border = self.border
        
        # Add category breakdown
        category_start_row = metrics_row + len(metrics) + 2
        ws.cell(row=category_start_row, column=1, value="Category Breakdown")
        ws.cell(row=category_start_row, column=1).font = Font(bold=True, size=14)
        
        # Count categories
        categories = {}
        indicators = assessment_data.get('indicators', [])
        for indicator in indicators:
            category = indicator.get('category', 'Unknown')
            categories[category] = categories.get(category, 0) + 1
        
        # Add category data
        category_data_row = category_start_row + 1
        ws.cell(row=category_data_row, column=1, value="Category")
        ws.cell(row=category_data_row, column=2, value="Count")
        ws.cell(row=category_data_row, column=1).font = self.header_font
        ws.cell(row=category_data_row, column=1).fill = self.header_fill
        ws.cell(row=category_data_row, column=2).font = self.header_font
        ws.cell(row=category_data_row, column=2).fill = self.header_fill
        
        category_data_row += 1
        for category, count in categories.items():
            ws.cell(row=category_data_row, column=1, value=category)
            ws.cell(row=category_data_row, column=2, value=count)
            category_data_row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def export_indicators_template(self, indicators: List[Dict[str, Any]], 
                                 filename: Optional[str] = None) -> HttpResponse:
        """
        Export indicators template for manual data entry.
        
        Args:
            indicators: List of indicators to include in template
            filename: Optional filename for the Excel file
            
        Returns:
            HttpResponse with Excel file attachment
        """
        try:
            workbook = Workbook()
            ws = workbook.active
            ws.title = "Indicators Template"
            
            # Add title
            ws.merge_cells('A1:D1')
            title_cell = ws['A1']
            title_cell.value = "Indicators Data Entry Template"
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = self.center_alignment
            
            # Add instructions
            ws['A3'] = "Instructions:"
            ws['A3'].font = Font(bold=True)
            ws['A4'] = "1. Fill in the Current Value column with actual data"
            ws['A5'] = "2. Fill in the Previous Value column with historical data"
            ws['A6'] = "3. Fill in the Target Value column with target data"
            ws['A7'] = "4. Save and upload this file to import the data"
            
            # Add headers
            headers = [
                'Indicator UID', 'Indicator Name', 'Current Value', 'Previous Value', 'Target Value'
            ]
            
            header_row = 9
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = self.center_alignment
            
            # Add indicator data
            data_row = header_row + 1
            for indicator in indicators:
                ws.cell(row=data_row, column=1, value=indicator.get('uid', ''))
                ws.cell(row=data_row, column=2, value=indicator.get('name', ''))
                ws.cell(row=data_row, column=3, value='')  # Current Value (to be filled)
                ws.cell(row=data_row, column=4, value='')  # Previous Value (to be filled)
                ws.cell(row=data_row, column=5, value='')  # Target Value (to be filled)
                
                # Apply borders
                for col in range(1, 6):
                    ws.cell(row=data_row, column=col).border = self.border
                
                data_row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Generate filename if not provided
            if not filename:
                filename = f"Indicators_Template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            # Save workbook to response
            workbook.save(response)
            return response
            
        except Exception as e:
            self.logger.error(f"Error exporting indicators template: {str(e)}")
            raise
    
    def export_dashboard_data(self, dashboard_data: Dict[str, Any], 
                            filename: Optional[str] = None) -> HttpResponse:
        """
        Export dashboard data to Excel format.
        
        Args:
            dashboard_data: Dashboard data including charts and metrics
            filename: Optional filename for the Excel file
            
        Returns:
            HttpResponse with Excel file attachment
        """
        try:
            workbook = Workbook()
            workbook.remove(workbook.active)  # Remove default sheet
            
            # Create metrics sheet
            self._create_metrics_sheet(workbook, dashboard_data)
            
            # Create trends sheet
            self._create_trends_sheet(workbook, dashboard_data)
            
            # Create comparisons sheet
            self._create_comparisons_sheet(workbook, dashboard_data)
            
            # Generate filename if not provided
            if not filename:
                filename = f"Dashboard_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            # Save workbook to response
            workbook.save(response)
            return response
            
        except Exception as e:
            self.logger.error(f"Error exporting dashboard data: {str(e)}")
            raise
    
    def _create_metrics_sheet(self, workbook: Workbook, dashboard_data: Dict[str, Any]) -> None:
        """Create metrics sheet with key performance indicators."""
        ws = workbook.create_sheet("Key Metrics")
        
        # Add title
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = "Key Performance Metrics"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = self.center_alignment
        
        # Add metrics
        metrics = dashboard_data.get('metrics', {})
        row = 3
        
        for metric_name, metric_value in metrics.items():
            ws.cell(row=row, column=1, value=metric_name)
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=metric_value)
            ws.cell(row=row, column=2).border = self.border
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_trends_sheet(self, workbook: Workbook, dashboard_data: Dict[str, Any]) -> None:
        """Create trends sheet with historical data."""
        ws = workbook.create_sheet("Trends")
        
        # Add title
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = "Historical Trends"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = self.center_alignment
        
        # Add trend data
        trends = dashboard_data.get('trends', [])
        if trends:
            # Add headers
            headers = ['Period', 'Score', 'Change']
            header_row = 3
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = self.center_alignment
            
            # Add data
            data_row = header_row + 1
            for trend in trends:
                ws.cell(row=data_row, column=1, value=trend.get('period', ''))
                ws.cell(row=data_row, column=2, value=trend.get('score', ''))
                ws.cell(row=data_row, column=3, value=trend.get('change', ''))
                
                # Apply borders
                for col in range(1, 4):
                    ws.cell(row=data_row, column=col).border = self.border
                
                data_row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_comparisons_sheet(self, workbook: Workbook, dashboard_data: Dict[str, Any]) -> None:
        """Create comparisons sheet with organization unit comparisons."""
        ws = workbook.create_sheet("Comparisons")
        
        # Add title
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = "Organization Unit Comparisons"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = self.center_alignment
        
        # Add comparison data
        comparisons = dashboard_data.get('comparisons', [])
        if comparisons:
            # Add headers
            headers = ['Organization Unit', 'Score', 'Rank']
            header_row = 3
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = self.center_alignment
            
            # Add data
            data_row = header_row + 1
            for comparison in comparisons:
                ws.cell(row=data_row, column=1, value=comparison.get('org_unit', ''))
                ws.cell(row=data_row, column=2, value=comparison.get('score', ''))
                ws.cell(row=data_row, column=3, value=comparison.get('rank', ''))
                
                # Apply borders
                for col in range(1, 4):
                    ws.cell(row=data_row, column=col).border = self.border
                
                data_row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
