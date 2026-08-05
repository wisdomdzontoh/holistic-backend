"""
Real-time DHIS2 Service

This module handles real-time data fetching from DHIS2 without persistent storage.
Uses the same synchronous approach as the working old_services.py
"""

import logging
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed, wait
from typing import Dict, List, Any, Optional
from datetime import datetime, timedelta
from django.conf import settings
from django.core.cache import cache
from django.utils import timezone

from ..models import TrackedIndicator
from .validation_service import ValidationService
from .cache_service import CacheService
from .data_processing_service import DataProcessingService
from .period_service import PeriodService
from dhis2_auth.dhis_client import DHIS2ClientFactory
from dhis2_auth.session import get_dhis2_session_data

logger = logging.getLogger(__name__)


class RealTimeDHIS2Service:
    """
    Service for real-time DHIS2 data fetching without database storage
    Uses the same synchronous approach as the working old_services.py
    """

    # Batched DHIS2 analytics responses are cached briefly so a re-export minutes
    # later (common during a working session) doesn't re-hit DHIS2 at all.
    DHIS2_BATCH_CACHE_TIMEOUT = 60 * 10  # 10 minutes

    # Upper bound on (indicators x periods) "cells" requested in a single DHIS2
    # analytics call. DHIS2's own server has to compute aggregates for every
    # cell, and a big enough request can exceed DHIS2's Cloudflare edge timeout
    # (~100s, outside our control) even though it's nowhere near any URL length
    # limit. Tuned conservatively; _fetch_chunk_with_backoff adapts below this
    # anyway if a particular request still turns out to be too expensive.
    MAX_BATCH_CELLS = 60
    # Stop auto-splitting a failing chunk below this size - let the normal
    # per-indicator fallback (with its full period-format retry logic) handle
    # whatever's left rather than recursing forever.
    MIN_BATCH_CHUNK_SIZE = 3

    def __init__(self, dhis2_client=None):
        self.client = dhis2_client
        self.logger = logging.getLogger(__name__)
        self.validation_service = ValidationService()
        self.cache_service = CacheService()
        self.data_processor = DataProcessingService()
        self.period_service = PeriodService()
    
    def _compute_objective_trend_from_indicators(self, indicators: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Compute objective-level trend from indicator data.
        
        Args:
            indicators: List of indicator data dictionaries
            
        Returns:
            Dictionary containing trend metadata
        """
        try:
            if not indicators:
                return {}
            
            # Calculate average scores and trends
            total_score = 0.0
            total_trend = 0.0
            valid_indicators = 0
            
            for indicator in indicators:
                if indicator.get('score') is not None:
                    total_score += indicator['score']
                    valid_indicators += 1
                
                # Extract trend information if available
                trend_info = indicator.get('trend', {})
                if isinstance(trend_info, dict) and trend_info.get('score') is not None:
                    total_trend += trend_info['score']
            
            if valid_indicators == 0:
                return {}
            
            avg_score = total_score / valid_indicators
            avg_trend = total_trend / valid_indicators if total_trend > 0 else 0.0
            
            # Determine trend direction
            if avg_trend > 0.5:
                trend_direction = 'improving'
            elif avg_trend < -0.5:
                trend_direction = 'declining'
            else:
                trend_direction = 'stable'
            
            return {
                'average_score': avg_score,
                'trend_score': avg_trend,
                'trend_direction': trend_direction,
                'indicator_count': valid_indicators
            }
            
        except Exception as e:
            self.logger.error(f"Error computing objective trend: {str(e)}")
            return {}
    
    def fetch_holistic_assessment_data(self, request, assessment_config: Dict[str, Any]) -> Dict[str, Any]:
        """
        Fetch holistic assessment data from DHIS2.
        
        Args:
            request: Django request object
            assessment_config: Assessment configuration data
            
        Returns:
            Dictionary containing assessment data
        """
        try:
            # Add debug logging
            self.logger.info(f"Received assessment_config: {assessment_config}")
            
            # Get DHIS2 user from session
            from dhis2_auth.session import get_dhis2_user_from_request
            dhis2_user = get_dhis2_user_from_request(request)
            if not dhis2_user:
                raise ValueError("No DHIS2 user found in session")
            
            # Extract configuration - handle both old and new payload formats
            org_unit_id = None
            period = None
            
            # Handle org_unit_ids array format
            if 'org_unit_ids' in assessment_config and assessment_config['org_unit_ids']:
                org_unit_id = assessment_config['org_unit_ids'][0]  # Take first org unit
                self.logger.info(f"Extracted org_unit_id from org_unit_ids: {org_unit_id}")
            
            # Handle org_unit_id string format (backward compatibility)
            elif 'org_unit_id' in assessment_config:
                org_unit_id = assessment_config['org_unit_id']
                self.logger.info(f"Extracted org_unit_id from org_unit_id: {org_unit_id}")
            
            # Handle periods array format
            if 'periods' in assessment_config and assessment_config['periods']:
                period_obj = assessment_config['periods'][0]  # Take first period
                self.logger.info(f"Processing period object: {period_obj}")
                
                # Check if period_obj is already a string (processed by serializer)
                if isinstance(period_obj, str):
                    period = period_obj
                    self.logger.info(f"Using period string directly: {period}")
                else:
                    # Convert period object to string format
                    if 'start_date' in period_obj and 'end_date' in period_obj:
                        # Convert date range to period string (e.g., "2023Q1")
                        start_date = period_obj['start_date']
                        end_date = period_obj['end_date']
                        # Extract year and quarter from dates
                        from datetime import datetime
                        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                        
                        # Determine quarter based on end date
                        if end_dt.month <= 3:
                            quarter = "Q1"
                        elif end_dt.month <= 6:
                            quarter = "Q2"
                        elif end_dt.month <= 9:
                            quarter = "Q3"
                        else:
                            quarter = "Q4"
                        
                        period = f"{start_dt.year}{quarter}"
                        self.logger.info(f"Converted date range to period: {period}")
                    elif 'code' in period_obj:
                        period = period_obj['code']
                        self.logger.info(f"Using period code: {period}")
                    elif 'name' in period_obj:
                        period = period_obj['name']
                        self.logger.info(f"Using period name: {period}")
            
            # Handle period string format (backward compatibility)
            elif 'period' in assessment_config:
                period = assessment_config['period']
                self.logger.info(f"Using direct period: {period}")
            
            # If we still don't have a period, try to construct one from the original request data
            if not period and hasattr(request, 'data'):
                original_data = request.data
                self.logger.info(f"Checking original request data: {original_data}")
                
                if 'periods' in original_data and original_data['periods']:
                    original_period = original_data['periods'][0]
                    self.logger.info(f"Found original period: {original_period}")
                    
                    if isinstance(original_period, dict) and 'start_date' in original_period and 'end_date' in original_period:
                        # Convert date range to period string
                        start_date = original_period['start_date']
                        end_date = original_period['end_date']
                        from datetime import datetime
                        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                        
                        # Determine quarter based on end date
                        if end_dt.month <= 3:
                            quarter = "Q1"
                        elif end_dt.month <= 6:
                            quarter = "Q2"
                        elif end_dt.month <= 9:
                            quarter = "Q3"
                        else:
                            quarter = "Q4"
                        
                        period = f"{start_dt.year}{quarter}"
                        self.logger.info(f"Constructed period from original data: {period}")
            
            self.logger.info(f"Final extracted values - org_unit_id: {org_unit_id}, period: {period}")
            
            if not org_unit_id or not period:
                raise ValueError(f"org_unit_id and period are required. Received: org_unit_id={org_unit_id}, period={period}")
            
            # Initialize client if not provided
            if not self.client:
                self.client = DHIS2ClientFactory.create_client_from_session(
                    dhis2_user.dhis2_instance_url,
                    request.session.session_key
                )
            
            # Extract configuration
            org_unit_ids = assessment_config.get('org_unit_ids', [])
            org_unit_names = assessment_config.get('org_unit_names', [])
            periods_raw = assessment_config.get('periods', [])
            indicator_uids = assessment_config.get('indicator_uids', [])
            manual_entries = assessment_config.get('manual_entries', {})
            pre_calculated_scores = assessment_config.get('pre_calculated_scores', {})
            self.logger.info(f"Org unit names received: {org_unit_names}")
            
            # Handle periods - they can be strings or objects with 'code' field
            periods = []
            for period in periods_raw:
                if isinstance(period, dict) and 'code' in period:
                    periods.append(period['code'])
                elif isinstance(period, str):
                    periods.append(period)
                else:
                    # Fallback: try to extract code from period object
                    period_code = getattr(period, 'code', str(period))
                    periods.append(period_code)
            
            # Sort periods chronologically to ensure correct change calculation
            periods.sort()
            
            if not org_unit_ids or not periods:
                raise ValueError("Organization units and periods are required")
            
            # Fetch active indicators if not specified; include manual indicators (no dhis2_uid)
            manual_indicators = []
            if not indicator_uids:
                indicators = TrackedIndicator.objects.filter(is_active=True)
                # Only include non-empty UIDs for DHIS2 fetch; keep track of manual ones
                indicator_uids = [ind.dhis2_uid for ind in indicators if ind.dhis2_uid]
                manual_indicators = [ind for ind in indicators if not ind.dhis2_uid]

            # Batch-fetch DHIS2 data for every (indicator, period) pair up front in a
            # handful of HTTP calls instead of one call per pair. The per-pair fetch
            # loops below still run, but they hit this pre-built lookup first and only
            # fall back to the slow single-call path for anything not covered here
            # (reporting-rate indicators, dataSet-type indicators, or a genuine miss).
            dhis2_indicators_for_batch = list(TrackedIndicator.objects.filter(
                dhis2_uid__in=indicator_uids, is_active=True
            ))
            dhis2_value_lookup = self._fetch_batch_indicator_data(
                dhis2_indicators_for_batch, org_unit_ids[0], periods
            )
            # Resolve everything the batch step missed concurrently too, so the
            # per-indicator assembly loops below never block on network I/O.
            dhis2_resolved_misses = self._prefetch_remaining_indicator_values(
                dhis2_indicators_for_batch, org_unit_ids[0], periods, dhis2_value_lookup
            )

            # Fetch data for each indicator
            assessment_data = {
                'indicators': [],
                'objectives': [],
                'milestones': [],
                'metadata': {
                    'org_units': org_unit_ids,
                    'periods': periods,
                    'fetched_at': timezone.now().isoformat()
                }
            }
            
            # Group indicators by objective
            from configurations.models import Objective
            objectives = Objective.objects.filter(is_active=True).prefetch_related('indicator_weights__indicator')
            
            # Check if we have indicator weights configured
            total_weights = sum(obj.indicator_weights.count() for obj in objectives)
            
            if total_weights == 0:
                # No indicator weights configured - fetch indicators directly and group them evenly
                all_indicators = list(TrackedIndicator.objects.filter(
                    dhis2_uid__in=indicator_uids,
                    is_active=True
                ))
                
                # Distribute indicators evenly across objectives
                indicators_per_objective = len(all_indicators) // max(len(objectives), 1)
                indicator_list = list(all_indicators)
                
                for i, objective in enumerate(objectives):
                    objective_data = {
                        'id': objective.id,
                        'name': objective.name,
                        'code': objective.code,
                        'color': objective.color,
                        'milestone': None,
                        'indicators': []
                    }
                    
                    # Assign indicators to this objective
                    start_idx = i * indicators_per_objective
                    end_idx = start_idx + indicators_per_objective if i < len(objectives) - 1 else len(indicator_list)
                    objective_indicators = indicator_list[start_idx:end_idx]
                    
                    for indicator in objective_indicators:
                        indicator_data = {
                            'id': indicator.id,
                            'name': indicator.name,
                            'dhis2_uid': indicator.dhis2_uid,
                            'description': indicator.description,
                            'indicator_number': indicator.indicator_number or f"{i+1}",
                            'display_order': indicator.display_order,
                            'target_value': float(indicator.target_value) if indicator.target_value else None,
                            'target_display': indicator.target_display,
                            'target_lower_limit': float(indicator.target_lower_limit) if indicator.target_lower_limit else None,
                            'target_upper_limit': float(indicator.target_upper_limit) if indicator.target_upper_limit else None,
                            'target_format': getattr(indicator, 'target_format', 'SINGLE'),
                            'target_type': indicator.target_type,
                            'weight': 1.0,  # Default weight when no indicator weights configured
                            'data_values': {},
                            'score': None
                        }
                        
                        # Fetch data for each period
                        for period_code in periods:
                            try:
                                if indicator.dhis2_uid:
                                    value = self._fetch_indicator_value_for_period(
                                        indicator, org_unit_ids[0], period_code, dhis2_value_lookup, dhis2_resolved_misses
                                    )
                                    clean_value = self._clean_numeric_value(value)
                                    # For DHIS2 data, if no value is found, assign 0 to help with scoring
                                    final_value = clean_value if clean_value is not None else 0
                                    indicator_data['data_values'][period_code] = {
                                        'value': final_value,
                                        'dhis2_value': clean_value,  # Keep original for reference
                                        'manual_override': None
                                    }
                                else:
                                    # Manual indicator – initialize empty value, editable on FE
                                    indicator_data['data_values'][period_code] = {
                                        'value': None,
                                        'dhis2_value': None,
                                        'manual_override': None
                                    }
                            except Exception as e:
                                self.logger.warning(f"Failed to process {indicator.name} for period {period_code}: {str(e)}")
                                indicator_data['data_values'][period_code] = {
                                    'value': None,
                                    'dhis2_value': None,
                                    'manual_override': None
                                }
                        
                        # Apply manual entries if available for this indicator
                        if manual_entries and str(indicator.id) in manual_entries:
                            indicator_manual_entries = manual_entries[str(indicator.id)]
                            logger.info(f"Applying manual entries for indicator {indicator.id}: {indicator_manual_entries}")
                            
                            for period_code, manual_value in indicator_manual_entries.items():
                                if period_code in indicator_data['data_values']:
                                    # Apply manual override - this is the key fix
                                    indicator_data['data_values'][period_code]['value'] = manual_value
                                    indicator_data['data_values'][period_code]['manual_override'] = manual_value
                                    logger.info(f"Applied manual value {manual_value} for indicator {indicator.id} period {period_code}")
                        
                        # Compute percent_change and target_gap for latest vs previous period
                        try:
                            if isinstance(periods, list) and len(periods) >= 1:
                                last_key = periods[-1]  # This is now a period code
                                prev_key = periods[-2] if len(periods) > 1 else None
                                curr_val = indicator_data['data_values'].get(last_key, {}).get('value')
                                prev_val = indicator_data['data_values'].get(prev_key, {}).get('value') if prev_key else None
                                
                                # Calculate percent change and target gap using the same logic as old_services.py
                                change_pct = None
                                gap_pct = None
                                
                                if prev_val not in (None, 0, 0.0) and curr_val is not None:
                                    try:
                                        # For range indicators, always use standard formula regardless of target_type
                                        if indicator_data.get('target_format') == 'RANGE':
                                            change = ((float(curr_val) - float(prev_val)) / abs(float(prev_val))) * 100.0
                                        else:
                                            # For non-range indicators, use target_type specific formula
                                            if indicator.target_type == 'decrease':
                                                # For decrease indicators: (previous_value - current_value) / abs(current_value) * 100
                                                change = ((float(prev_val) - float(curr_val)) / abs(float(curr_val))) * 100.0
                                            else:
                                                # For increase indicators: (current_value - previous_value) / abs(previous_value) * 100
                                                change = ((float(curr_val) - float(prev_val)) / abs(float(prev_val))) * 100.0
                                        
                                        if change == float('inf') or change == float('-inf'):
                                            change_pct = None
                                        else:
                                            change_pct = round(change, 2)
                                    except Exception:
                                        change_pct = None
                                
                                # Calculate target gap
                                if curr_val is not None and curr_val != 0:
                                    try:
                                        tgt = indicator_data.get('target_value')
                                        target_format = indicator_data.get('target_format', 'SINGLE')
                                        target_upper = indicator_data.get('target_upper_limit')
                                        
                                        if target_format == 'RANGE' and target_upper is not None:
                                            # For range indicators: (Target upper limit - Current Value) / Current Value * 100
                                            gap_calc = ((float(target_upper) - float(curr_val)) / float(curr_val)) * 100.0
                                        else:
                                            # For non-range indicators
                                            if tgt not in (None, 0, 0.0):
                                                if indicator.target_type == 'increase':
                                                    # For increase indicators: (Current Value - Target Value) / Target Value * 100
                                                    gap_calc = ((float(curr_val) - float(tgt)) / float(tgt)) * 100.0
                                                else:
                                                    # For decrease indicators: (Target Value - Current Value) / Current Value * 100
                                                    gap_calc = ((float(tgt) - float(curr_val)) / float(curr_val)) * 100.0
                                            else:
                                                gap_calc = None
                                        
                                        if gap_calc is not None and gap_calc != float('inf') and gap_calc != float('-inf'):
                                            gap_pct = round(gap_calc, 2)
                                        else:
                                            gap_pct = None
                                    except Exception:
                                        gap_pct = None
                                
                                # Derive categories and simple threshold flags for M/N
                                change_cat = self._classify_change_category(change_pct, indicator.target_type)
                                gap_cat = self._classify_gap_category(gap_pct)
                                
                                # Use comprehensive target achievement logic from HolisticScoringService
                                current_meets = self._check_target_achievement_comprehensive(curr_val, indicator)
                                previous_meets = self._check_target_achievement_comprehensive(prev_val, indicator)
                                
                                has_data = curr_val is not None
                                trend_score = self._compute_trend_score(has_data, current_meets, previous_meets, change_cat, gap_cat, indicator, curr_val, prev_val)
                                # Derive a simple indicator score from categories/trend if not provided by DB
                                derived_score = trend_score
                                color, label = self._score_color_label(derived_score)
                                
                                indicator_data['score'] = {
                                    'percent_change': change_pct,
                                    'target_gap': gap_pct,
                                    'change_category': change_cat,
                                    'gap_category': gap_cat,
                                    'trend_score': trend_score,
                                    'score': derived_score,
                                    'score_color': color,
                                    'score_label': label,
                                    'current_value': curr_val,
                                    'previous_value': prev_val,
                                    'is_manual_override': False
                                }
                        except Exception as e:
                            self.logger.warning(f"Failed computing change/gap for indicator {indicator.id}: {e}")
                        
                        objective_data['indicators'].append(indicator_data)
                    
                    assessment_data['objectives'].append(objective_data)
            
            else:
                # Use configured indicator weights
                self.logger.info("Using configured indicator weights")
                for objective in objectives:
                    objective_data = {
                        'id': objective.id,
                        'name': objective.name,
                        'code': objective.code,
                        'color': objective.color,
                        'milestone': None,
                        'indicators': []
                    }
                    
                    # Add milestone information if it exists
                    if objective.milestone:
                        objective_data['milestone'] = {
                            'id': objective.milestone.id,
                            'name': objective.milestone.name,
                            'code': objective.milestone.code,
                            'color': objective.milestone.color,
                            'score': -2,  # Default score for real-time data
                            'score_color': '#dc3545',
                            'score_label': 'Severely Underperforming'
                        }
                    
                    # Get indicators for this objective through indicator_weights relationship
                    objective_indicators = []
                    indicator_weights_map = {}
                    for indicator_weight in objective.indicator_weights.all():
                        indicator = indicator_weight.indicator
                        # Include DHIS2-backed indicators present in list and manual indicators (no UID)
                        if (indicator.dhis2_uid in indicator_uids or not indicator.dhis2_uid) and indicator.is_active:
                            objective_indicators.append(indicator)
                            indicator_weights_map[indicator.id] = indicator_weight.weight
                    
                    self.logger.info(f"Objective {objective.name} has {len(objective_indicators)} indicators")
                    
                    for indicator in objective_indicators:
                        indicator_data = {
                            'id': indicator.id,
                            'name': indicator.name,
                            'dhis2_uid': indicator.dhis2_uid,
                            'description': indicator.description,
                            'indicator_number': indicator.indicator_number or f"{objective.order}.{len(objective_data['indicators'])+1}",
                            'display_order': indicator.display_order,
                            'target_value': float(indicator.target_value) if indicator.target_value else None,
                            'target_display': indicator.target_display,
                            'target_lower_limit': float(indicator.target_lower_limit) if indicator.target_lower_limit else None,
                            'target_upper_limit': float(indicator.target_upper_limit) if indicator.target_upper_limit else None,
                            'target_format': getattr(indicator, 'target_format', 'SINGLE'),
                            'target_type': indicator.target_type,
                            'weight': float(indicator_weights_map.get(indicator.id, 1.0)),
                            'data_values': {},
                            'score': None
                        }
                        
                        # Fetch data for each period
                        for period in periods:
                            try:
                                if indicator.dhis2_uid:
                                    value = self._fetch_indicator_value_for_period(
                                        indicator, org_unit_ids[0], period, dhis2_value_lookup, dhis2_resolved_misses
                                    )
                                    clean_value = self._clean_numeric_value(value)
                                    # For DHIS2 data, if no value is found, assign 0 to help with scoring
                                    final_value = clean_value if clean_value is not None else 0
                                    indicator_data['data_values'][period] = {
                                        'value': final_value,
                                        'dhis2_value': clean_value,  # Keep original for reference
                                        'manual_override': None
                                    }
                                else:
                                    indicator_data['data_values'][period] = {
                                        'value': None,
                                        'dhis2_value': None,
                                        'manual_override': None
                                    }
                            except Exception as e:
                                self.logger.warning(f"Failed to process {indicator.name} for period {period}: {str(e)}")
                                indicator_data['data_values'][period] = {
                                    'value': None,
                                    'dhis2_value': None,
                                    'manual_override': None
                                }
                        
                        # Apply manual entries if available for this indicator
                        if manual_entries and str(indicator.id) in manual_entries:
                            indicator_manual_entries = manual_entries[str(indicator.id)]
                            logger.info(f"Applying manual entries for indicator {indicator.id}: {indicator_manual_entries}")
                            
                            for period_code, manual_value in indicator_manual_entries.items():
                                if period_code in indicator_data['data_values']:
                                    # Apply manual override - this is the key fix
                                    indicator_data['data_values'][period_code]['value'] = manual_value
                                    indicator_data['data_values'][period_code]['manual_override'] = manual_value
                                    logger.info(f"Applied manual value {manual_value} for indicator {indicator.id} period {period_code}")
                        
                        # Compute percent_change and target_gap for latest vs previous period
                        try:
                            if isinstance(periods, list) and len(periods) >= 1:
                                last_key = periods[-1]
                                prev_key = periods[-2] if len(periods) > 1 else None
                                curr_val = indicator_data['data_values'].get(last_key, {}).get('value')
                                prev_val = indicator_data['data_values'].get(prev_key, {}).get('value') if prev_key else None
                                
                                # Calculate percent change and target gap using the same logic as old_services.py
                                change_pct = None
                                gap_pct = None
                                
                                if prev_val not in (None, 0, 0.0) and curr_val is not None:
                                    try:
                                        # For range indicators, always use standard formula regardless of target_type
                                        if indicator_data.get('target_format') == 'RANGE':
                                            change = ((float(curr_val) - float(prev_val)) / abs(float(prev_val))) * 100.0
                                        else:
                                            # For non-range indicators, use target_type specific formula
                                            if indicator.target_type == 'decrease':
                                                # For decrease indicators: (previous_value - current_value) / abs(current_value) * 100
                                                change = ((float(prev_val) - float(curr_val)) / abs(float(curr_val))) * 100.0
                                            else:
                                                # For increase indicators: (current_value - previous_value) / abs(previous_value) * 100
                                                change = ((float(curr_val) - float(prev_val)) / abs(float(prev_val))) * 100.0
                                        
                                        if change == float('inf') or change == float('-inf'):
                                            change_pct = None
                                        else:
                                            change_pct = round(change, 2)
                                    except Exception:
                                        change_pct = None
                                
                                # Calculate target gap
                                if curr_val is not None and curr_val != 0:
                                    try:
                                        tgt = indicator_data.get('target_value')
                                        target_format = indicator_data.get('target_format', 'SINGLE')
                                        target_upper = indicator_data.get('target_upper_limit')
                                        
                                        if target_format == 'RANGE' and target_upper is not None:
                                            # For range indicators: (Target upper limit - Current Value) / Current Value * 100
                                            gap_calc = ((float(target_upper) - float(curr_val)) / float(curr_val)) * 100.0
                                        else:
                                            # For non-range indicators
                                            if tgt not in (None, 0, 0.0):
                                                if indicator.target_type == 'increase':
                                                    # For increase indicators: (Current Value - Target Value) / Target Value * 100
                                                    gap_calc = ((float(curr_val) - float(tgt)) / float(tgt)) * 100.0
                                                else:
                                                    # For decrease indicators: (Target Value - Current Value) / Current Value * 100
                                                    gap_calc = ((float(tgt) - float(curr_val)) / float(curr_val)) * 100.0
                                            else:
                                                gap_calc = None
                                        
                                        if gap_calc is not None and gap_calc != float('inf') and gap_calc != float('-inf'):
                                            gap_pct = round(gap_calc, 2)
                                        else:
                                            gap_pct = None
                                    except Exception:
                                        gap_pct = None
                                
                                # Derive categories and simple threshold flags for M/N
                                change_cat = self._classify_change_category(change_pct, indicator.target_type)
                                gap_cat = self._classify_gap_category(gap_pct)
                                
                                # Use comprehensive target achievement logic from HolisticScoringService
                                current_meets = self._check_target_achievement_comprehensive(curr_val, indicator)
                                previous_meets = self._check_target_achievement_comprehensive(prev_val, indicator)
                                
                                has_data = curr_val is not None
                                trend_score = self._compute_trend_score(has_data, current_meets, previous_meets, change_cat, gap_cat, indicator, curr_val, prev_val)
                                # Derive a simple indicator score from categories/trend if not provided by DB
                                derived_score = trend_score
                                color, label = self._score_color_label(derived_score)
                                
                                indicator_data['score'] = {
                                    'percent_change': change_pct,
                                    'target_gap': gap_pct,
                                    'change_category': change_cat,
                                    'gap_category': gap_cat,
                                    'trend_score': trend_score,
                                    'score': derived_score,
                                    'score_color': color,
                                    'score_label': label,
                                    'current_value': curr_val,
                                    'previous_value': prev_val,
                                    'is_manual_override': False
                                }
                        except Exception as e:
                            self.logger.warning(f"Failed computing change/gap for indicator {indicator.id}: {e}")
                        
                        objective_data['indicators'].append(indicator_data)
                    
                    assessment_data['objectives'].append(objective_data)
            
            # Return as array to match frontend expectations  
            # Use provided org unit names if available, otherwise fetch from DHIS2
            org_unit_name = None
            if org_unit_names and len(org_unit_names) > 0:
                org_unit_name = org_unit_names[0]
                self.logger.info(f"Using provided org unit name: {org_unit_name}")
            else:
                org_unit_name = self._get_org_unit_name(org_unit_ids[0])
                self.logger.info(f"Fetched org unit name from DHIS2: {org_unit_name}")
            
            return [{
                'org_unit_id': org_unit_ids[0],
                'org_unit_name': org_unit_name,
                'assessment_period': {
                    'id': 1,
                    'name': f"{periods[0]} to {periods[-1]}" if len(periods) > 1 else periods[0],
                    'start_date': periods[0],
                    'end_date': periods[-1]
                },
                'objectives': assessment_data['objectives']
            }]
                
        except Exception as e:
            self.logger.error(f"Error fetching holistic assessment data: {str(e)}")
            raise
    
    def generate_holistic_excel(self, assessment_payload: list, manual_entries: dict = None, pre_calculated_scores: dict = None) -> str:
        """
        Generate Excel file from assessment data.
        
        Args:
            assessment_payload: Assessment data list (format from fetch_holistic_assessment_data)
            manual_entries: Dict of manual entries by indicator ID and period
            pre_calculated_scores: Dict of pre-calculated scores by indicator ID
            
        Returns:
            Path to the generated Excel file
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.formatting.rule import FormulaRule
            import os
            from django.conf import settings
        except Exception as e:
            self.logger.error(f"openpyxl not available: {e}")
            raise

        try:
            # Apply manual entries to the assessment data before generating Excel
            if manual_entries:
                self.logger.info(f"Applying manual entries to Excel export: {manual_entries}")
                for indicator_id, period_entries in manual_entries.items():
                    # Find the indicator in the assessment data
                    for period_data in assessment_payload:
                        for objective in period_data.get('objectives', []):
                            for indicator in objective.get('indicators', []):
                                if str(indicator.get('id')) == str(indicator_id):
                                    # Apply manual entries for each period
                                    for period_code, manual_value in period_entries.items():
                                        if period_code in indicator.get('data_values', {}):
                                            # Update the data_values to include manual override
                                            indicator['data_values'][period_code]['manual_override'] = manual_value
                                            # Also update the main value to ensure it's used in Excel
                                            indicator['data_values'][period_code]['value'] = manual_value
                                            self.logger.info(f"Applied manual entry {manual_value} for indicator {indicator_id} period {period_code}")
                                    break

            wb = Workbook()
            ws = wb.active
            ws.title = 'Table'

            # Styling helpers
            header_fill = PatternFill('solid', fgColor='265380')  # #265380
            header_font = Font(color='FFFFFF', bold=True)
            center = Alignment(horizontal='center', vertical='center')
            left = Alignment(horizontal='left', vertical='center')
            thin = Side(border_style='thin', color='CCCCCC')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            orange_fill = PatternFill('solid', fgColor='FDBA74')  # approx #fd7e14 light
            yellow_fill = PatternFill('solid', fgColor='FFF3BF')  # light yellow
            green50 = PatternFill('solid', fgColor='E8F5E9')  # Light green for positive change
            yellow50 = PatternFill('solid', fgColor='FFFDE7')  # Light yellow for neutral change
            red50 = PatternFill('solid', fgColor='FFEBEE')   # Light red for negative change/gap

            def score_fill(score: float | None):
                if score is None:
                    return None
                # New color scheme: 2 (Dark Green), 1 (Light Green), 0 (Yellow), -1 (Light Red), -2 (Red)
                if score >= 2:
                    return PatternFill('solid', fgColor='548235')  # Dark Green for 2
                if score >= 1:
                    return PatternFill('solid', fgColor='A9D08E')  # Light Green for 1
                if score == 0:
                    return PatternFill('solid', fgColor='FFFF00')  # Yellow for 0
                if score == -1:
                    return PatternFill('solid', fgColor='FFC7CE')  # Light Red for -1
                return PatternFill('solid', fgColor='FF0000')  # Red for -2

            data = assessment_payload[0] if assessment_payload else None
            if not data:
                raise ValueError('Empty assessment data')

            periods = []
            # Derive periods by inspecting first objective/indicator
            if data.get('objectives'):
                for obj in data['objectives']:
                    if obj.get('indicators'):
                        first = obj['indicators'][0]
                        periods = list(first.get('data_values', {}).keys())
                        break

            # Header row
            headers = ['#', 'Indicator'] + periods + ['Change', 'P-T Gap Analysis', 'Target', 'Assessed score (-2, -1, 0, +1, +2)', 'Remarks']
            ws.append(headers)
            for idx, _ in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border
            ws.row_dimensions[1].height = 22

            row = 2
            # Sort objectives by order to match frontend
            sorted_objectives = sorted(data.get('objectives', []), key=lambda x: x.get('order', 0))
            for obj in sorted_objectives:
                # Objective row
                ws.append([None] * len(headers))
                for c in range(1, len(headers) + 1):
                    cell = ws.cell(row=row, column=c)
                    cell.fill = orange_fill
                    cell.border = border
                ws.cell(row=row, column=1, value=None)
                ws.cell(row=row, column=2, value=obj.get('name')).alignment = left
                row += 1

                # Indicator rows - sort by display_order to match frontend
                sorted_indicators = sorted(obj.get('indicators', []), key=lambda x: x.get('display_order', 0))
                for ind in sorted_indicators:
                    row_values = []
                    row_values.append(ind.get('indicator_number'))
                    row_values.append(ind.get('name'))
                    # periods
                    for p in periods:
                        data_value = ind.get('data_values', {}).get(p, {})
                        # Prioritize manual override over DHIS2 value
                        v = data_value.get('manual_override') if data_value.get('manual_override') is not None else data_value.get('value')
                        row_values.append(v)
                    # change/gap - format with % symbol
                    sc = ind.get('score') or {}
                    percent_change = sc.get('percent_change')
                    target_gap = sc.get('target_gap')
                    
                    # Format percentage values with proper handling of None and zero values
                    if percent_change is not None and percent_change != 0:
                        change_str = f"{percent_change:.1f}%" if percent_change != 0 else ""
                    else:
                        change_str = ""
                    
                    if target_gap is not None and target_gap != 0:
                        gap_str = f"{target_gap:.1f}%" if target_gap != 0 else ""
                    else:
                        gap_str = ""
                    
                    row_values.append(change_str)
                    row_values.append(gap_str)
                    row_values.append(ind.get('target_display', ''))
                    
                    # Score with color
                    score_value = sc.get('score')
                    row_values.append(score_value)
                    
                    # Remarks
                    row_values.append(sc.get('remarks', ''))
                    
                    ws.append(row_values)
                    
                    # Apply styling
                    for c in range(1, len(row_values) + 1):
                        cell = ws.cell(row=row, column=c)
                        cell.border = border
                        

                        
                        # Apply score color to score column (second to last column)
                        if c == len(row_values) - 1 and score_value is not None:  # Score column
                            fill = score_fill(score_value)
                            if fill:
                                cell.fill = fill
                        
                        # Apply change colors (Change column - after periods)
                        change_col = 2 + len(periods) + 1  # After indicator name and periods
                        if c == change_col and percent_change is not None and percent_change != 0:
                            if percent_change > 5:
                                cell.fill = green50
                            elif percent_change < -5:
                                cell.fill = red50
                            else:
                                cell.fill = yellow50
                        
                        # Apply gap colors (Gap column - after change column)
                        gap_col = change_col + 1
                        if c == gap_col and target_gap is not None and target_gap != 0:
                            if target_gap > 40:
                                cell.fill = green50
                            elif target_gap <= 10:
                                cell.fill = red50
                            else:
                                cell.fill = yellow50
                    
                    row += 1

                # Add milestone row if objective has a milestone
                if obj.get('milestone'):
                    milestone = obj['milestone']
                    milestone_row = []
                    milestone_row.append('MS')  # Milestone indicator
                    milestone_row.append(milestone.get('name', 'Milestone'))
                    
                    # Add empty values for periods
                    for _ in periods:
                        milestone_row.append('-')
                    
                    # Add empty values for change and gap
                    milestone_row.append('-')
                    milestone_row.append('-')
                    milestone_row.append('-')  # Target
                    
                    # Add milestone score
                    milestone_score = milestone.get('score', 0)
                    milestone_row.append(milestone_score)
                    
                    # Add empty remarks
                    milestone_row.append('')
                    
                    ws.append(milestone_row)
                    
                    # Apply milestone styling (yellow background)
                    for c in range(1, len(milestone_row) + 1):
                        cell = ws.cell(row=row, column=c)
                        cell.fill = yellow_fill
                        cell.border = border
                        if c == 1:  # MS column
                            cell.alignment = center
                        elif c == 2:  # Milestone name
                            cell.alignment = left
                        else:
                            cell.alignment = center
                    
                    # Apply score color to milestone score
                    score_col = 2 + len(periods) + 4  # Score column position
                    if milestone_score is not None:
                        fill = score_fill(milestone_score)
                        if fill:
                            ws.cell(row=row, column=score_col).fill = fill
                            ws.cell(row=row, column=score_col).font = Font(color='FFFFFF', bold=True)
                    
                    row += 1

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save file
            import tempfile
            import uuid
            from datetime import datetime
            import re

            temp_dir = getattr(settings, 'EXCEL_EXPORT_DIR', '/tmp')
            os.makedirs(temp_dir, exist_ok=True)
            
            # Generate filename with org unit name and date, e.g.
            # "Adabraka Polyclinic_05_08_2026.xlsx" - keep the org unit name
            # human-readable (spaces intact), only strip characters that are
            # actually illegal in a filename.
            org_unit_name = data.get('org_unit_name') or 'Unknown'
            safe_org_name = re.sub(r'[\\/:*?"<>|]', '_', org_unit_name).strip()
            current_date = datetime.now().strftime('%d_%m_%Y')
            filename = f"{safe_org_name}_{current_date}.xlsx"
            file_path = os.path.join(temp_dir, filename)
            wb.save(file_path)
            
            return file_path
            
        except Exception as e:
            self.logger.error(f"Error generating holistic Excel: {str(e)}")
            raise
    
    def close(self) -> None:
        """Close the service and clean up resources."""
        try:
            # No session to close in synchronous version
            pass
        except Exception as e:
            self.logger.error(f"Error closing service: {str(e)}")
    
    def _classify_change_category(self, change_pct: float | None, target_type: str = 'increase') -> str | None:
        """Map relative percent change to O category per scoring context."""
        if change_pct is None:
            return None
        
        # Now that we use the correct formulas, we don't need to invert
        # The change_pct passed here is already the correct performance change
        performance_change = change_pct
        
        # Categorize based on performance change - EXACTLY matches flowchart
        if performance_change > 5:
            return '>5%'
        if -5 < performance_change <= 5:
            return '-5%<C<=5%'  # Stagnation category
        if -10 < performance_change <= -5:
            return '-10%<C<=-5%'
        return '<=-10%'

    def _classify_gap_category(self, gap_pct: float | None) -> str | None:
        """Map gap to P category per Excel formula: =IF($I4<=10%,"<=10%",IF(AND($I4>10%,$I4<=40%),"10%<PT<=40%",IF($I4>40%,">40%","")))"""
        if gap_pct is None:
            return None
        # Use signed gap for categorization, matching Excel behavior
        if gap_pct <= 10:
            return '<=10%'
        if 10 < gap_pct <= 40:
            return '10%<PT<=40%'
        if gap_pct > 40:
            return '>40%'
        return None

    def _compute_trend_score(self, has_data: bool, current_meets: bool | None, previous_meets: bool | None,
                              change_cat: str | None, gap_cat: str | None, indicator=None, current_value=None, previous_value=None) -> int:
        """Use the updated HolisticScoringService algorithm for real-time scoring."""
        try:
            from .scoring_service import HolisticScoringService
            
            # If we have an indicator instance, use the proper HolisticScoringService
            if indicator and hasattr(indicator, 'target_type'):
                scoring_service = HolisticScoringService()
                
                result = scoring_service.calculate_indicator_score(
                    indicator=indicator,
                    current_value=current_value,
                    previous_value=previous_value,
                    data_provided=has_data
                )
                
                return result['score']
            
            # Fallback to simplified logic if no indicator instance
            # Step 1: Data provided check
            if not has_data:
                return -2
            
            # Step 2: First year check (simplified - assume not first year if we have previous data)
            is_first_year = previous_meets is None
            
            # Step 3: Target achieved check
            target_achieved = "Yes" if current_meets else "No"
            
            # Simplified version of the Excel outcome formula
            if is_first_year:
                if target_achieved == "Yes":
                    return 1
                else:
                    return 0
            else:
                # Not first year - use the complex logic
                if target_achieved == "Yes":
                    # Target WAS achieved - check performance change
                    if change_cat == ">5%":
                        return 2
                    elif change_cat == "-5%<C<=5%":
                        # For stagnation, when target is achieved, score should be 2
                        return 2
                    elif change_cat == "-10%<C<=-5%":
                        # For negative change, score lower even if target is achieved
                        return 1
                    elif change_cat == "<=-10%":
                        return 0
                    else:
                        return 0
                else:
                    # Target NOT achieved - check performance change
                    if change_cat == ">5%":
                        return 1
                    elif change_cat == "-5%<C<=5%":
                        # Stagnation - check how close to target
                        if gap_cat == "<=10%":
                            return 1
                        elif gap_cat == "10%<PT<=40%":
                            return 0
                        elif gap_cat == ">40%":
                            return -1
                        else:
                            return 0
                    elif change_cat == "-10%<C<=-5%":
                        return -1
                    elif change_cat == "<=-10%":
                        return -1
                    else:
                        return 0
        except Exception as e:
            self.logger.error(f"Error in _compute_trend_score: {str(e)}")
            return 0

    def _median(self, numbers: list[float]) -> float | None:
        vals = [float(x) for x in numbers if x is not None and isinstance(x, (int, float))]
        if not vals:
            return None
        vals.sort()
        n = len(vals)
        mid = n // 2
        if n % 2 == 0:
            return (vals[mid-1] + vals[mid]) / 2
        return vals[mid]

    def _score_color_label(self, score: int | None) -> tuple[str, str]:
        if score is None:
            return ('#6c757d', 'N/A')
        if score >= 2:
            return ('#548235', 'Highly Performing')
        if score == 1:
            return ('#A9D08E', 'Moderately Performing')
        if score == 0:
            return ('#FFFF00', 'Sustained')
        if score == -1:
            return ('#FFC7CE', 'Underperforming')
        return ('#FF0000', 'Severely Underperforming')

    def _clean_numeric_value(self, value):
        """Clean numeric values to ensure JSON compliance"""
        import math
        
        if value is None:
            return None
        
        try:
            # Convert to float if it's a string
            if isinstance(value, str):
                value = float(value)
            
            # Check for NaN, infinity, or other invalid values
            if isinstance(value, (int, float)):
                if math.isnan(value) or math.isinf(value):
                    self.logger.warning(f"Invalid numeric value detected: {value}, setting to None")
                    return None
                return value
            
            return value
            
        except (ValueError, TypeError) as e:
            self.logger.warning(f"Error cleaning numeric value {value}: {str(e)}")
            return None

    # Short fail-fast timeout for batch analytics attempts. A struggling DHIS2
    # origin can take a very long time to give up on its own (its Cloudflare
    # edge alone allows ~100s); waiting the full default 120s per attempt, times
    # however many split levels are tried, risks exceeding gunicorn's own worker
    # timeout and dropping the connection entirely instead of degrading gracefully.
    BATCH_ATTEMPT_TIMEOUT = 25

    def _fetch_chunk_with_backoff(self, kwarg_name, uids, converted_periods, org_unit_id, allow_split=True):
        """
        Fetch one chunk of a batched DHIS2 analytics call; on failure (e.g. the
        504 Cloudflare gateway timeout seen when a chunk is too computationally
        expensive for DHIS2 to aggregate in time), split the chunk in half and
        retry both halves CONCURRENTLY, one level only (allow_split=False on the
        retry). This adapts to DHIS2's actual capacity for this specific request
        instead of a fixed guess, while keeping worst-case wall time bounded to
        roughly 2x BATCH_ATTEMPT_TIMEOUT regardless of how many indicators are in
        the chunk - fetching the two halves sequentially, or recursing further,
        would let a persistently struggling DHIS2 origin blow the total request
        time up arbitrarily (observed in testing: even small chunks kept failing
        under sustained origin load, so unbounded recursion just compounds delay
        without ever succeeding).

        Returns a dict keyed by (dx_uid, pe_code) -> value, same shape as
        _extract_values_from_analytics_response_batch. Items that still fail after
        the one retry are simply omitted - the caller (_fetch_indicator_value_for_period)
        falls back to the single-item fetch path for anything missing from this lookup.
        """
        if not uids:
            return {}
        try:
            response = self.client.get_analytics_data(
                periods=list(converted_periods),
                org_units=[org_unit_id],
                timeout=self.BATCH_ATTEMPT_TIMEOUT,
                skip_meta=True,
                **{kwarg_name: uids}
            )
            return self._extract_values_from_analytics_response_batch(response)
        except Exception as e:
            if not allow_split or len(uids) <= self.MIN_BATCH_CHUNK_SIZE:
                self.logger.warning(
                    f"Batch DHIS2 analytics fetch failed for {kwarg_name} chunk "
                    f"({len(uids)} items): {e}. Falling back to individual fetches for these."
                )
                return {}

            mid = len(uids) // 2
            self.logger.info(
                f"Batch DHIS2 analytics fetch failed for {kwarg_name} chunk "
                f"({len(uids)} items): {e}. Retrying as {mid} + {len(uids) - mid} concurrently (one retry only)."
            )
            result: Dict[tuple, float] = {}
            with ThreadPoolExecutor(max_workers=2) as executor:
                futures = [
                    executor.submit(self._fetch_chunk_with_backoff, kwarg_name, uids[:mid], converted_periods, org_unit_id, False),
                    executor.submit(self._fetch_chunk_with_backoff, kwarg_name, uids[mid:], converted_periods, org_unit_id, False),
                ]
                for future in as_completed(futures):
                    try:
                        result.update(future.result())
                    except Exception as split_error:
                        self.logger.error(f"Unexpected error in batch split retry: {split_error}")
            return result

    def _fetch_batch_indicator_data(self, indicators, org_unit_id, periods):
        """
        Batch-fetch DHIS2 analytics data for many indicators x periods in as few
        HTTP round trips as possible (grouped by indicator type, chunked to stay
        under DHIS2 URL/response-size limits), instead of one call per
        (indicator, period) pair.

        Reporting-rate indicators (dhis2_uid containing '.REPORTING_RATE') and
        dataSet-type indicators use different endpoints/logic and are deliberately
        left out of the batch - they, along with any (indicator, period) pair the
        batch doesn't return a value for, fall back to the existing
        _fetch_single_indicator_data path (with its full period-format retry logic)
        via _fetch_indicator_value_for_period.

        Returns a dict keyed by (indicator.dhis2_uid, primary_dhis2_period) -> value.
        """
        lookup: Dict[tuple, float] = {}
        if not indicators or not periods or not self.client:
            return lookup

        # Short-TTL cache: a re-export minutes later (common during a working
        # session) reuses this instead of re-hitting DHIS2. Cache key covers the
        # exact indicator set + org unit + periods requested.
        cache_key = self.cache_service.generate_cache_key(
            "dhis2_batch_indicator_data",
            org_unit_id,
            sorted(ind.dhis2_uid for ind in indicators if ind.dhis2_uid),
            sorted(periods),
        )
        cached_pairs = self.cache_service.get_assessment_cache(cache_key)
        if cached_pairs is not None:
            return {(dx_uid, pe_code): value for dx_uid, pe_code, value in cached_pairs}

        try:
            # Convert each requested period to its primary DHIS2 format once, up front.
            converted_periods = set()
            for period in periods:
                converted = self._convert_to_dhis2_period(period)
                if converted:
                    converted_periods.add(converted)

            if not converted_periods:
                return lookup

            # Group batchable indicators by type - dataSet indicators use a
            # different endpoint (get_data_set_report) and stay on the per-call path.
            groups: Dict[str, List[str]] = {'indicator': [], 'dataElement': [], 'programIndicator': []}
            for indicator in indicators:
                uid = indicator.dhis2_uid
                if not uid or '.REPORTING_RATE' in uid:
                    continue
                indicator_type = getattr(indicator, 'indicator_type', 'indicator') or 'indicator'
                if indicator_type in groups:
                    groups[indicator_type].append(uid)

            fetch_kwarg_name = {
                'indicator': 'indicators',
                'dataElement': 'data_elements',
                'programIndicator': 'program_indicators',
            }
            # Chunk size is bounded by total "cells" (indicators x periods) in one
            # DHIS2 analytics call, not just indicator count - a request asking for
            # many indicators across several periods is much more expensive for
            # DHIS2 to compute than the same indicator count for one period. A
            # 35-indicator x 3-period call was observed timing out at DHIS2's own
            # Cloudflare edge (504, ~100s) even though it's well under any URL
            # length limit - the origin server itself couldn't compute it in time.
            num_periods = max(1, len(converted_periods))
            batch_chunk_size = max(self.MIN_BATCH_CHUNK_SIZE, self.MAX_BATCH_CELLS // num_periods)

            # Flatten every type-group's chunks into one job list and fetch them
            # ALL concurrently, rather than looping group-by-group (sequential
            # groups would each pay up to ~2x BATCH_ATTEMPT_TIMEOUT, and with more
            # than one active indicator_type that stacks up past gunicorn's worker
            # timeout - most projects only use one type, but nothing should assume
            # that).
            jobs = []
            for indicator_type, uids in groups.items():
                if not uids:
                    continue
                kwarg_name = fetch_kwarg_name[indicator_type]
                chunks = [uids[i:i + batch_chunk_size] for i in range(0, len(uids), batch_chunk_size)]
                jobs.extend((kwarg_name, chunk) for chunk in chunks)

            if jobs:
                # Fetch every chunk (across all type-groups) concurrently - each is
                # a separate DHIS2 call, and DHIS2Client's underlying requests.Session
                # is thread-safe for concurrent reads. Capped at 3 concurrent
                # top-level chunks (each of which may itself spawn 2 more on a split
                # retry, so worst case ~6 concurrent requests) to avoid piling
                # additional concurrent load onto a DHIS2 origin that's already
                # struggling, which would make things worse, not better.
                with ThreadPoolExecutor(max_workers=min(3, len(jobs))) as executor:
                    futures = [
                        executor.submit(self._fetch_chunk_with_backoff, kwarg_name, chunk, converted_periods, org_unit_id)
                        for kwarg_name, chunk in jobs
                    ]
                    # _fetch_chunk_with_backoff already catches everything it can
                    # anticipate (including recursively) - this is a pure safety
                    # net so one truly unexpected failure doesn't abandon the
                    # results of every other already-completed chunk.
                    for future in as_completed(futures):
                        try:
                            lookup.update(future.result())
                        except Exception as chunk_error:
                            self.logger.error(f"Unexpected error resolving a batch chunk future: {chunk_error}")

        except Exception as e:
            self.logger.error(f"Error batch-fetching DHIS2 indicator data: {str(e)}")

        # Tuple keys aren't JSON-serializable, so cache as a flat list of triples.
        self.cache_service.set_assessment_cache(
            cache_key,
            [[dx_uid, pe_code, value] for (dx_uid, pe_code), value in lookup.items()],
            timeout=self.DHIS2_BATCH_CACHE_TIMEOUT,
        )

        return lookup

    # Wall-clock ceiling for resolving every batch-miss concurrently. Chosen so
    # batch phase (~2x BATCH_ATTEMPT_TIMEOUT worst case) + this phase stays
    # comfortably under gunicorn's 120s worker timeout, leaving headroom for the
    # (now I/O-free) assembly loops and file generation that follow.
    MISS_PREFETCH_TIME_BUDGET = 45
    MISS_PREFETCH_MAX_WORKERS = 5

    def _prefetch_remaining_indicator_values(self, indicators, org_unit_id, periods, batch_lookup):
        """
        Concurrently resolve every (indicator, period) pair the batch step didn't
        cover - reporting-rate indicators, dataSet-type indicators (different
        DHIS2 endpoint, never batched), and genuine batch misses - instead of
        leaving them to the assembly loops below, which would otherwise fetch
        each one synchronously, one at a time. With dozens of indicators across
        several periods, that serial fallback is what turns a struggling DHIS2
        origin into a multi-minute request (each miss can cost up to
        BATCH_ATTEMPT_TIMEOUT seconds, and there can be 100+ misses).

        Bounded two ways: a small worker cap (matching the batch phase's own
        concurrency limit, so this doesn't pile extra load onto a DHIS2 origin
        that may already be struggling) and a hard wall-clock budget. Whatever
        isn't resolved by the deadline is recorded as None rather than left
        unattempted - the assembly loop trusts that (see
        _fetch_indicator_value_for_period) instead of retrying it synchronously,
        which would just reintroduce the same unbounded wait this exists to avoid.
        Threads still in flight when the deadline passes are abandoned (not
        cancelled - Python can't interrupt a blocking socket read) and simply
        discarded once they finish; harmless for a long-running worker process.

        Returns a dict keyed by (dhis2_uid, converted_period) -> value_or_None,
        covering every pair that was attempted.
        """
        resolved: Dict[tuple, Any] = {}
        if not indicators or not periods:
            return resolved

        pending = []
        seen_keys = set()
        for indicator in indicators:
            if not indicator.dhis2_uid:
                continue
            for period in periods:
                converted = self._convert_to_dhis2_period(period)
                if not converted:
                    continue
                key = (indicator.dhis2_uid, converted)
                if key in batch_lookup or key in seen_keys:
                    continue
                seen_keys.add(key)
                pending.append((indicator, period, key))

        if not pending:
            return resolved

        executor = ThreadPoolExecutor(max_workers=self.MISS_PREFETCH_MAX_WORKERS)
        try:
            future_map = {
                executor.submit(self._fetch_single_indicator_data, indicator, org_unit_id, period): key
                for indicator, period, key in pending
            }
            done, not_done = wait(future_map.keys(), timeout=self.MISS_PREFETCH_TIME_BUDGET)

            for future in done:
                key = future_map[future]
                try:
                    resolved[key] = future.result()
                except Exception as e:
                    self.logger.warning(f"Prefetch failed for {key}: {e}")
                    resolved[key] = None

            if not_done:
                self.logger.warning(
                    f"Miss-prefetch time budget ({self.MISS_PREFETCH_TIME_BUDGET}s) exhausted "
                    f"with {len(not_done)}/{len(pending)} pairs still in flight - treating as "
                    f"unavailable rather than waiting further."
                )
                for future in not_done:
                    resolved[future_map[future]] = None
        finally:
            # wait=False: don't block this request on threads whose results we've
            # already given up on; they'll finish (and be discarded) on their own.
            executor.shutdown(wait=False)

        return resolved

    def _extract_values_from_analytics_response_batch(self, response):
        """
        Parse a DHIS2 /api/analytics response into a {(dx_uid, pe_code): value}
        lookup covering every row, by reading the dx/pe/value column positions
        from the response headers. Unlike _extract_value_from_analytics_response
        (which only ever reads row 0 - correct for a single-indicator/single-period
        request), this is built for multi-indicator, multi-period batched responses.
        """
        lookup = {}
        try:
            if not response or not isinstance(response, dict):
                return lookup

            headers = response.get('headers', [])
            rows = response.get('rows', [])
            if not headers or not rows:
                return lookup

            dx_index = pe_index = value_index = None
            for i, header in enumerate(headers):
                name = (header.get('name') or '').lower()
                if name == 'dx':
                    dx_index = i
                elif name == 'pe':
                    pe_index = i
                elif name == 'value':
                    value_index = i

            if dx_index is None or pe_index is None:
                return lookup
            if value_index is None:
                value_index = len(headers) - 1  # DHIS2 analytics always puts value last

            for row in rows:
                if len(row) <= max(dx_index, pe_index, value_index):
                    continue
                raw_value = row[value_index]
                if raw_value is None or raw_value == '':
                    continue
                try:
                    value = float(raw_value)
                except (ValueError, TypeError):
                    continue
                lookup[(row[dx_index], row[pe_index])] = value

        except Exception as e:
            self.logger.error(f"Error parsing batched analytics response: {str(e)}")

        return lookup

    def _fetch_indicator_value_for_period(self, indicator, org_unit_id, period, batch_lookup, resolved_misses=None):
        """
        Look up a pre-fetched batched value first, then a concurrently-resolved
        miss (see _prefetch_remaining_indicator_values) before falling back to a
        live single-indicator/single-period fetch - which, unlike the other two
        paths, blocks this request on network I/O and should be a rare last
        resort, not the common case.
        """
        converted_period = self._convert_to_dhis2_period(period) if (batch_lookup or resolved_misses) else None

        if batch_lookup and converted_period:
            cached = batch_lookup.get((indicator.dhis2_uid, converted_period))
            if cached is not None:
                return cached

        if resolved_misses is not None and converted_period:
            key = (indicator.dhis2_uid, converted_period)
            if key in resolved_misses:
                # Already attempted concurrently, possibly with no result (None) -
                # trust it rather than paying for a redundant synchronous retry.
                return resolved_misses[key]

        return self._fetch_single_indicator_data(indicator, org_unit_id, period)

    def _fetch_single_indicator_data(self, indicator, org_unit_id, period):
        """
        Fetch data for a single indicator without storing in database
        Uses the same approach as the working old_services.py
        """
        self.logger.debug(f"Fetching real-time data for {indicator.name} ({indicator.dhis2_uid}) - {org_unit_id} - {period}")
        
        try:
            # Handle reporting rate indicators differently
            if '.REPORTING_RATE' in indicator.dhis2_uid:
                # Convert period to DHIS2 format for reporting rates
                dhis2_period = self._convert_to_dhis2_period(period)
                if not dhis2_period:
                    self.logger.warning(f"Could not convert period {period} to DHIS2 format for reporting rate")
                    return None
                
                self.logger.info(f"Fetching reporting rate for {indicator.dhis2_uid} with period {dhis2_period}")
                
                # For reporting rates, use the full UID including .REPORTING_RATE as a data element
                response = self.client.get_analytics_data(
                    data_elements=[indicator.dhis2_uid],
                    periods=[dhis2_period],
                    org_units=[org_unit_id],
                    skip_meta=True
                )
                value = self._extract_value_from_analytics_response(response, indicator.dhis2_uid)
                if value is not None:
                    self.logger.info(f"Successfully fetched reporting rate: {value}")
                    return value
                    
                self.logger.info(f"No reporting rate data found for {indicator.dhis2_uid}")
                return None

            # Convert period to DHIS2 format
            dhis2_period = self._convert_to_dhis2_period(period)
            self.logger.debug(f"Using DHIS2 period format: {dhis2_period}")
            
            # Try different period formats if needed
            periods_to_try = [dhis2_period]
            if '-' in period:  # If it's a date format
                try:
                    from datetime import datetime
                    date_obj = datetime.strptime(period, '%Y-%m-%d')
                    # Add alternative period formats
                    year = date_obj.strftime('%Y')
                    month = date_obj.strftime('%m')
                    quarter = ((int(month) - 1) // 3) + 1
                    semester = 1 if int(month) <= 6 else 2
                    periods_to_try.extend([
                        f"{year}{month}",  # YYYYMM
                        f"{year}Q{quarter}",  # YYYYQ#
                        f"{year}S{semester}",  # YYYY[S]#
                        year  # YYYY
                    ])
                except ValueError:
                    pass

            # Try each period format
            for try_period in periods_to_try:
                try:
                    # Make DHIS2 API request based on indicator type
                    indicator_type = getattr(indicator, 'indicator_type', 'indicator')  # Default to 'indicator' if not set
                    
                    self.logger.info(f"Making API request for {indicator.name} ({indicator.dhis2_uid}) with type '{indicator_type}' and period '{try_period}'")
                    
                    if indicator_type == 'indicator':
                        response = self.client.get_analytics_data(
                            indicators=[indicator.dhis2_uid],
                            periods=[try_period],
                            org_units=[org_unit_id],
                            timeout=self.BATCH_ATTEMPT_TIMEOUT,
                            skip_meta=True
                        )
                    elif indicator_type == 'dataElement':
                        response = self.client.get_analytics_data(
                            data_elements=[indicator.dhis2_uid],
                            periods=[try_period],
                            org_units=[org_unit_id],
                            timeout=self.BATCH_ATTEMPT_TIMEOUT,
                            skip_meta=True
                        )
                    elif indicator_type == 'dataSet':
                        response = self.client.get_data_set_report(
                            data_set_id=indicator.dhis2_uid,
                            periods=[try_period],
                            org_units=[org_unit_id],
                            timeout=self.BATCH_ATTEMPT_TIMEOUT
                        )
                    elif indicator_type == 'programIndicator':
                        response = self.client.get_analytics_data(
                            program_indicators=[indicator.dhis2_uid],
                            periods=[try_period],
                            org_units=[org_unit_id],
                            timeout=self.BATCH_ATTEMPT_TIMEOUT,
                            skip_meta=True
                        )
                    else:
                        # Fallback - try as indicator first, then data element
                        self.logger.info(f"Unknown indicator type '{indicator_type}', trying as indicator first")
                        try:
                            response = self.client.get_analytics_data(
                                indicators=[indicator.dhis2_uid],
                                periods=[try_period],
                                org_units=[org_unit_id],
                                timeout=self.BATCH_ATTEMPT_TIMEOUT,
                                skip_meta=True
                            )
                        except requests.exceptions.RequestException as e:
                            self.logger.info(f"Failed as indicator, trying as data element: {str(e)}")
                            response = self.client.get_analytics_data(
                                data_elements=[indicator.dhis2_uid],
                                periods=[try_period],
                                org_units=[org_unit_id],
                                timeout=self.BATCH_ATTEMPT_TIMEOUT,
                                skip_meta=True
                            )

                    # Extract value from response
                    if indicator_type == 'dataSet':
                        value = self._extract_value_from_dataset_response(response, indicator.dhis2_uid)
                    else:
                        value = self._extract_value_from_analytics_response(response, indicator.dhis2_uid)

                    if value is not None:
                        self.logger.info(f"Successfully fetched data for {indicator.name} using period {try_period}: {value}")
                        return value
                except requests.exceptions.RequestException as e:
                    # A request-level failure (timeout, connection error, 5xx) means DHIS2
                    # itself is unreachable/unhealthy right now - retrying with a different
                    # period *string* can't fix that, and doing so just multiplies the wait
                    # (each retry pays the same timeout again). Give up immediately instead
                    # of trying every remaining format or falling through to the alternative-
                    # format cascade below.
                    self.logger.warning(f"DHIS2 request failed for {indicator.name} (period {try_period}): {str(e)}")
                    return None
                except Exception as e:
                    self.logger.debug(f"Error fetching data for period {try_period}: {str(e)}")
                    continue

            # If no data found with any period format, try alternative period formats
            self.logger.info(f"No data found for {indicator.name} using any period format, trying alternative formats")
            return self._try_alternative_period_formats(indicator, org_unit_id, period)
                
        except Exception as e:
            self.logger.error(f"Error fetching data for {indicator.name}: {str(e)}")
            return None

    def _convert_to_dhis2_period(self, period):
        """
        Convert period to DHIS2 format
        
        Supported formats:
        - Yearly: 2024 -> 2024
        - Six-monthly: 2024S1, 2024S2
        - Quarterly: 2024Q1, 2024Q2, 2024Q3, 2024Q4
        - Monthly: 202401, 202402, etc.
        - Weekly: 2024W1, 2024W2, etc.
        - Date strings: 2024-01-01 -> 202401
        """
        try:
            if not period or not isinstance(period, (str, dict)):
                self.logger.warning(f"Invalid period format: {period}")
                return None

            # Handle period dict format
            if isinstance(period, dict):
                if 'code' in period:
                    period = period['code']
                else:
                    self.logger.warning(f"Invalid period dict format: {period}")
                    return None

            # Handle relative periods
            relative_periods = {
                'THIS_YEAR': lambda: datetime.now().strftime('%Y'),
                'LAST_YEAR': lambda: str(int(datetime.now().strftime('%Y')) - 1),
                'THIS_QUARTER': lambda: self._get_current_quarter(),
                'LAST_QUARTER': lambda: self._get_last_quarter(),
                'THIS_MONTH': lambda: datetime.now().strftime('%Y%m'),
                'LAST_MONTH': lambda: (datetime.now().replace(day=1) - timedelta(days=1)).strftime('%Y%m'),
            }

            if period in relative_periods:
                return relative_periods[period]()

            # Handle human-readable period formats like "January 2023"
            import re
            month_names = [
                'January', 'February', 'March', 'April', 'May', 'June',
                'July', 'August', 'September', 'October', 'November', 'December'
            ]
            
            # Pattern for "Month Year" format
            month_year_pattern = r'^(' + '|'.join(month_names) + r')\s+(\d{4})$'
            month_match = re.match(month_year_pattern, period, re.IGNORECASE)
            if month_match:
                try:
                    month_name = month_match.group(1)
                    year = month_match.group(2)
                    month_num = month_names.index(month_name.title()) + 1
                    # Convert to quarterly format
                    quarter = ((month_num - 1) // 3) + 1
                    dhis2_period = f"{year}Q{quarter}"
                    self.logger.info(f"Converted '{period}' to quarterly period {dhis2_period}")
                    return dhis2_period
                except (ValueError, IndexError):
                    self.logger.warning(f"Could not parse month from: {period}")
                    return None
            
            # Pattern for "Month, Year" format
            month_year_comma_pattern = r'^(' + '|'.join(month_names) + r'),\s*(\d{4})$'
            month_comma_match = re.match(month_year_comma_pattern, period, re.IGNORECASE)
            if month_comma_match:
                try:
                    month_name = month_comma_match.group(1)
                    year = month_comma_match.group(2)
                    month_num = month_names.index(month_name.title()) + 1
                    # Convert to quarterly format
                    quarter = ((month_num - 1) // 3) + 1
                    dhis2_period = f"{year}Q{quarter}"
                    self.logger.info(f"Converted '{period}' to quarterly period {dhis2_period}")
                    return dhis2_period
                except (ValueError, IndexError):
                    self.logger.warning(f"Could not parse month from: {period}")
                    return None

            # Handle date string format (YYYY-MM-DD)
            if re.match(r'^\d{4}-\d{2}-\d{2}$', period):
                try:
                    date_obj = datetime.strptime(period, '%Y-%m-%d')
                    # For date strings, determine the appropriate period format based on the date
                    # For now, convert to quarterly format since that's commonly used
                    year = date_obj.year
                    month = date_obj.month
                    quarter = ((month - 1) // 3) + 1
                    dhis2_period = f"{year}Q{quarter}"
                    self.logger.info(f"Converted date {period} to quarterly period {dhis2_period}")
                    return dhis2_period
                except ValueError:
                    self.logger.warning(f"Invalid date string format: {period}")
                    return None

            # Handle fixed period formats
            if re.match(r'^\d{4}$', period):  # Yearly: 2024
                return period
            elif re.match(r'^\d{4}S[1-2]$', period):  # Six-monthly: 2024S1
                return period
            elif re.match(r'^\d{4}Q[1-4]$', period):  # Quarterly: 2024Q1
                return period
            elif re.match(r'^\d{6}$', period):  # Monthly: 202401
                return period
            elif re.match(r'^\d{4}W[1-53]$', period):  # Weekly: 2024W1
                return period
            # Handle space-separated formats
            elif re.match(r'^\d{4}\s+S[1-2]$', period):  # Six-monthly: 2024 S1
                return period.replace(' ', '')  # Remove space
            elif re.match(r'^\d{4}\s+Q[1-4]$', period):  # Quarterly: 2024 Q1
                return period.replace(' ', '')  # Remove space
            elif re.match(r'^\d{4}\s+W[1-53]$', period):  # Weekly: 2024 W1
                return period.replace(' ', '')  # Remove space
            else:
                # Try to parse as date if it contains dashes
                if '-' in period:
                    try:
                        date_obj = datetime.strptime(period.split('T')[0], '%Y-%m-%d')
                        return date_obj.strftime('%Y%m')  # Convert to YYYYMM format
                    except ValueError:
                        pass
                self.logger.warning(f"Unrecognized period format: {period}")
                return None

        except Exception as e:
            self.logger.error(f"Error converting period {period}: {str(e)}")
            return None

    def _get_current_quarter(self):
        now = datetime.now()
        year = now.strftime('%Y')
        quarter = ((now.month - 1) // 3) + 1
        return f"{year}Q{quarter}"

    def _get_last_quarter(self):
        now = datetime.now()
        year = now.strftime('%Y')
        quarter = ((now.month - 1) // 3)
        
        if quarter == 0:
            year = str(int(year) - 1)
            quarter = 4
        
        return f"{year}Q{quarter}"

    def _try_alternative_period_formats(self, indicator, org_unit_id, period):
        """Try alternative period formats when no data is found"""
        try:
            # Extract year and period type
            year = period[:4]
            period_type = None
            
            if '-' in period:  # Handle date format like 2022-10-01
                try:
                    from datetime import datetime
                    date_obj = datetime.strptime(period, '%Y-%m-%d')
                    year = date_obj.strftime('%Y')
                    month = date_obj.strftime('%m')
                    period_type = 'monthly'
                    period = f"{year}{month}"  # Convert to YYYYMM format
                except ValueError:
                    self.logger.warning(f"Invalid date format: {period}")
                    return None
            elif 'Q' in period:
                period_type = 'quarterly'
            elif 'S' in period:
                period_type = 'sixmonthly'
            
            alternative_periods = []
            
            if period_type == 'monthly':
                # Try quarterly period for the corresponding month
                quarter = ((int(period[4:6]) - 1) // 3) + 1
                alternative_periods.append(f"{year}Q{quarter}")
                
                # Try six-monthly period
                semester = 1 if int(period[4:6]) <= 6 else 2
                alternative_periods.append(f"{year}S{semester}")
                
            elif period_type == 'quarterly':
                # Try monthly periods for the quarter
                # Extract quarter number properly (e.g., "2023Q1" -> quarter = 1)
                quarter_str = period.split('Q')[1] if 'Q' in period else period[5:]
                try:
                    quarter = int(quarter_str)
                    start_month = (quarter - 1) * 3 + 1
                    for month in range(start_month, start_month + 3):
                        alternative_periods.append(f"{year}{month:02d}")
                except (ValueError, IndexError):
                    self.logger.warning(f"Could not parse quarter from period: {period}")
                    pass
                    
            elif period_type == 'sixmonthly':
                # Try quarterly periods for the semester
                # Extract semester number properly (e.g., "2023S1" -> semester = 1)
                semester_str = period.split('S')[1] if 'S' in period else period[5:]
                try:
                    semester = int(semester_str)
                    start_quarter = (semester - 1) * 2 + 1
                    for quarter in range(start_quarter, start_quarter + 2):
                        alternative_periods.append(f"{year}Q{quarter}")
                except (ValueError, IndexError):
                    self.logger.warning(f"Could not parse semester from period: {period}")
                    pass
            
            # Always try yearly as fallback
            alternative_periods.append(year)
            
            self.logger.info(f"Trying alternative period formats for {indicator.name}: {alternative_periods}")
            
            for alt_period in alternative_periods:
                try:
                    # Make DHIS2 API request based on indicator type
                    if indicator.indicator_type == 'indicator':
                        response = self.client.get_analytics_data(
                            indicators=[indicator.dhis2_uid],
                            periods=[alt_period],
                            org_units=[org_unit_id],
                            timeout=self.BATCH_ATTEMPT_TIMEOUT,
                            skip_meta=True
                        )
                    elif indicator.indicator_type == 'dataElement':
                        response = self.client.get_analytics_data(
                            data_elements=[indicator.dhis2_uid],
                            periods=[alt_period],
                            org_units=[org_unit_id],
                            timeout=self.BATCH_ATTEMPT_TIMEOUT,
                            skip_meta=True
                        )
                    elif indicator.indicator_type == 'dataSet':
                        response = self.client.get_data_set_report(
                            data_set_id=indicator.dhis2_uid,
                            periods=[alt_period],
                            org_units=[org_unit_id],
                            timeout=self.BATCH_ATTEMPT_TIMEOUT
                        )
                    else:
                        continue
                    
                    # Extract value from response
                    if indicator.indicator_type == 'dataSet':
                        value = self._extract_value_from_dataset_response(response, indicator.dhis2_uid)
                    else:
                        value = self._extract_value_from_analytics_response(response, indicator.dhis2_uid)
                    
                    if value is not None:
                        self.logger.info(f"Found data using alternative period format: {alt_period}")
                        return value

                except requests.exceptions.RequestException as e:
                    # DHIS2 itself failed to respond - other alt period strings will hit the
                    # same unresponsive server, so stop here instead of paying the timeout
                    # again for every remaining candidate.
                    self.logger.warning(f"DHIS2 request failed for alternative period {alt_period}: {str(e)}")
                    return None
                except Exception as e:
                    self.logger.debug(f"Alternative period {alt_period} failed: {str(e)}")
                    continue
            
            self.logger.debug(f"All alternative period formats failed for {indicator.name}")
            return None
            
        except Exception as e:
            self.logger.error(f"Error trying alternative period formats for {indicator.name}: {str(e)}")
            return None

    def _extract_value_from_analytics_response(self, response, indicator_uid):
        """Extract value from DHIS2 analytics response - same logic as DataSyncService"""
        try:
            if not response or not isinstance(response, dict):
                self.logger.warning(f"Invalid response format for indicator {indicator_uid}")
                return None
            
            # Check for rows in response
            rows = response.get('rows', [])
            if not rows:
                # This is normal - some indicators don't have data for all periods/org units
                self.logger.info(f"No data available for indicator {indicator_uid} - this is normal if the indicator has no data for the specified period/org unit")
                return None
            
            # Get headers to understand the structure
            headers = response.get('headers', [])
            if not headers:
                self.logger.warning(f"No headers found in response for indicator {indicator_uid}")
                return None
            
            # Enhanced column detection
            # Look for the indicator UID in the headers
            indicator_column_index = None
            value_column_index = None
            
            for i, header in enumerate(headers):
                header_name = header.get('name', '').lower()
                header_column = header.get('column', '').lower()
                
                # Check if this header contains our indicator UID
                if indicator_uid.lower() in header_name or indicator_uid.lower() in header_column:
                    indicator_column_index = i
                    break
            
            # If we found the indicator column, the value should be in the next column
            if indicator_column_index is not None:
                value_column_index = indicator_column_index + 1
            else:
                # Fallback: look for value columns
                for i, header in enumerate(headers):
                    header_name = header.get('name', '').lower()
                    if 'value' in header_name or 'data' in header_name:
                        value_column_index = i
                        break
            
            # If still no value column found, use the last column
            if value_column_index is None and len(headers) > 1:
                value_column_index = len(headers) - 1
            
            self.logger.debug(f"Using value column index {value_column_index} for indicator {indicator_uid}")
            
            # Extract value from the first row
            if value_column_index is not None and len(rows) > 0:
                first_row = rows[0]
                if len(first_row) > value_column_index:
                    value = first_row[value_column_index]
                    self.logger.debug(f"Extracted value {value} from row {first_row}")
                    
                    # Convert to float if possible
                    try:
                        if isinstance(value, str):
                            value = float(value)
                        return value
                    except (ValueError, TypeError):
                        self.logger.warning(f"Could not convert value '{value}' to float for indicator {indicator_uid}")
                        return None
            
            # Try alternative parsing if standard parsing fails
            self.logger.info(f"Standard parsing failed, trying alternative parsing for indicator {indicator_uid}")
            return self._extract_value_alternative_parsing(response, indicator_uid, value_column_index)
            
        except Exception as e:
            self.logger.error(f"Error extracting value from analytics response for indicator {indicator_uid}: {str(e)}")
            return None

    def _extract_value_alternative_parsing(self, response, indicator_uid, value_column_index):
        """Alternative parsing method for analytics response - same logic as DataSyncService"""
        try:
            self.logger.info(f"Starting alternative parsing for {indicator_uid}")
            
            # Try to find the indicator in metadata
            meta_data = response.get('metaData', {})
            items = meta_data.get('items', {})
            
            # Look for the indicator in the items
            if indicator_uid in items:
                item_info = items[indicator_uid]
                self.logger.info(f"Found indicator info in metadata: {item_info}")
            
            # Process rows with more flexible matching
            rows = response.get('rows', [])
            self.logger.info(f"Alternative parsing: processing {len(rows)} rows")
            
            for i, row in enumerate(rows):
                if value_column_index is None or len(row) <= value_column_index:
                    self.logger.debug(f"Alternative parsing: skipping row {i} with insufficient columns")
                    continue
                
                # Try to match by checking if the indicator UID appears anywhere in the row
                row_str = ' '.join(str(cell) for cell in row)
                if indicator_uid in row_str:
                    self.logger.info(f"Alternative parsing: found indicator {indicator_uid} in row {i}: {row}")
                    raw_value = row[value_column_index]
                    
                    if raw_value is None or raw_value == '':
                        self.logger.warning(f"Alternative parsing: empty value found for {indicator_uid}")
                        return None
                    
                    try:
                        value = float(raw_value)
                        self.logger.info(f"Alternative parsing: successfully extracted value {value} for {indicator_uid}")
                        return value
                    except (ValueError, TypeError):
                        self.logger.warning(f"Alternative parsing: could not convert value '{raw_value}' to float for {indicator_uid}")
                        continue
            
            self.logger.warning(f"Alternative parsing: no value found for {indicator_uid}")
            return None
            
        except Exception as e:
            self.logger.error(f"Error in alternative parsing for indicator {indicator_uid}: {str(e)}")
            return None

    def _extract_value_from_dataset_response(self, response, indicator_uid):
        """Extract value from DHIS2 dataset response - same logic as DataSyncService"""
        try:
            if not response or not isinstance(response, dict):
                self.logger.warning(f"Invalid dataset response format for indicator {indicator_uid}")
                return None
            
            # Dataset responses have a different structure
            # Look for the indicator in the response
            if indicator_uid in response:
                value = response[indicator_uid]
                if value is None:
                    self.logger.debug(f"Dataset value is None for indicator {indicator_uid}")
                    return None
                try:
                    return float(value)
                except (ValueError, TypeError):
                    self.logger.warning(f"Could not convert dataset value '{value}' to float for indicator {indicator_uid}")
                    return None
            
            self.logger.warning(f"Indicator {indicator_uid} not found in dataset response")
            return None
            
        except Exception as e:
            self.logger.error(f"Error extracting value from dataset response for indicator {indicator_uid}: {str(e)}")
            return None

    def _get_org_unit_name(self, org_unit_id):
        """Get organization unit name from cache or DHIS2 API"""
        try:
            # Try to get from cache first
            cache_key = f"org_unit_name_{org_unit_id}"
            org_unit_name = cache.get(cache_key)
            
            if org_unit_name:
                return org_unit_name
            
            # For real-time service, try to fetch from DHIS2 API if client is available
            try:
                if self.client:
                    org_unit_data = self.client._make_request("GET", f"api/organisationUnits/{org_unit_id}", 
                                                             params={"fields": "id,name,displayName"})
                    org_unit_name = org_unit_data.get('displayName') or org_unit_data.get('name') or f"Org Unit {org_unit_id}"
                else:
                    org_unit_name = f"Org Unit {org_unit_id}"
            except Exception as e:
                self.logger.warning(f"Could not fetch org unit name from DHIS2: {str(e)}")
                org_unit_name = f"Org Unit {org_unit_id}"
            
            # Cache the result
            cache.set(cache_key, org_unit_name, timeout=3600)  # Cache for 1 hour
            
            return org_unit_name
            
        except Exception as e:
            self.logger.warning(f"Error getting org unit name for {org_unit_id}: {str(e)}")
            return f"Org Unit {org_unit_id}"
    
    def _check_target_achievement_comprehensive(self, current_value: Optional[float], indicator) -> bool:
        """
        Check if target was achieved using comprehensive logic from HolisticScoringService.
        
        Args:
            current_value: Current value
            indicator: Indicator instance
            
        Returns:
            True if target was achieved, False otherwise
        """
        if current_value is None:
            return False
        
        try:
            current_val = float(current_value)
            
            # Handle different target formats
            if hasattr(indicator, 'target_format') and indicator.target_format == 'RANGE':
                # Range target: check if current value is within the range
                if indicator.target_lower_limit is not None and indicator.target_upper_limit is not None:
                    lower_limit = float(indicator.target_lower_limit)
                    upper_limit = float(indicator.target_upper_limit)
                    return lower_limit <= current_val <= upper_limit
                else:
                    # Fallback to single target value
                    if indicator.target_value is not None:
                        target_float = float(indicator.target_value)
                        if indicator.target_type == 'decrease':
                            return current_val <= target_float
                        else:
                            return current_val >= target_float
            elif hasattr(indicator, 'target_format') and indicator.target_format == 'MINIMUM':
                # Minimum target: current value should be >= target_value
                if indicator.target_value is not None:
                    target_float = float(indicator.target_value)
                    return current_val >= target_float
            elif hasattr(indicator, 'target_format') and indicator.target_format == 'MAXIMUM':
                # Maximum target: current value should be <= target_value
                if indicator.target_value is not None:
                    target_float = float(indicator.target_value)
                    return current_val <= target_float
            else:
                # Single value target: use the target_operator
                if indicator.target_value is not None:
                    target_float = float(indicator.target_value)
                    
                    # Use the target_operator to determine achievement
                    if indicator.target_operator == '>=':
                        return current_val >= target_float
                    elif indicator.target_operator == '>':
                        return current_val > target_float
                    elif indicator.target_operator == '<=':
                        return current_val <= target_float
                    elif indicator.target_operator == '<':
                        return current_val < target_float
                    elif indicator.target_operator == '=':
                        return current_val == target_float
                    else:
                        # Default to >= for backward compatibility
                        return current_val >= target_float
            
            return False
            
        except Exception:
            return False
